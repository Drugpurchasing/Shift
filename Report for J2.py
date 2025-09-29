import streamlit as st
import pandas as pd
import io
import numpy as np
from PyPDF2 import PdfMerger
from openpyxl.styles import Alignment, Font
from PIL import Image

# ==============================================================================
# Page Configuration (‡∏ï‡πâ‡∏≠‡∏á‡πÄ‡∏õ‡πá‡∏ô‡∏Ñ‡∏≥‡∏™‡∏±‡πà‡∏á‡πÅ‡∏£‡∏Å)
# ==============================================================================
st.set_page_config(
    page_title="CRA Analytics Suite",
    page_icon="üî¨",
    layout="wide"
)


# ==============================================================================
# Functions 1-7 (‡πÇ‡∏Ñ‡πâ‡∏î‡∏ü‡∏±‡∏á‡∏Å‡πå‡∏ä‡∏±‡∏ô‡πÄ‡∏î‡∏¥‡∏°‡∏ó‡∏±‡πâ‡∏á‡∏´‡∏°‡∏î ‡πÑ‡∏°‡πà‡∏°‡∏µ‡∏Å‡∏≤‡∏£‡πÄ‡∏õ‡∏•‡∏µ‡πà‡∏¢‡∏ô‡πÅ‡∏õ‡∏•‡∏á Logic)
# ==============================================================================
def process_j2_report(uploaded_files):
    # (‡πÇ‡∏Ñ‡πâ‡∏î‡∏ü‡∏±‡∏á‡∏Å‡πå‡∏ä‡∏±‡∏ô‡∏ô‡∏µ‡πâ‡πÄ‡∏´‡∏°‡∏∑‡∏≠‡∏ô‡πÄ‡∏î‡∏¥‡∏°‡∏à‡∏≤‡∏Å‡∏Ñ‡∏£‡∏±‡πâ‡∏á‡∏Å‡πà‡∏≠‡∏ô)
    dfs = []
    for file_obj in uploaded_files:
        try:
            source_workbook = pd.ExcelFile(file_obj)
            for i, sheet_name in enumerate(source_workbook.sheet_names):
                df = source_workbook.parse(sheet_name, header=None)
                if i == 0: df = df.iloc[2:]
                dfs.append(df)
        except Exception as e:
            st.error(f"‡πÄ‡∏Å‡∏¥‡∏î‡∏Ç‡πâ‡∏≠‡∏ú‡∏¥‡∏î‡∏û‡∏•‡∏≤‡∏î‡πÉ‡∏ô‡∏Å‡∏≤‡∏£‡∏õ‡∏£‡∏∞‡∏°‡∏ß‡∏•‡∏ú‡∏•‡πÑ‡∏ü‡∏•‡πå {file_obj.name}: {e}"); return None
    if not dfs: st.warning("‡πÑ‡∏°‡πà‡∏û‡∏ö‡∏Ç‡πâ‡∏≠‡∏°‡∏π‡∏•‡πÉ‡∏ô‡πÑ‡∏ü‡∏•‡πå‡∏ó‡∏µ‡πà‡∏≠‡∏±‡∏õ‡πÇ‡∏´‡∏•‡∏î"); return None
    stacked_df = pd.concat(dfs, ignore_index=True)
    stacked_df = stacked_df.dropna(subset=[stacked_df.columns[12]])
    stacked_df[stacked_df.columns[12]] = pd.to_numeric(stacked_df[stacked_df.columns[12]], errors='coerce')
    stacked_df[stacked_df.columns[18]] = pd.to_numeric(stacked_df[stacked_df.columns[18]], errors='coerce')
    new_column_labels = ["‡∏•‡∏≥‡∏î‡∏±‡∏ö", "‡∏ß‡∏±‡∏ô‡∏ó‡∏µ‡πà‡∏à‡πà‡∏≤‡∏¢‡∏¢‡∏≤", "‡πÄ‡∏ß‡∏•‡∏≤", "‡πÄ‡∏•‡∏Ç‡∏ó‡∏µ‡πà‡πÄ‡∏≠‡∏Å‡∏™‡∏≤‡∏£", "VN / AN", "HN", "‡∏ä‡∏∑‡πà‡∏≠", "‡∏≠‡∏≤‡∏¢‡∏∏", "‡∏™‡∏¥‡∏ó‡∏ò‡∏¥‡πå",
                         "‡πÅ‡∏û‡∏ó‡∏¢‡πå", "Clinic", "Ward", "Material", "‡∏£‡∏≤‡∏¢‡∏Å‡∏≤‡∏£‡∏¢‡∏≤", "‡∏à‡∏≥‡∏ô‡∏ß‡∏ô", "‡∏´‡∏ô‡πà‡∏ß‡∏¢", "‡∏£‡∏≤‡∏Ñ‡∏≤‡∏Ç‡∏≤‡∏¢R", "‡∏£‡∏≤‡∏Ñ‡∏≤‡∏£‡∏ß‡∏°",
                         "Store"]
    if len(stacked_df.columns) != len(new_column_labels): st.error(
        f"‡∏à‡∏≥‡∏ô‡∏ß‡∏ô‡∏Ñ‡∏≠‡∏•‡∏±‡∏°‡∏ô‡πå‡πÑ‡∏°‡πà‡∏ï‡∏£‡∏á‡∏Å‡∏±‡∏ô ‡∏Ñ‡∏≤‡∏î‡∏ß‡πà‡∏≤‡∏ï‡πâ‡∏≠‡∏á‡∏°‡∏µ {len(new_column_labels)} ‡πÅ‡∏ï‡πà‡∏û‡∏ö {len(stacked_df.columns)}"); return None
    stacked_df.columns = new_column_labels
    valid_material_values = [1400000010, 1400000020, 1400000021, 1400000025, 1400000029, 1400000030, 1400000040,
                             1400000044, 1400000052, 1400000053, 1400000055, 1400000098, 1400000099, 1400000148,
                             1400000187, 1400000201, 1400000220, 1400000221, 1400000228, 1400000247, 1400000264,
                             1400000068, 1400000069, 1400000093, 1400000106, 1400000113, 1400000115, 1400000116,
                             1400000118, 1400000124, 1400000126, 1400000130, 1400000165, 1400000166, 1400000167,
                             1400000168, 1400000169, 1400000170, 1400000171, 1400000172, 1400000194, 1400000284,
                             1400000288, 1400000294, 1400000295, 1400000331, 1400000335, 1400000344, 1400000345,
                             1400000265]
    merged_df = stacked_df[stacked_df["Material"].isin(valid_material_values)].copy()
    final_cols = ['‡∏ß‡∏±‡∏ô‡∏ó‡∏µ‡πà‡∏à‡πà‡∏≤‡∏¢‡∏¢‡∏≤', 'VN / AN', 'HN', '‡∏ä‡∏∑‡πà‡∏≠', '‡∏™‡∏¥‡∏ó‡∏ò‡∏¥‡πå', "‡πÅ‡∏û‡∏ó‡∏¢‡πå", 'Material', '‡∏£‡∏≤‡∏¢‡∏Å‡∏≤‡∏£‡∏¢‡∏≤', '‡∏à‡∏≥‡∏ô‡∏ß‡∏ô']
    merged_df = merged_df[final_cols]
    return merged_df


def process_drug_rate_analysis(data_files, master_file):
    dfs = [];
    for file_obj in data_files:
        try:
            source_workbook = pd.ExcelFile(file_obj)
            for i, sheet_name in enumerate(source_workbook.sheet_names):
                df = source_workbook.parse(sheet_name, header=None)
                if i == 0: df = df.iloc[2:]
                dfs.append(df)
        except Exception as e:
            st.error(f"‡πÄ‡∏Å‡∏¥‡∏î‡∏Ç‡πâ‡∏≠‡∏ú‡∏¥‡∏î‡∏û‡∏•‡∏≤‡∏î‡πÉ‡∏ô‡∏Å‡∏≤‡∏£‡∏õ‡∏£‡∏∞‡∏°‡∏ß‡∏•‡∏ú‡∏•‡πÑ‡∏ü‡∏•‡πå {file_obj.name}: {e}"); return None, {}
    if not dfs: st.warning("‡πÑ‡∏°‡πà‡∏û‡∏ö‡∏Ç‡πâ‡∏≠‡∏°‡∏π‡∏•‡πÉ‡∏ô‡πÑ‡∏ü‡∏•‡πå‡∏ó‡∏µ‡πà‡∏≠‡∏±‡∏õ‡πÇ‡∏´‡∏•‡∏î"); return None, {}
    stacked_df = pd.concat(dfs, ignore_index=True)
    try:
        dfmaster = pd.read_excel(master_file, sheet_name="Drug master")
    except Exception as e:
        st.error(f"‡πÑ‡∏°‡πà‡∏™‡∏≤‡∏°‡∏≤‡∏£‡∏ñ‡∏≠‡πà‡∏≤‡∏ô‡∏ä‡∏µ‡∏ó 'Drug master' ‡∏à‡∏≤‡∏Å‡πÑ‡∏ü‡∏•‡πå Master ‡πÑ‡∏î‡πâ: {e}"); return None, {}
    stacked_df = stacked_df.dropna(subset=[stacked_df.columns[12]])
    stacked_df[stacked_df.columns[12]] = pd.to_numeric(stacked_df[stacked_df.columns[12]], errors='coerce')
    stacked_df[stacked_df.columns[18]] = pd.to_numeric(stacked_df[stacked_df.columns[18]], errors='coerce')
    new_column_labels = ["‡∏•‡∏≥‡∏î‡∏±‡∏ö", "‡∏ß‡∏±‡∏ô‡∏ó‡∏µ‡πà‡∏à‡πà‡∏≤‡∏¢‡∏¢‡∏≤", "‡πÄ‡∏ß‡∏•‡∏≤", "‡πÄ‡∏•‡∏Ç‡∏ó‡∏µ‡πà‡πÄ‡∏≠‡∏Å‡∏™‡∏≤‡∏£", "VN / AN", "HN", "‡∏ä‡∏∑‡πà‡∏≠", "‡∏≠‡∏≤‡∏¢‡∏∏", "‡∏™‡∏¥‡∏ó‡∏ò‡∏¥‡πå",
                         "‡πÅ‡∏û‡∏ó‡∏¢‡πå", "Clinic", "Ward", "Material", "‡∏£‡∏≤‡∏¢‡∏Å‡∏≤‡∏£‡∏¢‡∏≤", "‡∏à‡∏≥‡∏ô‡∏ß‡∏ô", "‡∏´‡∏ô‡πà‡∏ß‡∏¢", "‡∏£‡∏≤‡∏Ñ‡∏≤‡∏Ç‡∏≤‡∏¢R", "‡∏£‡∏≤‡∏Ñ‡∏≤‡∏£‡∏ß‡∏°",
                         "Store"]
    if len(stacked_df.columns) != len(new_column_labels): st.error(
        f"‡∏à‡∏≥‡∏ô‡∏ß‡∏ô‡∏Ñ‡∏≠‡∏•‡∏±‡∏°‡∏ô‡πå‡πÑ‡∏°‡πà‡∏ï‡∏£‡∏á‡∏Å‡∏±‡∏ô ‡∏Ñ‡∏≤‡∏î‡∏ß‡πà‡∏≤‡∏ï‡πâ‡∏≠‡∏á‡∏°‡∏µ {len(new_column_labels)} ‡πÅ‡∏ï‡πà‡∏û‡∏ö {len(stacked_df.columns)}"); return None, {}
    stacked_df.columns = new_column_labels;
    merged_df = pd.merge(stacked_df, dfmaster, on="Material", how="left");
    merged_df['Store'] = merged_df['Store'].astype('object')
    valid_store_values = [2403, 2401, 2408, 2409, 2417, 2402];
    merged_df.loc[~merged_df["Store"].isin(valid_store_values), "Store"] = "‡∏≠‡∏∑‡πà‡∏ô‡πÜ"
    merged_df["‡∏£‡∏≤‡∏Ñ‡∏≤‡∏ó‡∏∏‡∏ô‡∏£‡∏ß‡∏°"] = pd.to_numeric(merged_df["‡∏à‡∏≥‡∏ô‡∏ß‡∏ô"], errors='coerce') * pd.to_numeric(merged_df["‡∏ï‡πâ‡∏ô‡∏ó‡∏∏‡∏ô"],
                                                                                                 errors='coerce')
    merged_df['‡∏ß‡∏±‡∏ô‡∏ó‡∏µ‡πà‡∏à‡πà‡∏≤‡∏¢‡∏¢‡∏≤'] = pd.to_datetime(merged_df['‡∏ß‡∏±‡∏ô‡∏ó‡∏µ‡πà‡∏à‡πà‡∏≤‡∏¢‡∏¢‡∏≤'], errors='coerce');
    merged_df['Month'] = merged_df['‡∏ß‡∏±‡∏ô‡∏ó‡∏µ‡πà‡∏à‡πà‡∏≤‡∏¢‡∏¢‡∏≤'].dt.to_period('M')
    merged_df = merged_df[merged_df['Store'] != "‡∏≠‡∏∑‡πà‡∏ô‡πÜ"]
    direct_map = {
        '(‡∏ï‡∏£‡∏ß‡∏à‡∏ó‡∏µ‡πà‡∏£‡∏û.‡∏à‡∏∏‡∏¨‡∏≤‡∏†‡∏£‡∏ì‡πå) ‡πÇ‡∏Ñ‡∏£‡∏á‡∏Å‡∏≤‡∏£‡∏Ñ‡∏±‡∏î‡∏Å‡∏£‡∏≠‡∏á‡∏°‡∏∞‡πÄ‡∏£‡πá‡∏á‡∏õ‡∏≤‡∏Å‡∏°‡∏î‡∏•‡∏π‡∏Å ‡∏ì ‡∏£‡∏û.‡∏à‡∏∏‡∏¨‡∏≤‡∏†‡∏£‡∏ì‡πå‡πÅ‡∏•‡∏∞‡∏Ñ‡∏ì‡∏∞‡πÅ‡∏û‡∏ó‡∏¢‡πå‡∏®‡∏≤‡∏™‡∏ï‡∏£‡πå‡∏ß‡∏ä‡∏¥‡∏£‡∏û‡∏¢‡∏≤‡∏ö‡∏≤‡∏•': '‡∏à‡πà‡∏≤‡∏¢‡πÄ‡∏≠‡∏á',
        '[TopUp] ‡∏™‡∏ß‡∏±‡∏™‡∏î‡∏¥‡∏Å‡∏≤‡∏£‡πÄ‡∏à‡πâ‡∏≤‡∏´‡∏ô‡πâ‡∏≤‡∏ó‡∏µ‡πà‡∏£‡∏≤‡∏ä‡∏ß‡∏¥‡∏ó‡∏¢‡∏≤‡∏•‡∏±‡∏¢‡∏à‡∏∏‡∏¨‡∏≤‡∏†‡∏£‡∏ì‡πå': '‡∏™‡∏ß‡∏±‡∏™‡∏î‡∏¥‡∏Å‡∏≤‡∏£‡πÄ‡∏à‡πâ‡∏≤‡∏´‡∏ô‡πâ‡∏≤‡∏ó‡∏µ‡πà',
        '‡∏≠‡∏á‡∏Ñ‡πå‡∏Å‡∏≤‡∏£‡∏õ‡∏Å‡∏Ñ‡∏£‡∏≠‡∏á‡∏™‡πà‡∏ß‡∏ô‡∏ó‡πâ‡∏≠‡∏á‡∏ñ‡∏¥‡πà‡∏ô‡∏ö‡∏≥‡∏ô‡∏≤‡∏ç(‡πÄ‡∏ö‡∏¥‡∏Å‡∏à‡πà‡∏≤‡∏¢‡∏ï‡∏£‡∏á)': '‡∏Ç‡πâ‡∏≤‡∏£‡∏≤‡∏ä‡∏Å‡∏≤‡∏£'}
    merged_df["‡∏™‡∏¥‡∏ó‡∏ò‡∏¥‡πå"] = merged_df["‡∏™‡∏¥‡∏ó‡∏ò‡∏¥‡πå"].map(direct_map).fillna(merged_df["‡∏™‡∏¥‡∏ó‡∏ò‡∏¥‡πå"])
    opd_merged_df = merged_df[
        merged_df['‡πÄ‡∏•‡∏Ç‡∏ó‡∏µ‡πà‡πÄ‡∏≠‡∏Å‡∏™‡∏≤‡∏£'].isna() | (merged_df['‡πÄ‡∏•‡∏Ç‡∏ó‡∏µ‡πà‡πÄ‡∏≠‡∏Å‡∏™‡∏≤‡∏£'].astype(str).str.strip().isin(['', '0']))]
    opd_2409 = opd_merged_df[opd_merged_df['Store'].notna() & (opd_merged_df['Store'] == 2409)];
    opd_not_2409 = opd_merged_df[opd_merged_df['Store'].notna() & (opd_merged_df['Store'] != 2409)]
    ipd_merged_df = merged_df[
        merged_df['Clinic'].isna() | (merged_df['Clinic'].astype(str).str.strip().isin(['', '0']))]
    ipd_2409 = ipd_merged_df[ipd_merged_df['Store'].notna() & (ipd_merged_df['Store'] == 2409)];
    ipd_not_2409 = ipd_merged_df[ipd_merged_df['Store'].notna() & (ipd_merged_df['Store'] != 2409)]

    def count_unique_by_month(df, subset_cols):
        return df.drop_duplicates(subset=subset_cols).groupby('Month').size().reset_index(name='Unique_Count')

    uniqueOPD = count_unique_by_month(opd_not_2409, ['VN / AN', 'HN', 'Clinic', 'Month']);
    uniqueOPD2409 = count_unique_by_month(opd_2409, ['VN / AN', 'HN', 'Clinic', 'Month'])
    uniqueIPD = count_unique_by_month(ipd_not_2409, ['‡πÄ‡∏•‡∏Ç‡∏ó‡∏µ‡πà‡πÄ‡∏≠‡∏Å‡∏™‡∏≤‡∏£', 'HN', 'Ward', 'Month']);
    uniqueIPD2409 = count_unique_by_month(ipd_2409, ['‡πÄ‡∏•‡∏Ç‡∏ó‡∏µ‡πà‡πÄ‡∏≠‡∏Å‡∏™‡∏≤‡∏£', 'HN', 'Ward', 'Month'])
    merged_df["‡∏´‡∏ô‡πà‡∏ß‡∏¢"] = pd.to_numeric(merged_df["‡∏´‡∏ô‡πà‡∏ß‡∏¢"].astype(str).str.replace(r'.*/ ', '', regex=True),
                                       errors='coerce').fillna(1).astype(int)
    merged_df["‡∏à‡∏≥‡∏ô‡∏ß‡∏ô"] = merged_df["‡∏à‡∏≥‡∏ô‡∏ß‡∏ô"] * merged_df["‡∏´‡∏ô‡πà‡∏ß‡∏¢"];
    merged_df['HN'] = merged_df['HN'].astype(str).str.replace('.0', '', regex=False)
    grouped_countHN_df = merged_df.pivot_table(index=['Material', 'Material description'], columns='Month', values='HN',
                                               aggfunc=pd.Series.nunique).reset_index()
    grouped_sumRate_df = merged_df.pivot_table(index=['Material', 'Material description', '‡∏´‡∏ô‡πà‡∏ß‡∏¢‡∏¢‡πà‡∏≠‡∏¢'], columns='Month',
                                               values='‡∏à‡∏≥‡∏ô‡∏ß‡∏ô', aggfunc='sum').reset_index()
    grouped_sumRateSplit_df = merged_df.pivot_table(index=['Material', "Store", 'Material description', '‡∏´‡∏ô‡πà‡∏ß‡∏¢‡∏¢‡πà‡∏≠‡∏¢'],
                                                    columns='Month', values='‡∏à‡∏≥‡∏ô‡∏ß‡∏ô', aggfunc='sum').reset_index()
    output_dfs = {"Rate ‡πÅ‡∏¢‡∏Å‡πÄ‡∏î‡∏∑‡∏≠‡∏ô": grouped_sumRate_df, "Rate (M-Sloc)": grouped_sumRateSplit_df,
                  "‡∏à‡∏≥‡∏ô‡∏ß‡∏ô‡πÄ‡∏Ñ‡∏™‡∏ï‡πà‡∏≠‡πÄ‡∏î‡∏∑‡∏≠‡∏ô": grouped_countHN_df, "Raw": merged_df,
                  "Summary_Data": {'‡∏à‡∏≥‡∏ô‡∏ß‡∏ô‡πÉ‡∏ö‡∏¢‡∏≤ OPD': uniqueOPD, '‡∏à‡∏≥‡∏ô‡∏ß‡∏ô‡πÉ‡∏ö‡∏¢‡∏≤ OPD 2409': uniqueOPD2409,
                                   '‡∏à‡∏≥‡∏ô‡∏ß‡∏ô‡πÉ‡∏ö‡∏¢‡∏≤ IPD': uniqueIPD, '‡∏à‡∏≥‡∏ô‡∏ß‡∏ô‡πÉ‡∏ö‡∏¢‡∏≤ IPD 2409': uniqueIPD2409, }}
    return merged_df, output_dfs


def process_epi_usage(uploaded_files):
    dfs = [];
    for file_obj in uploaded_files:
        try:
            source_workbook = pd.ExcelFile(file_obj)
            for i, sheet_name in enumerate(source_workbook.sheet_names):
                df = source_workbook.parse(sheet_name, header=None)
                if i == 0: df = df.iloc[2:]
                dfs.append(df)
        except Exception as e:
            st.error(f"‡πÄ‡∏Å‡∏¥‡∏î‡∏Ç‡πâ‡∏≠‡∏ú‡∏¥‡∏î‡∏û‡∏•‡∏≤‡∏î‡πÉ‡∏ô‡∏Å‡∏≤‡∏£‡∏õ‡∏£‡∏∞‡∏°‡∏ß‡∏•‡∏ú‡∏•‡πÑ‡∏ü‡∏•‡πå {file_obj.name}: {e}"); return None
    if not dfs: st.warning("‡πÑ‡∏°‡πà‡∏û‡∏ö‡∏Ç‡πâ‡∏≠‡∏°‡∏π‡∏•‡πÉ‡∏ô‡πÑ‡∏ü‡∏•‡πå‡∏ó‡∏µ‡πà‡∏≠‡∏±‡∏õ‡πÇ‡∏´‡∏•‡∏î"); return None
    stacked_df = pd.concat(dfs, ignore_index=True)
    stacked_df = stacked_df.dropna(subset=[stacked_df.columns[12]])
    stacked_df[stacked_df.columns[12]] = pd.to_numeric(stacked_df[stacked_df.columns[12]], errors='coerce')
    stacked_df[stacked_df.columns[18]] = pd.to_numeric(stacked_df[stacked_df.columns[18]], errors='coerce')
    new_column_labels = ["‡∏•‡∏≥‡∏î‡∏±‡∏ö", "‡∏ß‡∏±‡∏ô‡∏ó‡∏µ‡πà‡∏à‡πà‡∏≤‡∏¢‡∏¢‡∏≤", "‡πÄ‡∏ß‡∏•‡∏≤", "‡πÄ‡∏•‡∏Ç‡∏ó‡∏µ‡πà‡πÄ‡∏≠‡∏Å‡∏™‡∏≤‡∏£", "VN / AN", "HN", "‡∏ä‡∏∑‡πà‡∏≠", "‡∏≠‡∏≤‡∏¢‡∏∏", "‡∏™‡∏¥‡∏ó‡∏ò‡∏¥‡πå",
                         "‡πÅ‡∏û‡∏ó‡∏¢‡πå", "Clinic", "Ward", "Material", "‡∏£‡∏≤‡∏¢‡∏Å‡∏≤‡∏£‡∏¢‡∏≤", "‡∏à‡∏≥‡∏ô‡∏ß‡∏ô", "‡∏´‡∏ô‡πà‡∏ß‡∏¢", "‡∏£‡∏≤‡∏Ñ‡∏≤‡∏Ç‡∏≤‡∏¢R", "‡∏£‡∏≤‡∏Ñ‡∏≤‡∏£‡∏ß‡∏°",
                         "Store"]
    if len(stacked_df.columns) != len(new_column_labels): st.error(
        f"‡∏à‡∏≥‡∏ô‡∏ß‡∏ô‡∏Ñ‡∏≠‡∏•‡∏±‡∏°‡∏ô‡πå‡πÑ‡∏°‡πà‡∏ï‡∏£‡∏á‡∏Å‡∏±‡∏ô ‡∏Ñ‡∏≤‡∏î‡∏ß‡πà‡∏≤‡∏ï‡πâ‡∏≠‡∏á‡∏°‡∏µ {len(new_column_labels)} ‡πÅ‡∏ï‡πà‡∏û‡∏ö {len(stacked_df.columns)}"); return None
    stacked_df.columns = new_column_labels
    valid_epi_materials = [1400000084, 1400000083, 1400000087, 1400000086, 1400000088, 1400000081, 1400000082,
                           1400000090, 1400000085, 1400000089]
    epi_df = stacked_df[stacked_df["Material"].isin(valid_epi_materials)].copy()
    summary_df = epi_df.groupby(['Material', '‡∏£‡∏≤‡∏¢‡∏Å‡∏≤‡∏£‡∏¢‡∏≤'])['‡∏à‡∏≥‡∏ô‡∏ß‡∏ô'].sum().reset_index()
    summary_df.rename(columns={'‡∏à‡∏≥‡∏ô‡∏ß‡∏ô': '‡∏à‡∏≥‡∏ô‡∏ß‡∏ô‡∏£‡∏ß‡∏°'}, inplace=True)
    return summary_df


def process_narcotics_report(xls_files, receipt_report_file, master_file):
    def convert_date_to_thai(date_str):
        if not pd.isna(date_str):
            try:
                date_obj = pd.to_datetime(date_str);
                month_mapping = {1: '‡∏°‡∏Å‡∏£‡∏≤‡∏Ñ‡∏°', 2: '‡∏Å‡∏∏‡∏°‡∏†‡∏≤‡∏û‡∏±‡∏ô‡∏ò‡πå', 3: '‡∏°‡∏µ‡∏ô‡∏≤‡∏Ñ‡∏°', 4: '‡πÄ‡∏°‡∏©‡∏≤‡∏¢‡∏ô', 5: '‡∏û‡∏§‡∏©‡∏†‡∏≤‡∏Ñ‡∏°', 6: '‡∏°‡∏¥‡∏ñ‡∏∏‡∏ô‡∏≤‡∏¢‡∏ô',
                                 7: '‡∏Å‡∏£‡∏Å‡∏é‡∏≤‡∏Ñ‡∏°', 8: '‡∏™‡∏¥‡∏á‡∏´‡∏≤‡∏Ñ‡∏°', 9: '‡∏Å‡∏±‡∏ô‡∏¢‡∏≤‡∏¢‡∏ô', 10: '‡∏ï‡∏∏‡∏•‡∏≤‡∏Ñ‡∏°', 11: '‡∏û‡∏§‡∏®‡∏à‡∏¥‡∏Å‡∏≤‡∏¢‡∏ô', 12: '‡∏ò‡∏±‡∏ô‡∏ß‡∏≤‡∏Ñ‡∏°'}
                return f"{date_obj.strftime('%d')} {month_mapping.get(date_obj.month, date_obj.month)} {str(date_obj.year + 543)}"
            except (ValueError, TypeError):
                return ''
        return ''

    stacked_df_list = []
    for file_obj in xls_files:
        try:
            df = pd.read_excel(file_obj)
            df['‡πÇ‡∏£‡∏á‡∏û‡∏¢‡∏≤‡∏ö‡∏≤‡∏•‡∏à‡∏∏‡∏¨‡∏≤‡∏†‡∏£‡∏ì‡πå'] = pd.to_datetime(df['‡πÇ‡∏£‡∏á‡∏û‡∏¢‡∏≤‡∏ö‡∏≤‡∏•‡∏à‡∏∏‡∏¨‡∏≤‡∏†‡∏£‡∏ì‡πå'], errors='coerce')
            df = df.dropna(subset=['‡πÇ‡∏£‡∏á‡∏û‡∏¢‡∏≤‡∏ö‡∏≤‡∏•‡∏à‡∏∏‡∏¨‡∏≤‡∏†‡∏£‡∏ì‡πå']).sort_values(by='‡πÇ‡∏£‡∏á‡∏û‡∏¢‡∏≤‡∏ö‡∏≤‡∏•‡∏à‡∏∏‡∏¨‡∏≤‡∏†‡∏£‡∏ì‡πå').reset_index(drop=True)
            df.columns = range(df.shape[1]);
            value_to_expand = str(df.at[0, 1]).replace("‡∏£‡∏ß‡∏°", "").strip();
            df[1] = value_to_expand
            df = df[df[4].apply(lambda x: isinstance(x, str) and x.strip() != '')]
            df[4] = pd.to_numeric(df[4], errors='coerce').dropna().astype(int);
            df = df.drop(0, axis=1)
            negative_values = df[6] < 0;
            df.insert(6, '6.5', 0);
            df.loc[negative_values, '6.5'] = df.loc[negative_values, 6];
            df.loc[df[6] < 0, 6] = 0
            new_row = pd.DataFrame(
                {1: [value_to_expand], 5: ["‡∏£‡∏ß‡∏°‡∏ó‡∏±‡πâ‡∏á‡∏™‡∏¥‡πâ‡∏ô"], 6: [df[6].sum()], '6.5': [df['6.5'].sum()],
                 7: [df.iat[0, 7]], 9: [""]})
            df = pd.concat([df, new_row], ignore_index=True)
            df.columns = ['‡∏ä‡∏∑‡πà‡∏≠‡∏¢‡∏≤‡πÄ‡∏™‡∏û‡∏ï‡∏¥‡∏î‡πÉ‡∏´‡πâ‡πÇ‡∏ó‡∏©‡∏õ‡∏£‡∏∞‡πÄ‡∏†‡∏ó 2', '‡∏ß‡∏±‡∏ô ‡πÄ‡∏î‡∏∑‡∏≠‡∏ô ‡∏õ‡∏µ', 'AN/VN', 'HN', '‡∏ä‡∏∑‡πà‡∏≠', '‡∏à‡πà‡∏≤‡∏¢', '‡∏£‡∏±‡∏ö', '‡∏´‡∏ô‡πà‡∏ß‡∏¢',
                          '‡∏£‡∏≤‡∏Ñ‡∏≤', '‡∏ó‡∏µ‡πà‡∏≠‡∏¢‡∏π‡πà']
            df = df[['‡∏ß‡∏±‡∏ô ‡πÄ‡∏î‡∏∑‡∏≠‡∏ô ‡∏õ‡∏µ', '‡∏ä‡∏∑‡πà‡∏≠‡∏¢‡∏≤‡πÄ‡∏™‡∏û‡∏ï‡∏¥‡∏î‡πÉ‡∏´‡πâ‡πÇ‡∏ó‡∏©‡∏õ‡∏£‡∏∞‡πÄ‡∏†‡∏ó 2', '‡∏ä‡∏∑‡πà‡∏≠', '‡∏£‡∏±‡∏ö', '‡∏à‡πà‡∏≤‡∏¢', '‡∏´‡∏ô‡πà‡∏ß‡∏¢', '‡∏ó‡∏µ‡πà‡∏≠‡∏¢‡∏π‡πà']]
            df['‡∏à‡πà‡∏≤‡∏¢‡πÑ‡∏õ'] = df['‡∏ä‡∏∑‡πà‡∏≠'].astype(str) + " " + df['‡∏ó‡∏µ‡πà‡∏≠‡∏¢‡∏π‡πà'].astype(str)
            df = df[['‡∏ß‡∏±‡∏ô ‡πÄ‡∏î‡∏∑‡∏≠‡∏ô ‡∏õ‡∏µ', '‡∏ä‡∏∑‡πà‡∏≠‡∏¢‡∏≤‡πÄ‡∏™‡∏û‡∏ï‡∏¥‡∏î‡πÉ‡∏´‡πâ‡πÇ‡∏ó‡∏©‡∏õ‡∏£‡∏∞‡πÄ‡∏†‡∏ó 2', '‡∏à‡πà‡∏≤‡∏¢‡πÑ‡∏õ', '‡∏´‡∏ô‡πà‡∏ß‡∏¢', '‡∏£‡∏±‡∏ö', '‡∏´‡∏ô‡πà‡∏ß‡∏¢', '‡∏à‡πà‡∏≤‡∏¢', '‡∏´‡∏ô‡πà‡∏ß‡∏¢']]
            df['‡∏ß‡∏±‡∏ô ‡πÄ‡∏î‡∏∑‡∏≠‡∏ô ‡∏õ‡∏µ'] = df['‡∏ß‡∏±‡∏ô ‡πÄ‡∏î‡∏∑‡∏≠‡∏ô ‡∏õ‡∏µ'].apply(convert_date_to_thai);
            df.insert(3, '‡∏£‡∏±‡∏ö‡∏à‡∏≤‡∏Å ‡∏≠‡∏¢', '');
            df.insert(2, '‡∏£‡∏´‡∏±‡∏™', '')
            df.columns = ['‡∏ß‡∏±‡∏ô ‡πÄ‡∏î‡∏∑‡∏≠‡∏ô ‡∏õ‡∏µ', '‡∏ä‡∏∑‡πà‡∏≠‡∏¢‡∏≤‡πÄ‡∏™‡∏û‡∏ï‡∏¥‡∏î‡πÉ‡∏´‡πâ‡πÇ‡∏ó‡∏©‡∏õ‡∏£‡∏∞‡πÄ‡∏†‡∏ó 2', '‡∏£‡∏´‡∏±‡∏™', '‡∏à‡πà‡∏≤‡∏¢‡πÑ‡∏õ', '‡∏´‡∏ô‡πà‡∏ß‡∏¢1', '‡∏£‡∏±‡∏ö', '‡∏´‡∏ô‡πà‡∏ß‡∏¢2',
                          '‡∏à‡πà‡∏≤‡∏¢', '‡∏´‡∏ô‡πà‡∏ß‡∏¢3', '‡∏£‡∏±‡∏ö‡∏à‡∏≤‡∏Å ‡∏≠‡∏¢']
            df = df[
                ['‡∏ß‡∏±‡∏ô ‡πÄ‡∏î‡∏∑‡∏≠‡∏ô ‡∏õ‡∏µ', '‡∏ä‡∏∑‡πà‡∏≠‡∏¢‡∏≤‡πÄ‡∏™‡∏û‡∏ï‡∏¥‡∏î‡πÉ‡∏´‡πâ‡πÇ‡∏ó‡∏©‡∏õ‡∏£‡∏∞‡πÄ‡∏†‡∏ó 2', '‡∏£‡∏´‡∏±‡∏™', '‡∏£‡∏±‡∏ö‡∏à‡∏≤‡∏Å ‡∏≠‡∏¢', '‡∏à‡πà‡∏≤‡∏¢‡πÑ‡∏õ', '‡∏´‡∏ô‡πà‡∏ß‡∏¢1', '‡∏£‡∏±‡∏ö', '‡∏´‡∏ô‡πà‡∏ß‡∏¢2',
                 '‡∏à‡πà‡∏≤‡∏¢', '‡∏´‡∏ô‡πà‡∏ß‡∏¢3']]
            stacked_df_list.append(df)
        except Exception as e:
            st.warning(f"‡πÑ‡∏°‡πà‡∏™‡∏≤‡∏°‡∏≤‡∏£‡∏ñ‡∏õ‡∏£‡∏∞‡∏°‡∏ß‡∏•‡∏ú‡∏•‡πÑ‡∏ü‡∏•‡πå {file_obj.name}: {e}"); continue
    if not stacked_df_list: st.error("‡πÑ‡∏°‡πà‡∏™‡∏≤‡∏°‡∏≤‡∏£‡∏ñ‡∏õ‡∏£‡∏∞‡∏°‡∏ß‡∏•‡∏ú‡∏•‡πÑ‡∏ü‡∏•‡πå‡∏Ç‡πâ‡∏≠‡∏°‡∏π‡∏•‡∏Å‡∏≤‡∏£‡∏à‡πà‡∏≤‡∏¢‡∏¢‡∏≤‡πÑ‡∏î‡πâ‡πÄ‡∏•‡∏¢"); return None
    stacked_df = pd.concat(stacked_df_list, axis=0, ignore_index=True)
    try:
        dfT = pd.read_excel(receipt_report_file, sheet_name='Sheet1')
        dfmaster = pd.read_excel(master_file, sheet_name="Drug master")[["Material", "TradeName"]]
        dfT = pd.merge(dfT, dfmaster, how="left")[
            ["Posting Date", "TradeName", "Batch", 'Receiving stor. loc.', "Quantity"]]
        dfT.columns = ['‡∏ß‡∏±‡∏ô ‡πÄ‡∏î‡∏∑‡∏≠‡∏ô ‡∏õ‡∏µ', "‡∏ä‡∏∑‡πà‡∏≠‡∏¢‡∏≤‡πÄ‡∏™‡∏û‡∏ï‡∏¥‡∏î‡πÉ‡∏´‡πâ‡πÇ‡∏ó‡∏©‡∏õ‡∏£‡∏∞‡πÄ‡∏†‡∏ó 2", '‡∏£‡∏´‡∏±‡∏™', '‡∏à‡πà‡∏≤‡∏¢‡πÑ‡∏õ', '‡∏£‡∏±‡∏ö‡∏à‡∏≤‡∏Å ‡∏≠‡∏¢']
        dfT['‡∏ß‡∏±‡∏ô ‡πÄ‡∏î‡∏∑‡∏≠‡∏ô ‡∏õ‡∏µ'] = dfT['‡∏ß‡∏±‡∏ô ‡πÄ‡∏î‡∏∑‡∏≠‡∏ô ‡∏õ‡∏µ'].apply(convert_date_to_thai);
        dfT.insert(5, '‡∏´‡∏ô‡πà‡∏ß‡∏¢', '');
        dfT.insert(6, '‡∏£‡∏±‡∏ö', '');
        dfT.insert(7, '‡∏à‡πà‡∏≤‡∏¢', '')
        dfT = dfT[['‡∏ß‡∏±‡∏ô ‡πÄ‡∏î‡∏∑‡∏≠‡∏ô ‡∏õ‡∏µ', '‡∏ä‡∏∑‡πà‡∏≠‡∏¢‡∏≤‡πÄ‡∏™‡∏û‡∏ï‡∏¥‡∏î‡πÉ‡∏´‡πâ‡πÇ‡∏ó‡∏©‡∏õ‡∏£‡∏∞‡πÄ‡∏†‡∏ó 2', '‡∏£‡∏´‡∏±‡∏™', '‡∏à‡πà‡∏≤‡∏¢‡πÑ‡∏õ', '‡∏£‡∏±‡∏ö‡∏à‡∏≤‡∏Å ‡∏≠‡∏¢', '‡∏´‡∏ô‡πà‡∏ß‡∏¢', '‡∏£‡∏±‡∏ö', '‡∏´‡∏ô‡πà‡∏ß‡∏¢',
                   '‡∏à‡πà‡∏≤‡∏¢', '‡∏´‡∏ô‡πà‡∏ß‡∏¢']]
    except Exception as e:
        st.error(f"‡πÄ‡∏Å‡∏¥‡∏î‡∏Ç‡πâ‡∏≠‡∏ú‡∏¥‡∏î‡∏û‡∏•‡∏≤‡∏î‡πÉ‡∏ô‡∏Å‡∏≤‡∏£‡∏õ‡∏£‡∏∞‡∏°‡∏ß‡∏•‡∏ú‡∏•‡πÑ‡∏ü‡∏•‡πå‡∏£‡∏≤‡∏¢‡∏á‡∏≤‡∏ô‡∏£‡∏±‡∏ö‡πÄ‡∏Ç‡πâ‡∏≤: {e}"); return None
    total_df = stacked_df[stacked_df['‡∏à‡πà‡∏≤‡∏¢‡πÑ‡∏õ'].str.strip() == "‡∏£‡∏ß‡∏°‡∏ó‡∏±‡πâ‡∏á‡∏™‡∏¥‡πâ‡∏ô"].copy()
    return {'‡∏£‡∏≤‡∏¢‡∏á‡∏≤‡∏ô‡πÅ‡∏¢‡∏Å': stacked_df, '‡∏£‡∏≤‡∏¢‡∏á‡∏≤‡∏ô‡∏£‡∏ß‡∏°': total_df, '‡∏£‡∏≤‡∏¢‡∏á‡∏≤‡∏ô‡∏£‡∏±‡∏ö‡πÄ‡∏Ç‡πâ‡∏≤': dfT}


def process_kpi_report(rate_file, inventory_file, master_file):
    try:
        remain = pd.read_excel(inventory_file, sheet_name="Sheet1");
        remain = remain.groupby('Storage location')['Stock Value on Period End'].sum().reset_index()
        remain = remain.rename(columns={'Storage location': 'Store'})
        source_workbook = pd.ExcelFile(rate_file);
        dfs = [source_workbook.parse(sheet_name, header=None) for sheet_name in source_workbook.sheet_names]
        dfs[0] = dfs[0].iloc[2:];
        stacked_df = pd.concat(dfs, ignore_index=True)
        dfmaster = pd.read_excel(master_file, sheet_name="Drug master")
    except Exception as e:
        st.error(f"‡πÄ‡∏Å‡∏¥‡∏î‡∏Ç‡πâ‡∏≠‡∏ú‡∏¥‡∏î‡∏û‡∏•‡∏≤‡∏î‡πÉ‡∏ô‡∏Å‡∏≤‡∏£‡πÇ‡∏´‡∏•‡∏î‡πÑ‡∏ü‡∏•‡πå‡πÄ‡∏£‡∏¥‡πà‡∏°‡∏ï‡πâ‡∏ô: {e}"); return None
    stacked_df = stacked_df.dropna(subset=[stacked_df.columns[12]])
    stacked_df[stacked_df.columns[12]] = pd.to_numeric(stacked_df[stacked_df.columns[12]], errors='coerce');
    stacked_df[stacked_df.columns[18]] = pd.to_numeric(stacked_df[stacked_df.columns[18]], errors='coerce')
    new_column_labels = ["‡∏•‡∏≥‡∏î‡∏±‡∏ö", "‡∏ß‡∏±‡∏ô‡∏ó‡∏µ‡πà‡∏à‡πà‡∏≤‡∏¢‡∏¢‡∏≤", "‡πÄ‡∏ß‡∏•‡∏≤", "‡πÄ‡∏•‡∏Ç‡∏ó‡∏µ‡πà‡πÄ‡∏≠‡∏Å‡∏™‡∏≤‡∏£", "VN / AN", "HN", "‡∏ä‡∏∑‡πà‡∏≠", "‡∏≠‡∏≤‡∏¢‡∏∏", "‡∏™‡∏¥‡∏ó‡∏ò‡∏¥‡πå",
                         "‡πÅ‡∏û‡∏ó‡∏¢‡πå", "Clinic", "Ward", "Material", "‡∏£‡∏≤‡∏¢‡∏Å‡∏≤‡∏£‡∏¢‡∏≤", "‡∏à‡∏≥‡∏ô‡∏ß‡∏ô", "‡∏´‡∏ô‡πà‡∏ß‡∏¢", "‡∏£‡∏≤‡∏Ñ‡∏≤‡∏Ç‡∏≤‡∏¢R", "‡∏£‡∏≤‡∏Ñ‡∏≤‡∏£‡∏ß‡∏°",
                         "Store"]
    stacked_df.columns = new_column_labels;
    merged_df = pd.merge(stacked_df, dfmaster, on="Material", how="left");
    merged_df['Store'] = merged_df['Store'].astype('object')
    valid_store_values = [2403, 2401, 2408, 2409, 2417, 2402];
    merged_df.loc[~merged_df["Store"].isin(valid_store_values), "Store"] = "‡∏≠‡∏∑‡πà‡∏ô‡πÜ";
    merged_df = merged_df[merged_df['Store'] != "‡∏≠‡∏∑‡πà‡∏ô‡πÜ"]
    merged_df["‡∏´‡∏ô‡πà‡∏ß‡∏¢"] = pd.to_numeric(merged_df["‡∏´‡∏ô‡πà‡∏ß‡∏¢"].astype(str).str.replace(r'.*/ ', '', regex=True),
                                       errors='coerce').fillna(1).astype(int);
    merged_df["‡∏à‡∏≥‡∏ô‡∏ß‡∏ô"] = merged_df["‡∏à‡∏≥‡∏ô‡∏ß‡∏ô"] * merged_df["‡∏´‡∏ô‡πà‡∏ß‡∏¢"]
    if "‡∏ï‡πâ‡∏ô‡∏ó‡∏∏‡∏ô" not in merged_df.columns: st.error("‡πÑ‡∏°‡πà‡∏û‡∏ö‡∏Ñ‡∏≠‡∏•‡∏±‡∏°‡∏ô‡πå '‡∏ï‡πâ‡∏ô‡∏ó‡∏∏‡∏ô' ‡πÉ‡∏ô‡∏Ç‡πâ‡∏≠‡∏°‡∏π‡∏•"); return None
    merged_df["‡∏£‡∏≤‡∏Ñ‡∏≤‡∏ó‡∏∏‡∏ô‡∏£‡∏ß‡∏°"] = merged_df["‡∏à‡∏≥‡∏ô‡∏ß‡∏ô"] * merged_df["‡∏ï‡πâ‡∏ô‡∏ó‡∏∏‡∏ô"]
    grouped_sumRate_df = merged_df.pivot_table(index=['Material', "Store", 'Material description', '‡∏´‡∏ô‡πà‡∏ß‡∏¢‡∏¢‡πà‡∏≠‡∏¢'],
                                               values=['‡∏à‡∏≥‡∏ô‡∏ß‡∏ô', "‡∏£‡∏≤‡∏Ñ‡∏≤‡∏ó‡∏∏‡∏ô‡∏£‡∏ß‡∏°", "‡∏£‡∏≤‡∏Ñ‡∏≤‡∏£‡∏ß‡∏°"], aggfunc='sum').reset_index()
    grouped_Valuesum_df = merged_df.groupby('Store')[['‡∏£‡∏≤‡∏Ñ‡∏≤‡∏ó‡∏∏‡∏ô‡∏£‡∏ß‡∏°', '‡∏£‡∏≤‡∏Ñ‡∏≤‡∏£‡∏ß‡∏°']].sum().reset_index()
    grouped_Valuesum_df.columns = ['Store', 'Sum of Cost price', 'Sum of sale price'];
    remainFinal = pd.merge(remain, grouped_Valuesum_df, on='Store', how='left')
    with np.errstate(divide='ignore', invalid='ignore'):
        remainFinal["‡∏ß‡∏±‡∏ô‡∏™‡∏≥‡∏£‡∏≠‡∏á‡∏Ñ‡∏á‡∏Ñ‡∏•‡∏±‡∏á"] = (remainFinal["Stock Value on Period End"] / remainFinal[
            "Sum of Cost price"]) * 30
    remainFinal.replace([np.inf, -np.inf], 0, inplace=True);
    remainFinal["‡∏ß‡∏±‡∏ô‡∏™‡∏≥‡∏£‡∏≠‡∏á‡∏Ñ‡∏á‡∏Ñ‡∏•‡∏±‡∏á"].fillna(0, inplace=True)
    return {'‡∏¢‡∏≠‡∏î‡∏Ç‡∏≤‡∏¢-‡∏Ñ‡∏á‡∏Ñ‡∏•‡∏±‡∏á-‡∏™‡∏≥‡∏£‡∏≠‡∏á‡∏Ñ‡∏á‡∏Ñ‡∏•‡∏±‡∏á': remainFinal, '‡∏¢‡∏≠‡∏î‡∏Ç‡∏≤‡∏¢': grouped_sumRate_df, 'Raw': merged_df}


def process_abc_analysis(inventory_files, master_file):
    try:
        all_dfs = [pd.read_excel(fp) for fp in inventory_files]
        consolidated_df = pd.concat(all_dfs, ignore_index=True)
        master_df = pd.read_excel(master_file, sheet_name='Drug master', usecols=['Material', 'Drug group'])
        master_df['Material'] = master_df['Material'].astype(str)
    except Exception as e:
        st.error(f"‡πÄ‡∏Å‡∏¥‡∏î‡∏Ç‡πâ‡∏≠‡∏ú‡∏¥‡∏î‡∏û‡∏•‡∏≤‡∏î‡πÉ‡∏ô‡∏Å‡∏≤‡∏£‡πÇ‡∏´‡∏•‡∏î‡πÑ‡∏ü‡∏•‡πå: {e}"); return None
    df = consolidated_df;
    df['Posting Date'] = pd.to_datetime(df['Posting Date'], errors='coerce');
    df.dropna(subset=['Posting Date'], inplace=True)
    df['MonthYear'] = df['Posting Date'].dt.to_period('M');
    df['Amt.in Loc.Cur.'] = pd.to_numeric(df['Amt.in Loc.Cur.'], errors='coerce').fillna(0)
    df['Quantity'] = pd.to_numeric(df['Quantity'], errors='coerce').fillna(0);
    df['Material'] = df['Material'].astype(str)
    monthly_data = df.groupby(['Material', 'Material description', 'Storage location', 'MonthYear']).agg(
        MonthlyNetConsumption=('Amt.in Loc.Cur.', 'sum'), MonthlyNetQuantity=('Quantity', 'sum')).reset_index()
    monthly_qty_pivot = monthly_data.pivot_table(index=['Material', 'Material description', 'Storage location'],
                                                 columns='MonthYear', values='MonthlyNetQuantity', fill_value=0)
    monthly_qty_pivot.columns = [f"Qty_{str(col)}" for col in monthly_qty_pivot.columns];
    monthly_qty_pivot = monthly_qty_pivot.abs()
    final_agg = monthly_data.groupby(['Material', 'Material description', 'Storage location']).agg(
        AvgMonthlyNetQuantity=('MonthlyNetQuantity', 'mean'),
        TotalNetConsumption=('MonthlyNetConsumption', 'sum')).reset_index()
    final_agg['AvgMonthlyNetQuantity'] = final_agg['AvgMonthlyNetQuantity'].abs()
    final_agg = pd.merge(final_agg, monthly_qty_pivot, on=['Material', 'Material description', 'Storage location'],
                         how='left')
    final_agg['NetConsumptionValue'] = final_agg['TotalNetConsumption'].abs()
    abc_data_no_class = final_agg[final_agg['NetConsumptionValue'] > 0].copy()
    if abc_data_no_class.empty: st.warning("‡πÑ‡∏°‡πà‡∏û‡∏ö‡∏Ç‡πâ‡∏≠‡∏°‡∏π‡∏•‡∏Å‡∏≤‡∏£‡πÉ‡∏ä‡πâ‡∏á‡∏≤‡∏ô (consumption data) ‡∏ó‡∏µ‡πà‡∏°‡∏µ‡∏°‡∏π‡∏•‡∏Ñ‡πà‡∏≤‡∏°‡∏≤‡∏Å‡∏Å‡∏ß‡πà‡∏≤ 0"); return None
    abc_data_no_class = pd.merge(abc_data_no_class, master_df, on='Material', how='left');
    abc_data_no_class['Drug group'].fillna('N/A', inplace=True)
    all_locations_classified = []
    for location in abc_data_no_class['Storage location'].unique():
        loc_df = abc_data_no_class[abc_data_no_class['Storage location'] == location].copy();
        total_value_loc = loc_df['NetConsumptionValue'].sum()
        loc_df = loc_df.sort_values(by='NetConsumptionValue', ascending=False).reset_index(drop=True)
        loc_df['PercentageValue'] = loc_df['NetConsumptionValue'] / total_value_loc if total_value_loc > 0 else 0;
        loc_df['CumulativePercentage'] = loc_df['PercentageValue'].cumsum()

        def assign_abc_class(cum_perc):
            if cum_perc <= 0.80:
                return 'A'
            elif cum_perc <= 0.95:
                return 'B'
            else:
                return 'C'

        loc_df['ABC_Class'] = loc_df['CumulativePercentage'].apply(assign_abc_class);
        all_locations_classified.append(loc_df)
    final_results = pd.concat(all_locations_classified)
    output_buffer = io.BytesIO()
    with pd.ExcelWriter(output_buffer, engine='openpyxl') as writer:
        def apply_formats_and_hide_cols(writer, sheet_name, df):
            worksheet = writer.sheets[sheet_name];
            center_align = Alignment(horizontal='center', vertical='center')
            col_map = {'AvgMonthlyNetQuantity': '#,##0', 'NetConsumptionValue': '#,##0.00', 'PercentageValue': '0.00%',
                       'CumulativePercentage': '0.00%'}
            for col in df.columns:
                if isinstance(col, str) and col.startswith('Qty_'): col_map[col] = '#,##0'
            col_letters = {col_name: chr(65 + i) for i, col_name in enumerate(df.columns)}
            for col_name, num_format in col_map.items():
                if col_name in col_letters:
                    col_letter = col_letters[col_name]
                    for row in range(2, worksheet.max_row + 1): worksheet[
                        f'{col_letter}{row}'].number_format = num_format
            for row in range(2, worksheet.max_row + 1): worksheet[
                f'{col_letters["ABC_Class"]}{row}'].alignment = center_align
            for col_name in df.columns:
                if isinstance(col_name, str) and col_name.startswith('Qty_'): worksheet.column_dimensions[
                    col_letters[col_name]].hidden = True
            for col in worksheet.columns:
                if not worksheet.column_dimensions[col[0].column_letter].hidden:
                    max_length = max(len(str(cell.value)) for cell in col if cell.value);
                    worksheet.column_dimensions[col[0].column_letter].width = max_length + 2

        worksheet = writer.book.create_sheet("Executive Summary", 0);
        writer.sheets['Executive Summary'] = worksheet;
        current_row = 1
        summary_abc_count = final_results.groupby(['Storage location', 'ABC_Class']).size().unstack(fill_value=0)
        for c in ['A', 'B', 'C']:
            if c not in summary_abc_count: summary_abc_count[c] = 0
        summary_abc_count = summary_abc_count[['A', 'B', 'C']];
        summary_abc_count['Total'] = summary_abc_count.sum(axis=1);
        summary_abc_count.loc['Total'] = summary_abc_count.sum()
        worksheet.cell(row=current_row, column=1, value='‡∏™‡∏£‡∏∏‡∏õ‡∏à‡∏≥‡∏ô‡∏ß‡∏ô‡∏£‡∏≤‡∏¢‡∏Å‡∏≤‡∏£ A, B, C ‡πÉ‡∏ô‡πÅ‡∏ï‡πà‡∏•‡∏∞‡∏Ñ‡∏•‡∏±‡∏á').font = Font(bold=True);
        current_row += 1
        summary_abc_count.to_excel(writer, sheet_name='Executive Summary', startrow=current_row, startcol=0);
        current_row += summary_abc_count.shape[0] + 3
        worksheet.cell(row=current_row, column=1,
                       value='‡∏Å‡∏•‡∏∏‡πà‡∏°‡∏¢‡∏≤ (Drug Group) ‡∏ó‡∏µ‡πà‡∏°‡∏µ‡∏°‡∏π‡∏•‡∏Ñ‡πà‡∏≤‡∏Å‡∏≤‡∏£‡πÉ‡∏ä‡πâ‡∏á‡∏≤‡∏ô‡∏™‡∏π‡∏á‡∏™‡∏∏‡∏î 3 ‡∏≠‡∏±‡∏ô‡∏î‡∏±‡∏ö‡πÅ‡∏£‡∏Å (‡πÅ‡∏¢‡∏Å‡∏ï‡∏≤‡∏°‡∏Ñ‡∏•‡∏±‡∏á)').font = Font(
            bold=True);
        current_row += 1
        top_groups = final_results.groupby('Storage location').apply(
            lambda x: x.groupby('Drug group')['NetConsumptionValue'].sum().nlargest(3)).reset_index()
        top_groups['NetConsumptionValue'] = top_groups['NetConsumptionValue'].map('{:,.2f}'.format)
        top_groups.to_excel(writer, sheet_name='Executive Summary', startrow=current_row, startcol=0, index=False);
        current_row += top_groups.shape[0] + 3
        worksheet.cell(row=current_row, column=1,
                       value='‡∏£‡∏≤‡∏¢‡∏Å‡∏≤‡∏£‡∏¢‡∏≤‡∏ó‡∏µ‡πà‡∏°‡∏µ‡∏°‡∏π‡∏•‡∏Ñ‡πà‡∏≤‡∏Å‡∏≤‡∏£‡πÉ‡∏ä‡πâ‡∏á‡∏≤‡∏ô‡∏™‡∏π‡∏á‡∏™‡∏∏‡∏î 5 ‡∏≠‡∏±‡∏ô‡∏î‡∏±‡∏ö‡πÅ‡∏£‡∏Å (‡πÅ‡∏¢‡∏Å‡∏ï‡∏≤‡∏°‡∏Ñ‡∏•‡∏±‡∏á)').font = Font(bold=True);
        current_row += 1
        top_items = final_results.groupby('Storage location').apply(
            lambda x: x.groupby(['Material', 'Material description'])['NetConsumptionValue'].sum().nlargest(
                5)).reset_index()
        top_items['NetConsumptionValue'] = top_items['NetConsumptionValue'].map('{:,.2f}'.format)
        top_items.to_excel(writer, sheet_name='Executive Summary', startrow=current_row, startcol=0, index=False)
        for location in final_results['Storage location'].unique():
            sheet_df = final_results[final_results['Storage location'] == location].copy();
            sheet_name = f'SLoc_{location}'
            monthly_cols = sorted([col for col in sheet_df.columns if isinstance(col, str) and col.startswith('Qty_')])
            output_columns = ['Material', 'Material description', 'Storage location'] + monthly_cols + [
                'AvgMonthlyNetQuantity', 'NetConsumptionValue', 'PercentageValue', 'CumulativePercentage', 'ABC_Class',
                'Drug group']
            sheet_df = sheet_df[output_columns];
            sheet_df.to_excel(writer, sheet_name=sheet_name, index=False)
            apply_formats_and_hide_cols(writer, sheet_name, sheet_df)
    return output_buffer.getvalue()


# ==============================================================================
# STREAMLIT USER INTERFACE (UI)
# ==============================================================================

# --- Sidebar ---
# st.sidebar.image("path/to/your/logo.png", width=150) # <-- ‡πÉ‡∏™‡πà Path ‡∏£‡∏π‡∏õ‡πÇ‡∏•‡πÇ‡∏Å‡πâ‡∏Ç‡∏≠‡∏á‡∏Ñ‡∏∏‡∏ì‡∏ó‡∏µ‡πà‡∏ô‡∏µ‡πà
st.sidebar.title("CRA Analytics Suite")
st.sidebar.markdown("---")

app_mode = st.sidebar.selectbox(
    "‡πÇ‡∏õ‡∏£‡∏î‡πÄ‡∏•‡∏∑‡∏≠‡∏Å‡∏ü‡∏±‡∏á‡∏Å‡πå‡∏ä‡∏±‡∏ô‡∏ó‡∏µ‡πà‡∏ï‡πâ‡∏≠‡∏á‡∏Å‡∏≤‡∏£:",
    ["üè† ‡∏´‡∏ô‡πâ‡∏≤‡∏´‡∏•‡∏±‡∏Å", "üìä 1. ‡∏£‡∏≤‡∏¢‡∏á‡∏≤‡∏ô‡∏¢‡∏≤ ‡∏à2", "üìà 2. ‡∏£‡∏≤‡∏¢‡∏á‡∏≤‡∏ô‡∏Ç‡∏≤‡∏¢‡∏¢‡∏≤‡∏õ‡∏£‡∏∞‡∏à‡∏≥‡πÄ‡∏î‡∏∑‡∏≠‡∏ô", "üíâ 3. ‡∏£‡∏≤‡∏¢‡∏á‡∏≤‡∏ô‡∏¢‡∏≤ EPI", "üíä 4. ‡∏£‡∏≤‡∏¢‡∏á‡∏≤‡∏ô‡∏¢‡∏≤‡πÄ‡∏™‡∏û‡∏ï‡∏¥‡∏î‡∏Ø",
     "üéØ 5. ‡∏£‡∏≤‡∏¢‡∏á‡∏≤‡∏ô KPI", "üìÑ 6. ‡∏£‡∏ß‡∏°‡πÑ‡∏ü‡∏•‡πå PDF", "üî§ 7. ‡∏ß‡∏¥‡πÄ‡∏Ñ‡∏£‡∏≤‡∏∞‡∏´‡πå ABC"]
)

# --- Main Page ---
if app_mode == "üè† ‡∏´‡∏ô‡πâ‡∏≤‡∏´‡∏•‡∏±‡∏Å":
    st.title("üî¨ CRA Analytics Suite")
    st.markdown("‡∏¢‡∏¥‡∏ô‡∏î‡∏µ‡∏ï‡πâ‡∏≠‡∏ô‡∏£‡∏±‡∏ö‡∏™‡∏π‡πà‡πÄ‡∏Ñ‡∏£‡∏∑‡πà‡∏≠‡∏á‡∏°‡∏∑‡∏≠‡∏ß‡∏¥‡πÄ‡∏Ñ‡∏£‡∏≤‡∏∞‡∏´‡πå‡πÅ‡∏•‡∏∞‡∏õ‡∏£‡∏∞‡∏°‡∏ß‡∏•‡∏ú‡∏•‡∏Ç‡πâ‡∏≠‡∏°‡∏π‡∏•‡∏™‡∏≥‡∏´‡∏£‡∏±‡∏ö‡∏á‡∏≤‡∏ô‡πÄ‡∏†‡∏™‡∏±‡∏ä‡∏Å‡∏£‡∏£‡∏°")
    st.markdown("---")

    st.subheader("‡πÄ‡∏Ñ‡∏£‡∏∑‡πà‡∏≠‡∏á‡∏°‡∏∑‡∏≠‡∏ó‡∏±‡πâ‡∏á‡∏´‡∏°‡∏î")

    col1, col2, col3 = st.columns(3)
    with col1:
        st.info("üìä **‡∏£‡∏≤‡∏¢‡∏á‡∏≤‡∏ô‡∏¢‡∏≤ ‡∏à2**")
        st.write("‡∏£‡∏ß‡∏°‡πÑ‡∏ü‡∏•‡πå‡πÅ‡∏•‡∏∞‡∏Å‡∏£‡∏≠‡∏á‡∏¢‡∏≤‡∏ï‡∏≤‡∏°‡∏£‡∏≤‡∏¢‡∏Å‡∏≤‡∏£‡∏ó‡∏µ‡πà‡∏Å‡∏≥‡∏´‡∏ô‡∏î (J2)")

        st.info("üíä **‡∏£‡∏≤‡∏¢‡∏á‡∏≤‡∏ô‡∏¢‡∏≤‡πÄ‡∏™‡∏û‡∏ï‡∏¥‡∏î‡∏Ø**")
        st.write("‡∏™‡∏£‡πâ‡∏≤‡∏á‡∏£‡∏≤‡∏¢‡∏á‡∏≤‡∏ô‡∏¢‡∏≤‡πÄ‡∏™‡∏û‡∏ï‡∏¥‡∏î‡∏à‡∏≤‡∏Å‡∏Å‡∏≤‡∏£‡∏à‡πà‡∏≤‡∏¢, ‡∏Å‡∏≤‡∏£‡∏£‡∏±‡∏ö‡πÄ‡∏Ç‡πâ‡∏≤, ‡πÅ‡∏•‡∏∞‡πÑ‡∏ü‡∏•‡πå Master")

        st.info("üî§ **‡∏ß‡∏¥‡πÄ‡∏Ñ‡∏£‡∏≤‡∏∞‡∏´‡πå ABC**")
        st.write("‡∏à‡∏±‡∏î‡∏Å‡∏•‡∏∏‡πà‡∏°‡∏¢‡∏≤‡∏ï‡∏≤‡∏°‡∏°‡∏π‡∏•‡∏Ñ‡πà‡∏≤‡∏Å‡∏≤‡∏£‡πÉ‡∏ä‡πâ‡∏á‡∏≤‡∏ô‡πÅ‡∏•‡∏∞‡∏™‡∏£‡πâ‡∏≤‡∏á‡∏£‡∏≤‡∏¢‡∏á‡∏≤‡∏ô‡∏™‡∏£‡∏∏‡∏õ")

    with col2:
        st.info("üìà **‡∏£‡∏≤‡∏¢‡∏á‡∏≤‡∏ô‡∏Ç‡∏≤‡∏¢‡∏¢‡∏≤‡∏õ‡∏£‡∏∞‡∏à‡∏≥‡πÄ‡∏î‡∏∑‡∏≠‡∏ô**")
        st.write("‡∏ß‡∏¥‡πÄ‡∏Ñ‡∏£‡∏≤‡∏∞‡∏´‡πå‡∏Ç‡πâ‡∏≠‡∏°‡∏π‡∏•‡∏¢‡∏≤‡πÇ‡∏î‡∏¢‡∏•‡∏∞‡πÄ‡∏≠‡∏µ‡∏¢‡∏î‡∏û‡∏£‡πâ‡∏≠‡∏°‡πÑ‡∏ü‡∏•‡πå Master")

        st.info("üéØ **‡∏£‡∏≤‡∏¢‡∏á‡∏≤‡∏ô KPI**")
        st.write("‡∏Ñ‡∏≥‡∏ô‡∏ß‡∏ì‡∏ß‡∏±‡∏ô‡∏™‡∏≥‡∏£‡∏≠‡∏á‡∏Ñ‡∏á‡∏Ñ‡∏•‡∏±‡∏á‡∏à‡∏≤‡∏Å‡∏Ç‡πâ‡∏≠‡∏°‡∏π‡∏•‡∏¢‡∏≠‡∏î‡∏Ç‡∏≤‡∏¢‡πÅ‡∏•‡∏∞‡∏¢‡∏≠‡∏î‡∏Ñ‡∏á‡∏Ñ‡∏•‡∏±‡∏á")

    with col3:
        st.info("üíâ **‡∏£‡∏≤‡∏¢‡∏á‡∏≤‡∏ô‡∏¢‡∏≤ EPI**")
        st.write("‡∏™‡∏£‡∏∏‡∏õ‡∏¢‡∏≠‡∏î‡∏Å‡∏≤‡∏£‡πÉ‡∏ä‡πâ‡∏¢‡∏≤‡∏ï‡∏≤‡∏°‡∏£‡∏≤‡∏¢‡∏Å‡∏≤‡∏£ EPI")

        st.info("üìÑ **‡∏£‡∏ß‡∏°‡πÑ‡∏ü‡∏•‡πå PDF**")
        st.write("‡∏£‡∏ß‡∏°‡πÑ‡∏ü‡∏•‡πå PDF ‡∏´‡∏•‡∏≤‡∏¢‡πÑ‡∏ü‡∏•‡πå‡πÉ‡∏´‡πâ‡πÄ‡∏õ‡πá‡∏ô‡πÑ‡∏ü‡∏•‡πå‡πÄ‡∏î‡∏µ‡∏¢‡∏ß")

elif "1. ‡∏£‡∏≤‡∏¢‡∏á‡∏≤‡∏ô‡∏¢‡∏≤ ‡∏à2" in app_mode:
    st.header("üìä 1. ‡∏£‡∏≤‡∏¢‡∏á‡∏≤‡∏ô‡∏¢‡∏≤ ‡∏à2")
    st.markdown("---")
    st.info("**‡∏Ç‡∏±‡πâ‡∏ô‡∏ï‡∏≠‡∏ô:** ‡∏≠‡∏±‡∏õ‡πÇ‡∏´‡∏•‡∏î‡πÑ‡∏ü‡∏•‡πå‡∏Ç‡πâ‡∏≠‡∏°‡∏π‡∏•‡∏î‡∏¥‡∏ö (.xls) ‡∏à‡∏≤‡∏Å‡∏ô‡∏±‡πâ‡∏ô‡∏Å‡∏î‡∏õ‡∏∏‡πà‡∏°‡∏õ‡∏£‡∏∞‡∏°‡∏ß‡∏•‡∏ú‡∏•")
    uploaded_files_j2 = st.file_uploader("‡∏≠‡∏±‡∏õ‡πÇ‡∏´‡∏•‡∏î‡πÑ‡∏ü‡∏•‡πå‡∏Ç‡πâ‡∏≠‡∏°‡∏π‡∏•‡∏î‡∏¥‡∏ö‡∏Ç‡∏≠‡∏á‡∏Ñ‡∏∏‡∏ì (*.xls)", type="xls", accept_multiple_files=True,
                                         key="j2_uploader")
    if st.button("üöÄ ‡∏õ‡∏£‡∏∞‡∏°‡∏ß‡∏•‡∏ú‡∏•‡∏£‡∏≤‡∏¢‡∏á‡∏≤‡∏ô ‡∏à2", key="j2_button", use_container_width=True):
        if uploaded_files_j2:
            with st.spinner("‡∏Å‡∏≥‡∏•‡∏±‡∏á‡∏õ‡∏£‡∏∞‡∏°‡∏ß‡∏•‡∏ú‡∏•..."):
                final_df = process_j2_report(uploaded_files_j2)
            if final_df is not None:
                st.success("‚úÖ ‡∏õ‡∏£‡∏∞‡∏°‡∏ß‡∏•‡∏ú‡∏•‡∏™‡∏≥‡πÄ‡∏£‡πá‡∏à!")
                st.dataframe(final_df)
                output_buffer = io.BytesIO()
                with pd.ExcelWriter(output_buffer, engine='xlsxwriter') as writer: final_df.to_excel(writer,
                                                                                                     sheet_name='Raw',
                                                                                                     index=False)
                st.download_button(label="üì• ‡∏î‡∏≤‡∏ß‡∏ô‡πå‡πÇ‡∏´‡∏•‡∏î‡πÑ‡∏ü‡∏•‡πå J2.xlsx", data=output_buffer.getvalue(), file_name="J2.xlsx",
                                   mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        else:
            st.warning("‚ö†Ô∏è ‡∏Å‡∏£‡∏∏‡∏ì‡∏≤‡∏≠‡∏±‡∏õ‡πÇ‡∏´‡∏•‡∏î‡πÑ‡∏ü‡∏•‡πå‡∏Ç‡πâ‡∏≠‡∏°‡∏π‡∏•")

elif "2. ‡∏£‡∏≤‡∏¢‡∏á‡∏≤‡∏ô‡∏Ç‡∏≤‡∏¢‡∏¢‡∏≤‡∏õ‡∏£‡∏∞‡∏à‡∏≥‡πÄ‡∏î‡∏∑‡∏≠‡∏ô" in app_mode:
    st.header("üìà 2. ‡∏£‡∏≤‡∏¢‡∏á‡∏≤‡∏ô‡∏Ç‡∏≤‡∏¢‡∏¢‡∏≤‡∏õ‡∏£‡∏∞‡∏à‡∏≥‡πÄ‡∏î‡∏∑‡∏≠‡∏ô")
    st.markdown("---")
    st.info("""
        **‡∏Ç‡∏±‡πâ‡∏ô‡∏ï‡∏≠‡∏ô‡∏Å‡∏≤‡∏£‡πÉ‡∏ä‡πâ‡∏á‡∏≤‡∏ô:**
        1. **‡∏≠‡∏±‡∏õ‡πÇ‡∏´‡∏•‡∏î‡πÑ‡∏ü‡∏•‡πå‡∏Ç‡πâ‡∏≠‡∏°‡∏π‡∏•‡∏î‡∏¥‡∏ö:** ‡πÄ‡∏•‡∏∑‡∏≠‡∏Å‡πÑ‡∏ü‡∏•‡πå Excel (.xls) ‡∏ó‡∏µ‡πà‡∏°‡∏µ‡∏Ç‡πâ‡∏≠‡∏°‡∏π‡∏•‡∏Å‡∏≤‡∏£‡πÄ‡∏ö‡∏¥‡∏Å‡∏à‡πà‡∏≤‡∏¢‡∏¢‡∏≤ (‡πÄ‡∏•‡∏∑‡∏≠‡∏Å‡∏´‡∏•‡∏≤‡∏¢‡πÑ‡∏ü‡∏•‡πå‡πÑ‡∏î‡πâ)
        2. **‡∏≠‡∏±‡∏õ‡πÇ‡∏´‡∏•‡∏î‡πÑ‡∏ü‡∏•‡πå Drug Master:** ‡πÄ‡∏•‡∏∑‡∏≠‡∏Å‡πÑ‡∏ü‡∏•‡πå Master ‡∏ó‡∏µ‡πà‡∏°‡∏µ‡∏Ç‡πâ‡∏≠‡∏°‡∏π‡∏•‡∏¢‡∏≤
        3. **‡∏Å‡∏î‡∏õ‡∏∏‡πà‡∏° '‡πÄ‡∏£‡∏¥‡πà‡∏°‡∏Å‡∏≤‡∏£‡∏ß‡∏¥‡πÄ‡∏Ñ‡∏£‡∏≤‡∏∞‡∏´‡πå'**
    """)
    col1, col2 = st.columns(2)
    with col1:
        uploaded_files_raw = st.file_uploader("1. ‡∏≠‡∏±‡∏õ‡πÇ‡∏´‡∏•‡∏î‡πÑ‡∏ü‡∏•‡πå‡∏Ç‡πâ‡∏≠‡∏°‡∏π‡∏•‡∏î‡∏¥‡∏ö (*.xls)", type="xls", accept_multiple_files=True,
                                              key="raw_uploader")
    with col2:
        master_file = st.file_uploader("2. ‡∏≠‡∏±‡∏õ‡πÇ‡∏´‡∏•‡∏î‡πÑ‡∏ü‡∏•‡πå Drug Master (*.xlsx)", type=["xlsx"], key="master_uploader")
    if st.button("üöÄ ‡πÄ‡∏£‡∏¥‡πà‡∏°‡∏Å‡∏≤‡∏£‡∏ß‡∏¥‡πÄ‡∏Ñ‡∏£‡∏≤‡∏∞‡∏´‡πå", key="analysis_button", use_container_width=True):
        if uploaded_files_raw and master_file:
            with st.spinner("‡∏Å‡∏≥‡∏•‡∏±‡∏á‡∏ß‡∏¥‡πÄ‡∏Ñ‡∏£‡∏≤‡∏∞‡∏´‡πå‡∏Ç‡πâ‡∏≠‡∏°‡∏π‡∏•..."):
                raw_df, output_dfs = process_drug_rate_analysis(uploaded_files_raw, master_file)
            if raw_df is not None:
                st.success("‚úÖ ‡∏ß‡∏¥‡πÄ‡∏Ñ‡∏£‡∏≤‡∏∞‡∏´‡πå‡∏™‡∏≥‡πÄ‡∏£‡πá‡∏à!")
                output_buffer = io.BytesIO()
                with pd.ExcelWriter(output_buffer, engine='xlsxwriter') as writer:
                    for sheet_name, df_to_save in output_dfs.items():
                        if sheet_name != "Summary_Data": df_to_save.to_excel(writer, sheet_name=sheet_name, index=False)
                    startrow = 0
                    for label, df_summary in output_dfs["Summary_Data"].items():
                        summary_pivot = df_summary.set_index('Month').T;
                        summary_pivot.index = [label]
                        summary_pivot.to_excel(writer, sheet_name='Summary', startrow=startrow);
                        startrow += summary_pivot.shape[0] + 2
                st.download_button(label="üì• ‡∏î‡∏≤‡∏ß‡∏ô‡πå‡πÇ‡∏´‡∏•‡∏î‡∏£‡∏≤‡∏¢‡∏á‡∏≤‡∏ô‡∏ß‡∏¥‡πÄ‡∏Ñ‡∏£‡∏≤‡∏∞‡∏´‡πå", data=output_buffer.getvalue(),
                                   file_name="Drugstore_Rate.xlsx",
                                   mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                st.subheader("üìä ‡∏î‡∏π‡∏ï‡∏±‡∏ß‡∏≠‡∏¢‡πà‡∏≤‡∏á‡∏ú‡∏•‡∏•‡∏±‡∏û‡∏ò‡πå")
                with st.expander("‡∏Ñ‡∏•‡∏¥‡∏Å‡πÄ‡∏û‡∏∑‡πà‡∏≠‡∏î‡∏π‡∏ï‡∏±‡∏ß‡∏≠‡∏¢‡πà‡∏≤‡∏á‡∏Ç‡πâ‡∏≠‡∏°‡∏π‡∏•"):
                    tab1, tab2, tab3 = st.tabs(["Rate by Month", "Cases per Month", "Raw Merged Data"])
                    with tab1: st.dataframe(output_dfs["Rate ‡πÅ‡∏¢‡∏Å‡πÄ‡∏î‡∏∑‡∏≠‡∏ô"])
                    with tab2: st.dataframe(output_dfs["‡∏à‡∏≥‡∏ô‡∏ß‡∏ô‡πÄ‡∏Ñ‡∏™‡∏ï‡πà‡∏≠‡πÄ‡∏î‡∏∑‡∏≠‡∏ô"])
                    with tab3: st.dataframe(raw_df)
        else:
            st.warning("‚ö†Ô∏è ‡∏Å‡∏£‡∏∏‡∏ì‡∏≤‡∏≠‡∏±‡∏õ‡πÇ‡∏´‡∏•‡∏î‡∏ó‡∏±‡πâ‡∏á‡πÑ‡∏ü‡∏•‡πå‡∏Ç‡πâ‡∏≠‡∏°‡∏π‡∏•‡∏î‡∏¥‡∏ö‡πÅ‡∏•‡∏∞‡πÑ‡∏ü‡∏•‡πå Drug Master")

elif "3. ‡∏£‡∏≤‡∏¢‡∏á‡∏≤‡∏ô‡∏¢‡∏≤ EPI" in app_mode:
    st.header("üíâ 3. ‡∏£‡∏≤‡∏¢‡∏á‡∏≤‡∏ô‡∏¢‡∏≤ EPI")
    st.markdown("---")
    st.info("**‡∏Ç‡∏±‡πâ‡∏ô‡∏ï‡∏≠‡∏ô:** ‡∏≠‡∏±‡∏õ‡πÇ‡∏´‡∏•‡∏î‡πÑ‡∏ü‡∏•‡πå‡∏Ç‡πâ‡∏≠‡∏°‡∏π‡∏•‡∏î‡∏¥‡∏ö (.xls) ‡∏à‡∏≤‡∏Å‡∏ô‡∏±‡πâ‡∏ô‡∏Å‡∏î‡∏õ‡∏∏‡πà‡∏°‡∏õ‡∏£‡∏∞‡∏°‡∏ß‡∏•‡∏ú‡∏•")
    uploaded_files_epi = st.file_uploader("‡∏≠‡∏±‡∏õ‡πÇ‡∏´‡∏•‡∏î‡πÑ‡∏ü‡∏•‡πå‡∏Ç‡πâ‡∏≠‡∏°‡∏π‡∏•‡∏î‡∏¥‡∏ö‡∏Ç‡∏≠‡∏á‡∏Ñ‡∏∏‡∏ì (*.xls)", type="xls", accept_multiple_files=True,
                                          key="epi_uploader")
    if st.button("üöÄ ‡∏õ‡∏£‡∏∞‡∏°‡∏ß‡∏•‡∏ú‡∏•‡∏£‡∏≤‡∏¢‡∏á‡∏≤‡∏ô EPI", key="epi_button", use_container_width=True):
        if uploaded_files_epi:
            with st.spinner("‡∏Å‡∏≥‡∏•‡∏±‡∏á‡∏õ‡∏£‡∏∞‡∏°‡∏ß‡∏•‡∏ú‡∏•..."):
                final_df = process_epi_usage(uploaded_files_epi)
            if final_df is not None:
                st.success("‚úÖ ‡∏õ‡∏£‡∏∞‡∏°‡∏ß‡∏•‡∏ú‡∏•‡∏™‡∏≥‡πÄ‡∏£‡πá‡∏à!");
                st.subheader("‡∏ï‡∏≤‡∏£‡∏≤‡∏á‡∏™‡∏£‡∏∏‡∏õ‡∏¢‡∏≠‡∏î‡πÉ‡∏ä‡πâ‡∏¢‡∏≤ EPI");
                st.dataframe(final_df)
                output_buffer = io.BytesIO()
                with pd.ExcelWriter(output_buffer, engine='xlsxwriter') as writer: final_df.to_excel(writer,
                                                                                                     sheet_name='Raw',
                                                                                                     index=False)
                st.download_button(label="üì• ‡∏î‡∏≤‡∏ß‡∏ô‡πå‡πÇ‡∏´‡∏•‡∏î‡πÑ‡∏ü‡∏•‡πå EPI usage.xlsx", data=output_buffer.getvalue(),
                                   file_name="EPI usage.xlsx",
                                   mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        else:
            st.warning("‚ö†Ô∏è ‡∏Å‡∏£‡∏∏‡∏ì‡∏≤‡∏≠‡∏±‡∏õ‡πÇ‡∏´‡∏•‡∏î‡πÑ‡∏ü‡∏•‡πå‡∏Ç‡πâ‡∏≠‡∏°‡∏π‡∏•")

elif "4. ‡∏£‡∏≤‡∏¢‡∏á‡∏≤‡∏ô‡∏¢‡∏≤‡πÄ‡∏™‡∏û‡∏ï‡∏¥‡∏î‡∏Ø" in app_mode:
    st.header("üíä 4. ‡∏£‡∏≤‡∏¢‡∏á‡∏≤‡∏ô‡∏¢‡∏≤‡πÄ‡∏™‡∏û‡∏ï‡∏¥‡∏î‡πÅ‡∏•‡∏∞‡∏ß‡∏±‡∏ï‡∏ñ‡∏∏‡∏≠‡∏≠‡∏Å‡∏§‡∏ó‡∏ò‡∏¥‡πå")
    st.markdown("---")
    st.info("""
        **‡∏Ç‡∏±‡πâ‡∏ô‡∏ï‡∏≠‡∏ô‡∏Å‡∏≤‡∏£‡πÉ‡∏ä‡πâ‡∏á‡∏≤‡∏ô:**
        1. **‡∏≠‡∏±‡∏õ‡πÇ‡∏´‡∏•‡∏î‡πÑ‡∏ü‡∏•‡πå‡∏Ç‡πâ‡∏≠‡∏°‡∏π‡∏•‡∏Å‡∏≤‡∏£‡∏à‡πà‡∏≤‡∏¢‡∏¢‡∏≤:** ‡πÄ‡∏•‡∏∑‡∏≠‡∏Å‡πÑ‡∏ü‡∏•‡πå Excel (.xls) ‡∏ó‡∏µ‡πà‡∏°‡∏µ‡∏Ç‡πâ‡∏≠‡∏°‡∏π‡∏•‡∏Å‡∏≤‡∏£‡∏à‡πà‡∏≤‡∏¢‡∏¢‡∏≤ (‡πÄ‡∏•‡∏∑‡∏≠‡∏Å‡∏´‡∏•‡∏≤‡∏¢‡πÑ‡∏ü‡∏•‡πå‡πÑ‡∏î‡πâ)
        2. **‡∏≠‡∏±‡∏õ‡πÇ‡∏´‡∏•‡∏î‡πÑ‡∏ü‡∏•‡πå‡∏£‡∏≤‡∏¢‡∏á‡∏≤‡∏ô‡∏£‡∏±‡∏ö‡πÄ‡∏Ç‡πâ‡∏≤:** ‡πÄ‡∏•‡∏∑‡∏≠‡∏Å‡πÑ‡∏ü‡∏•‡πå Excel (.xlsx) ‡∏ó‡∏µ‡πà‡πÄ‡∏õ‡πá‡∏ô‡∏£‡∏≤‡∏¢‡∏á‡∏≤‡∏ô‡∏Å‡∏≤‡∏£‡∏£‡∏±‡∏ö‡πÄ‡∏Ç‡πâ‡∏≤
        3. **‡∏≠‡∏±‡∏õ‡πÇ‡∏´‡∏•‡∏î‡πÑ‡∏ü‡∏•‡πå Drug Master:** ‡πÄ‡∏•‡∏∑‡∏≠‡∏Å‡πÑ‡∏ü‡∏•‡πå Master ‡∏ó‡∏µ‡πà‡∏°‡∏µ‡∏Ç‡πâ‡∏≠‡∏°‡∏π‡∏•‡∏¢‡∏≤
        4. **‡∏Å‡∏î‡∏õ‡∏∏‡πà‡∏° '‡∏õ‡∏£‡∏∞‡∏°‡∏ß‡∏•‡∏ú‡∏•‡∏£‡∏≤‡∏¢‡∏á‡∏≤‡∏ô‡∏¢‡∏≤‡πÄ‡∏™‡∏û‡∏ï‡∏¥‡∏î'**
    """)
    col1, col2, col3 = st.columns(3)
    with col1:
        xls_files = st.file_uploader("1. ‡∏≠‡∏±‡∏õ‡πÇ‡∏´‡∏•‡∏î‡πÑ‡∏ü‡∏•‡πå‡∏Ç‡πâ‡∏≠‡∏°‡∏π‡∏•‡∏Å‡∏≤‡∏£‡∏à‡πà‡∏≤‡∏¢‡∏¢‡∏≤ (*.xls)", type="xls", accept_multiple_files=True,
                                     key="narcotics_xls_uploader")
    with col2:
        receipt_file = st.file_uploader("2. ‡∏≠‡∏±‡∏õ‡πÇ‡∏´‡∏•‡∏î‡πÑ‡∏ü‡∏•‡πå‡∏£‡∏≤‡∏¢‡∏á‡∏≤‡∏ô‡∏£‡∏±‡∏ö‡πÄ‡∏Ç‡πâ‡∏≤ (*.xlsx)", type="xlsx",
                                        key="narcotics_receipt_uploader")
    with col3:
        master_file_narcotics = st.file_uploader("3. ‡∏≠‡∏±‡∏õ‡πÇ‡∏´‡∏•‡∏î‡πÑ‡∏ü‡∏•‡πå Drug Master (*.xlsx)", type="xlsx",
                                                 key="narcotics_master_uploader")
    if st.button("üöÄ ‡∏õ‡∏£‡∏∞‡∏°‡∏ß‡∏•‡∏ú‡∏•‡∏£‡∏≤‡∏¢‡∏á‡∏≤‡∏ô‡∏¢‡∏≤‡πÄ‡∏™‡∏û‡∏ï‡∏¥‡∏î", key="narcotics_button", use_container_width=True):
        if xls_files and receipt_file and master_file_narcotics:
            with st.spinner("‡∏Å‡∏≥‡∏•‡∏±‡∏á‡∏õ‡∏£‡∏∞‡∏°‡∏ß‡∏•‡∏ú‡∏•..."):
                output_data = process_narcotics_report(xls_files, receipt_file, master_file_narcotics)
            if output_data:
                st.success("‚úÖ ‡∏õ‡∏£‡∏∞‡∏°‡∏ß‡∏•‡∏ú‡∏•‡∏™‡∏≥‡πÄ‡∏£‡πá‡∏à!")
                output_buffer = io.BytesIO()
                with pd.ExcelWriter(output_buffer, engine='xlsxwriter') as writer:
                    for sheet_name, df in output_data.items(): df.to_excel(writer, sheet_name=sheet_name, index=False)
                st.download_button(label="üì• ‡∏î‡∏≤‡∏ß‡∏ô‡πå‡πÇ‡∏´‡∏•‡∏î‡∏£‡∏≤‡∏¢‡∏á‡∏≤‡∏ô‡∏¢‡∏≤‡πÄ‡∏™‡∏û‡∏ï‡∏¥‡∏î.xlsx", data=output_buffer.getvalue(),
                                   file_name="‡∏£‡∏≤‡∏¢‡∏á‡∏≤‡∏ô‡∏Å‡∏≤‡∏£‡∏£‡∏±‡∏ö‡πÄ‡∏Ç‡πâ‡∏≤‡πÅ‡∏•‡∏∞‡∏à‡πà‡∏≤‡∏¢.xlsx",
                                   mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                with st.expander("‡∏Ñ‡∏•‡∏¥‡∏Å‡πÄ‡∏û‡∏∑‡πà‡∏≠‡∏î‡∏π‡∏ï‡∏±‡∏ß‡∏≠‡∏¢‡πà‡∏≤‡∏á‡∏Ç‡πâ‡∏≠‡∏°‡∏π‡∏•"):
                    tab1, tab2, tab3 = st.tabs(output_data.keys())
                    with tab1: st.dataframe(output_data['‡∏£‡∏≤‡∏¢‡∏á‡∏≤‡∏ô‡πÅ‡∏¢‡∏Å']);
                    with tab2: st.dataframe(output_data['‡∏£‡∏≤‡∏¢‡∏á‡∏≤‡∏ô‡∏£‡∏ß‡∏°']);
                    with tab3: st.dataframe(output_data['‡∏£‡∏≤‡∏¢‡∏á‡∏≤‡∏ô‡∏£‡∏±‡∏ö‡πÄ‡∏Ç‡πâ‡∏≤'])
        else:
            st.warning("‚ö†Ô∏è ‡∏Å‡∏£‡∏∏‡∏ì‡∏≤‡∏≠‡∏±‡∏õ‡πÇ‡∏´‡∏•‡∏î‡πÑ‡∏ü‡∏•‡πå‡πÉ‡∏´‡πâ‡∏Ñ‡∏£‡∏ö‡∏ó‡∏±‡πâ‡∏á 3 ‡∏™‡πà‡∏ß‡∏ô")

elif "5. ‡∏£‡∏≤‡∏¢‡∏á‡∏≤‡∏ô KPI" in app_mode:
    st.header("üéØ 5. ‡∏£‡∏≤‡∏¢‡∏á‡∏≤‡∏ô KPI")
    st.markdown("---")
    st.info("""
        **‡∏Ç‡∏±‡πâ‡∏ô‡∏ï‡∏≠‡∏ô‡∏Å‡∏≤‡∏£‡πÉ‡∏ä‡πâ‡∏á‡∏≤‡∏ô:**
        1. **‡∏≠‡∏±‡∏õ‡πÇ‡∏´‡∏•‡∏î‡πÑ‡∏ü‡∏•‡πå Rate:** ‡πÄ‡∏•‡∏∑‡∏≠‡∏Å‡πÑ‡∏ü‡∏•‡πå Excel (.xls) ‡∏ó‡∏µ‡πà‡∏°‡∏µ‡∏Ç‡πâ‡∏≠‡∏°‡∏π‡∏•‡∏Å‡∏≤‡∏£‡πÄ‡∏ö‡∏¥‡∏Å‡∏à‡πà‡∏≤‡∏¢‡∏¢‡∏≤
        2. **‡∏≠‡∏±‡∏õ‡πÇ‡∏´‡∏•‡∏î‡πÑ‡∏ü‡∏•‡πå‡∏¢‡∏≠‡∏î‡∏Ñ‡∏á‡∏Ñ‡∏•‡∏±‡∏á:** ‡πÄ‡∏•‡∏∑‡∏≠‡∏Å‡πÑ‡∏ü‡∏•‡πå Excel (.xlsx) ‡∏ó‡∏µ‡πà‡∏°‡∏µ‡∏¢‡∏≠‡∏î‡∏Ñ‡∏á‡∏Ñ‡∏•‡∏±‡∏á‡∏™‡∏¥‡πâ‡∏ô‡πÄ‡∏î‡∏∑‡∏≠‡∏ô
        3. **‡∏≠‡∏±‡∏õ‡πÇ‡∏´‡∏•‡∏î‡πÑ‡∏ü‡∏•‡πå Drug Master:** ‡πÄ‡∏•‡∏∑‡∏≠‡∏Å‡πÑ‡∏ü‡∏•‡πå Master ‡∏ó‡∏µ‡πà‡∏°‡∏µ‡∏Ç‡πâ‡∏≠‡∏°‡∏π‡∏•‡∏¢‡∏≤
        4. **‡∏Å‡∏î‡∏õ‡∏∏‡πà‡∏° '‡∏õ‡∏£‡∏∞‡∏°‡∏ß‡∏•‡∏ú‡∏• KPI'**
    """)
    col1, col2, col3 = st.columns(3)
    with col1:
        rate_file = st.file_uploader("1. ‡∏≠‡∏±‡∏õ‡πÇ‡∏´‡∏•‡∏î‡πÑ‡∏ü‡∏•‡πå Rate (*.xls)", type="xls", key="kpi_rate_uploader")
    with col2:
        inventory_file = st.file_uploader("2. ‡∏≠‡∏±‡∏õ‡πÇ‡∏´‡∏•‡∏î‡πÑ‡∏ü‡∏•‡πå‡∏¢‡∏≠‡∏î‡∏Ñ‡∏á‡∏Ñ‡∏•‡∏±‡∏á‡∏™‡∏¥‡πâ‡∏ô‡πÄ‡∏î‡∏∑‡∏≠‡∏ô (*.xlsx)", type="xlsx",
                                          key="kpi_inventory_uploader")
    with col3:
        master_file_kpi = st.file_uploader("3. ‡∏≠‡∏±‡∏õ‡πÇ‡∏´‡∏•‡∏î‡πÑ‡∏ü‡∏•‡πå Drug Master (*.xlsx)", type="xlsx",
                                           key="kpi_master_uploader")
    if st.button("üöÄ ‡∏õ‡∏£‡∏∞‡∏°‡∏ß‡∏•‡∏ú‡∏• KPI", key="kpi_button", use_container_width=True):
        if rate_file and inventory_file and master_file_kpi:
            with st.spinner("‡∏Å‡∏≥‡∏•‡∏±‡∏á‡∏Ñ‡∏≥‡∏ô‡∏ß‡∏ì KPI..."):
                output_data = process_kpi_report(rate_file, inventory_file, master_file_kpi)
            if output_data:
                st.success("‚úÖ ‡∏õ‡∏£‡∏∞‡∏°‡∏ß‡∏•‡∏ú‡∏•‡∏™‡∏≥‡πÄ‡∏£‡πá‡∏à!")
                output_buffer = io.BytesIO()
                with pd.ExcelWriter(output_buffer, engine='xlsxwriter') as writer:
                    for sheet_name, df in output_data.items(): df.to_excel(writer, sheet_name=sheet_name, index=False)
                st.download_button(label="üì• ‡∏î‡∏≤‡∏ß‡∏ô‡πå‡πÇ‡∏´‡∏•‡∏î‡∏£‡∏≤‡∏¢‡∏á‡∏≤‡∏ô KPI.xlsx", data=output_buffer.getvalue(),
                                   file_name="KPI_Report.xlsx",
                                   mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                with st.expander("‡∏Ñ‡∏•‡∏¥‡∏Å‡πÄ‡∏û‡∏∑‡πà‡∏≠‡∏î‡∏π‡∏ï‡∏±‡∏ß‡∏≠‡∏¢‡πà‡∏≤‡∏á‡∏Ç‡πâ‡∏≠‡∏°‡∏π‡∏•"):
                    tab1, tab2, tab3 = st.tabs(output_data.keys())
                    with tab1: st.dataframe(output_data['‡∏¢‡∏≠‡∏î‡∏Ç‡∏≤‡∏¢-‡∏Ñ‡∏á‡∏Ñ‡∏•‡∏±‡∏á-‡∏™‡∏≥‡∏£‡∏≠‡∏á‡∏Ñ‡∏á‡∏Ñ‡∏•‡∏±‡∏á']);
                    with tab2: st.dataframe(output_data['‡∏¢‡∏≠‡∏î‡∏Ç‡∏≤‡∏¢']);
                    with tab3: st.dataframe(output_data['Raw'])
        else:
            st.warning("‚ö†Ô∏è ‡∏Å‡∏£‡∏∏‡∏ì‡∏≤‡∏≠‡∏±‡∏õ‡πÇ‡∏´‡∏•‡∏î‡πÑ‡∏ü‡∏•‡πå‡πÉ‡∏´‡πâ‡∏Ñ‡∏£‡∏ö‡∏ó‡∏±‡πâ‡∏á 3 ‡∏™‡πà‡∏ß‡∏ô")

elif "6. ‡∏£‡∏ß‡∏°‡πÑ‡∏ü‡∏•‡πå PDF" in app_mode:
    st.header("üìÑ 6. ‡∏£‡∏ß‡∏°‡πÑ‡∏ü‡∏•‡πå PDF")
    st.markdown("---")
    st.info("""
        **‡∏Ç‡∏±‡πâ‡∏ô‡∏ï‡∏≠‡∏ô‡∏Å‡∏≤‡∏£‡πÉ‡∏ä‡πâ‡∏á‡∏≤‡∏ô:**
        1. **‡πÄ‡∏•‡∏∑‡∏≠‡∏Å‡πÑ‡∏ü‡∏•‡πå PDF:** ‡πÄ‡∏•‡∏∑‡∏≠‡∏Å‡πÑ‡∏ü‡∏•‡πå PDF ‡∏ó‡∏µ‡πà‡∏ï‡πâ‡∏≠‡∏á‡∏Å‡∏≤‡∏£‡∏£‡∏ß‡∏° (‡πÄ‡∏•‡∏∑‡∏≠‡∏Å‡∏´‡∏•‡∏≤‡∏¢‡πÑ‡∏ü‡∏•‡πå‡πÑ‡∏î‡πâ)
        2. **‡∏ï‡∏±‡πâ‡∏á‡∏ä‡∏∑‡πà‡∏≠‡πÑ‡∏ü‡∏•‡πå‡∏ú‡∏•‡∏•‡∏±‡∏û‡∏ò‡πå:** ‡∏Å‡∏£‡∏≠‡∏Å‡∏ä‡∏∑‡πà‡∏≠‡πÑ‡∏ü‡∏•‡πå‡πÉ‡∏´‡∏°‡πà‡∏ó‡∏µ‡πà‡∏ï‡πâ‡∏≠‡∏á‡∏Å‡∏≤‡∏£ (‡πÑ‡∏°‡πà‡∏ï‡πâ‡∏≠‡∏á‡πÉ‡∏™‡πà‡∏ô‡∏≤‡∏°‡∏™‡∏Å‡∏∏‡∏• .pdf)
        3. **‡∏Å‡∏î‡∏õ‡∏∏‡πà‡∏° '‡∏£‡∏ß‡∏°‡πÑ‡∏ü‡∏•‡πå PDF'**
    """)
    uploaded_pdfs = st.file_uploader("1. ‡πÄ‡∏•‡∏∑‡∏≠‡∏Å‡πÑ‡∏ü‡∏•‡πå PDF ‡∏ó‡∏µ‡πà‡∏ï‡πâ‡∏≠‡∏á‡∏Å‡∏≤‡∏£‡∏£‡∏ß‡∏°", type="pdf", accept_multiple_files=True,
                                     key="pdf_uploader")
    output_filename = st.text_input("2. ‡∏ï‡∏±‡πâ‡∏á‡∏ä‡∏∑‡πà‡∏≠‡πÑ‡∏ü‡∏•‡πå‡∏ú‡∏•‡∏•‡∏±‡∏û‡∏ò‡πå", "merged_output", key="pdf_output_name")
    if st.button("üöÄ ‡∏£‡∏ß‡∏°‡πÑ‡∏ü‡∏•‡πå PDF", key="pdf_merge_button", use_container_width=True):
        if uploaded_pdfs and output_filename:
            with st.spinner("‡∏Å‡∏≥‡∏•‡∏±‡∏á‡∏£‡∏ß‡∏°‡πÑ‡∏ü‡∏•‡πå PDF..."):
                merger = PdfMerger();
                for pdf_file in uploaded_pdfs: merger.append(pdf_file)
                pdf_buffer = io.BytesIO();
                merger.write(pdf_buffer);
                merger.close()
                final_filename = f"{output_filename.strip()}.pdf"
                st.success(f"‚úÖ ‡∏£‡∏ß‡∏°‡πÑ‡∏ü‡∏•‡πå PDF ‡∏™‡∏≥‡πÄ‡∏£‡πá‡∏à!")
                st.download_button(label=f"üì• ‡∏î‡∏≤‡∏ß‡∏ô‡πå‡πÇ‡∏´‡∏•‡∏î‡πÑ‡∏ü‡∏•‡πå {final_filename}", data=pdf_buffer.getvalue(),
                                   file_name=final_filename, mime="application/pdf")
        else:
            st.warning("‚ö†Ô∏è ‡∏Å‡∏£‡∏∏‡∏ì‡∏≤‡∏≠‡∏±‡∏õ‡πÇ‡∏´‡∏•‡∏î‡πÑ‡∏ü‡∏•‡πå PDF ‡πÅ‡∏•‡∏∞‡∏ï‡∏±‡πâ‡∏á‡∏ä‡∏∑‡πà‡∏≠‡πÑ‡∏ü‡∏•‡πå‡∏ú‡∏•‡∏•‡∏±‡∏û‡∏ò‡πå")

elif "7. ‡∏ß‡∏¥‡πÄ‡∏Ñ‡∏£‡∏≤‡∏∞‡∏´‡πå ABC" in app_mode:
    st.header("üî§ 7. ‡∏ß‡∏¥‡πÄ‡∏Ñ‡∏£‡∏≤‡∏∞‡∏´‡πå ABC (ABC Analysis)")
    st.markdown("---")
    st.info("""
        **‡∏Ç‡∏±‡πâ‡∏ô‡∏ï‡∏≠‡∏ô‡∏Å‡∏≤‡∏£‡πÉ‡∏ä‡πâ‡∏á‡∏≤‡∏ô:**
        1. **‡∏≠‡∏±‡∏õ‡πÇ‡∏´‡∏•‡∏î‡πÑ‡∏ü‡∏•‡πå‡∏Ç‡πâ‡∏≠‡∏°‡∏π‡∏•‡∏Å‡∏≤‡∏£‡πÉ‡∏ä‡πâ‡∏á‡∏≤‡∏ô:** ‡πÄ‡∏•‡∏∑‡∏≠‡∏Å‡πÑ‡∏ü‡∏•‡πå Excel (.xls, .xlsx) ‡∏ó‡∏µ‡πà‡∏°‡∏µ‡∏Ç‡πâ‡∏≠‡∏°‡∏π‡∏•‡∏Å‡∏≤‡∏£‡πÄ‡∏ö‡∏¥‡∏Å‡∏à‡πà‡∏≤‡∏¢‡∏¢‡∏≤ (‡πÄ‡∏•‡∏∑‡∏≠‡∏Å‡∏´‡∏•‡∏≤‡∏¢‡πÑ‡∏ü‡∏•‡πå‡πÑ‡∏î‡πâ)
        2. **‡∏≠‡∏±‡∏õ‡πÇ‡∏´‡∏•‡∏î‡πÑ‡∏ü‡∏•‡πå Drug Master:** ‡πÄ‡∏•‡∏∑‡∏≠‡∏Å‡πÑ‡∏ü‡∏•‡πå Master ‡∏ó‡∏µ‡πà‡∏°‡∏µ‡∏Ç‡πâ‡∏≠‡∏°‡∏π‡∏• 'Drug group'
        3. **‡∏Å‡∏î‡∏õ‡∏∏‡πà‡∏° '‡πÄ‡∏£‡∏¥‡πà‡∏°‡∏Å‡∏≤‡∏£‡∏ß‡∏¥‡πÄ‡∏Ñ‡∏£‡∏≤‡∏∞‡∏´‡πå ABC'**
    """)
    col1, col2 = st.columns(2)
    with col1:
        inventory_files = st.file_uploader("1. ‡∏≠‡∏±‡∏õ‡πÇ‡∏´‡∏•‡∏î‡πÑ‡∏ü‡∏•‡πå‡∏Ç‡πâ‡∏≠‡∏°‡∏π‡∏•‡∏Å‡∏≤‡∏£‡πÉ‡∏ä‡πâ‡∏á‡∏≤‡∏ô", type=["xlsx", "xls"],
                                           accept_multiple_files=True, key="abc_inventory_uploader")
    with col2:
        master_file_abc = st.file_uploader("2. ‡∏≠‡∏±‡∏õ‡πÇ‡∏´‡∏•‡∏î‡πÑ‡∏ü‡∏•‡πå Drug Master (*.xlsx)", type="xlsx",
                                           key="abc_master_uploader")
    if st.button("üöÄ ‡πÄ‡∏£‡∏¥‡πà‡∏°‡∏Å‡∏≤‡∏£‡∏ß‡∏¥‡πÄ‡∏Ñ‡∏£‡∏≤‡∏∞‡∏´‡πå ABC", key="abc_button", use_container_width=True):
        if inventory_files and master_file_abc:
            with st.spinner("‡∏Å‡∏≥‡∏•‡∏±‡∏á‡∏ó‡∏≥‡∏Å‡∏≤‡∏£‡∏ß‡∏¥‡πÄ‡∏Ñ‡∏£‡∏≤‡∏∞‡∏´‡πå ABC... ‡∏Å‡∏£‡∏∞‡∏ö‡∏ß‡∏ô‡∏Å‡∏≤‡∏£‡∏ô‡∏µ‡πâ‡∏≠‡∏≤‡∏à‡πÉ‡∏ä‡πâ‡πÄ‡∏ß‡∏•‡∏≤‡∏™‡∏±‡∏Å‡∏Ñ‡∏£‡∏π‡πà"):
                report_bytes = process_abc_analysis(inventory_files, master_file_abc)
            if report_bytes:
                st.success("‚úÖ ‡∏Å‡∏≤‡∏£‡∏ß‡∏¥‡πÄ‡∏Ñ‡∏£‡∏≤‡∏∞‡∏´‡πå ABC ‡πÄ‡∏™‡∏£‡πá‡∏à‡∏™‡∏°‡∏ö‡∏π‡∏£‡∏ì‡πå‡πÅ‡∏•‡∏∞‡∏™‡∏£‡πâ‡∏≤‡∏á‡∏£‡∏≤‡∏¢‡∏á‡∏≤‡∏ô‡∏™‡∏≥‡πÄ‡∏£‡πá‡∏à!")
                st.download_button(
                    label="üì• ‡∏î‡∏≤‡∏ß‡∏ô‡πå‡πÇ‡∏´‡∏•‡∏î‡∏£‡∏≤‡∏¢‡∏á‡∏≤‡∏ô ABC Analysis",
                    data=report_bytes,
                    file_name="Consolidated_ABC_Report.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
        else:
            st.warning("‚ö†Ô∏è ‡∏Å‡∏£‡∏∏‡∏ì‡∏≤‡∏≠‡∏±‡∏õ‡πÇ‡∏´‡∏•‡∏î‡πÑ‡∏ü‡∏•‡πå‡πÉ‡∏´‡πâ‡∏Ñ‡∏£‡∏ö‡∏ó‡∏±‡πâ‡∏á 2 ‡∏™‡πà‡∏ß‡∏ô")