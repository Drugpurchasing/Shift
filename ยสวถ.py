import tkinter as tk
from tkinter import filedialog, messagebox, simpledialog
import pandas as pd
import os
from openpyxl.styles import PatternFill

# Create a tkinter root window (it won't be shown)
root = tk.Tk()
root.withdraw()  # Hide the root window

while True:
    report_type = simpledialog.askstring("เลือกประเภทรายงาน", "กรอก ยส หรือ วถ:")
    if report_type is None:
        raise SystemExit("ยกเลิกการทำงาน")

    report_type = report_type.strip()
    if report_type in ("ยส", "วถ"):
        break

    messagebox.showerror("ข้อมูลไม่ถูกต้อง", "กรุณากรอกเฉพาะ ยส หรือ วถ")

# Open a file dialog to select a folder
folder_path = filedialog.askdirectory(
    title="Select Folder Containing Excel Files"
)

file_path = filedialog.askopenfilename(filetypes=[('Excel Files', '*.xlsx')])
source_workbook = pd.ExcelFile(file_path)
sheet_name = 'Sheet1'

source_workbook2 = pd.ExcelFile(
    r"C:\Users\813703\OneDrive - Chulabhorn Royal Academy\Shared Documents\..Drug Master คลังยา\Drug master.xlsx")
dfmaster = source_workbook2.parse("Drug master")
# Get a list of Excel files in the selected folder
excel_files = [f for f in os.listdir(folder_path) if f.endswith('.xls')]

# Initialize an empty DataFrame to store the combined data
stacked_df = pd.DataFrame()
total_source_df = pd.DataFrame()


def parse_consumption(value):
    if pd.isna(value):
        return 0

    text = str(value).strip().replace(',', '')
    if text == '':
        return 0

    is_parentheses_negative = text.startswith('(') and text.endswith(')')
    if is_parentheses_negative:
        text = text[1:-1].strip()

    number = pd.to_numeric(text, errors='coerce')
    if pd.isna(number):
        return 0

    if is_parentheses_negative:
        return -abs(number)

    return number


def parse_report_date(value):
    if pd.isna(value):
        return pd.NaT

    text = ' '.join(str(value).strip().split())
    for date_format in ('%d %b %y', '%d/%m/%Y %I:%M:%S %p', '%d/%m/%Y %H:%M:%S', '%d/%m/%Y'):
        parsed_date = pd.to_datetime(text, format=date_format, errors='coerce')
        if not pd.isna(parsed_date):
            return parsed_date.normalize()

    parsed_date = pd.to_datetime(text, errors='coerce')
    if pd.isna(parsed_date):
        return pd.NaT

    return parsed_date.normalize()


def convert_date(date_str):
    if not pd.isna(date_str):

        date_obj = pd.to_datetime(date_str)

        month_mapping = {
            1: 'มกราคม', 2: 'กุมภาพันธ์', 3: 'มีนาคม',
            4: 'เมษายน', 5: 'พฤษภาคม', 6: 'มิถุนายน',
            7: 'กรกฎาคม', 8: 'สิงหาคม', 9: 'กันยายน',
            10: 'ตุลาคม', 11: 'พฤศจิกายน', 12: 'ธันวาคม'
        }

        # Format the date in Thai format
        thai_day = date_obj.strftime('%d')
        thai_month = month_mapping.get(date_obj.month, date_obj.month)
        thai_year = str(date_obj.year + 543)
        return f"{thai_day} {thai_month} {thai_year}"
    else:
        return ''


for excel_file in excel_files:
    # Construct the full path to the Excel file
    file_path = os.path.join(folder_path, excel_file)
    drug_name = os.path.splitext(excel_file)[0].strip()

    df = pd.read_excel(file_path, header=None, skiprows=4)
    df = df.iloc[:, :9]
    df.columns = ['Date', 'AN/VN', 'HN', 'Name', 'pres id', 'consumption', 'unit', 'price', 'Address']
    df = df.dropna(how='all')

    df['Date'] = df['Date'].apply(parse_report_date)
    df = df.dropna(subset=['Date'])
    df['HN'] = df['HN'].astype(str).str.strip().str.replace(r'\.0$', '', regex=True)
    df = df[df['HN'].str.fullmatch(r'\d{9,10}', na=False)].reset_index(drop=True)
    if df.empty:
        continue

    df['consumption'] = df['consumption'].apply(parse_consumption)
    df['Name'] = df['Name'].fillna('').astype(str).str.strip()
    df['Address'] = df['Address'].fillna('').astype(str).str.strip()
    df['unit'] = df['unit'].fillna('').astype(str).str.strip()

    df = df.groupby(['Date', 'HN'], as_index=False).agg(
        Name=('Name', 'first'),
        Address=('Address', 'first'),
        unit=('unit', 'first'),
        consumption=('consumption', 'sum')
    )
    df = df[df['consumption'] != 0].reset_index(drop=True)
    if df.empty:
        continue

    df['ชื่อยาเสพติดให้โทษประเภท 2'] = drug_name
    df['จ่ายไป'] = (df['Name'] + ' ' + df['Address']).str.strip()
    total_source_df = pd.concat([
        total_source_df,
        df[['ชื่อยาเสพติดให้โทษประเภท 2', 'HN', 'จ่ายไป', 'unit', 'consumption']]
    ], axis=0, ignore_index=True)
    report_df = pd.DataFrame({
        'วัน เดือน ปี': df['Date'].apply(convert_date),
        'ชื่อยาเสพติดให้โทษประเภท 2': df['ชื่อยาเสพติดให้โทษประเภท 2'],
        'รหัส': df['HN'],
        'จ่ายไป': df['จ่ายไป'],
        'รับจาก อย': 0,
        'หน่วย_รับจาก_อย': df['unit'],
        'รับ': df['consumption'].apply(lambda x: abs(x) if x < 0 else 0),
        'หน่วย_รับ': df['unit'],
        'จ่าย': df['consumption'].apply(lambda x: x if x > 0 else 0),
        'หน่วย_จ่าย': df['unit']
    })
    summary_unit = df['unit'].iloc[0] if not df.empty else ''
    summary_row = pd.DataFrame([{
        'วัน เดือน ปี': '',
        'ชื่อยาเสพติดให้โทษประเภท 2': drug_name,
        'รหัส': '',
        'จ่ายไป': 'รวมทั้งสิ้น',
        'รับจาก อย': 0,
        'หน่วย_รับจาก_อย': summary_unit,
        'รับ': report_df['รับ'].sum(),
        'หน่วย_รับ': summary_unit,
        'จ่าย': report_df['จ่าย'].sum(),
        'หน่วย_จ่าย': summary_unit
    }])
    report_df = pd.concat([report_df, summary_row], axis=0, ignore_index=True)
    report_df.columns = [
        'วัน เดือน ปี',
        'ชื่อยาเสพติดให้โทษประเภท 2',
        'รหัส',
        'จ่ายไป',
        'รับจาก อย',
        'หน่วย',
        'รับ',
        'หน่วย',
        'จ่าย',
        'หน่วย'
    ]
    stacked_df = pd.concat([stacked_df, report_df], axis=0, ignore_index=True)

dfT = pd.read_excel(source_workbook, sheet_name)

dfmaster = dfmaster[["Material", "TradeName"]]

dfT = pd.merge(dfT, dfmaster, how="left")

dfT = dfT[["Posting Date", "TradeName", "Batch", 'Receiving stor. loc.', "Quantity"]]
dfT.columns = ['วัน เดือน ปี', "ชื่อยาเสพติดให้โทษประเภท 2", 'รหัส', 'จ่ายไป', 'รับจาก อย']


dfT['วัน เดือน ปี'] = dfT['วัน เดือน ปี'].apply(convert_date)

dfT.insert(5, 'หน่วย', '')
dfT.insert(6, 'รับ', '')
dfT.insert(7, 'จ่าย', '')

dfT = dfT[['วัน เดือน ปี', 'ชื่อยาเสพติดให้โทษประเภท 2', 'รหัส', 'จ่ายไป', 'รับจาก อย', 'หน่วย', 'รับ', 'หน่วย', 'จ่าย', 'หน่วย']]

output_file_name = "รายงานการรับเข้าและจ่าย ยส.xlsx" if report_type == "ยส" else "รายงานการรับเข้าและจ่าย.xlsx"
combined_file_path = os.path.join(folder_path, output_file_name)
if total_source_df.empty:
    Total_df = pd.DataFrame(columns=[
        'วัน เดือน ปี',
        'ชื่อยาเสพติดให้โทษประเภท 2',
        'รหัส',
        'จ่ายไป',
        'รับจาก อย',
        'หน่วย',
        'รับ',
        'หน่วย',
        'จ่าย',
        'หน่วย'
    ])
else:
    total_source_df = total_source_df.groupby(
        ['ชื่อยาเสพติดให้โทษประเภท 2', 'HN', 'จ่ายไป'], as_index=False
    ).agg(
        unit=('unit', 'first'),
        consumption=('consumption', 'sum'),
    )
    total_source_df = total_source_df[total_source_df['consumption'] != 0]
    Total_df = pd.DataFrame({
        'วัน เดือน ปี': 'รวม',
        'ชื่อยาเสพติดให้โทษประเภท 2': total_source_df['ชื่อยาเสพติดให้โทษประเภท 2'],
        'รหัส': total_source_df['HN'],
        'จ่ายไป': total_source_df['จ่ายไป'],
        'รับจาก อย': 0,
        'หน่วย_รับจาก_อย': total_source_df['unit'],
        'รับ': total_source_df['consumption'].apply(lambda x: abs(x) if x < 0 else 0),
        'หน่วย_รับ': total_source_df['unit'],
        'จ่าย': total_source_df['consumption'].apply(lambda x: x if x > 0 else 0),
        'หน่วย_จ่าย': total_source_df['unit']
    })
    Total_df.columns = [
        'วัน เดือน ปี',
        'ชื่อยาเสพติดให้โทษประเภท 2',
        'รหัส',
        'จ่ายไป',
        'รับจาก อย',
        'หน่วย',
        'รับ',
        'หน่วย',
        'จ่าย',
        'หน่วย'
    ]

with pd.ExcelWriter(combined_file_path, engine='openpyxl') as writer:
    stacked_df.to_excel(writer, index=False, sheet_name='รายงานแยก')
    Total_df.to_excel(writer, index=False, sheet_name='รายงานรวม')
    dfT.to_excel(writer, index=False, sheet_name='รายงานรับเข้า')

    light_grey_fill = PatternFill(fill_type='solid', fgColor='D9D9D9')
    worksheet = writer.sheets['รายงานแยก']
    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
        if row[3].value == 'รวมทั้งสิ้น':
            for cell in row:
                cell.fill = light_grey_fill
