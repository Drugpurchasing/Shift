# ==============================================================================
# CRA Pharmacy Shift Scheduler - Streamlit Single File App
# Run: streamlit run streamlit_scheduler_single_app.py
# Required packages: streamlit pandas numpy openpyxl tqdm
# ==============================================================================

import pandas as pd

import numpy as np
from datetime import datetime, timedelta, time
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter
import random
from statistics import stdev
# The 'drive' import is specific to Google Colab.
# If running locally, you might need to comment it out and adjust file paths.
try:
    from tqdm import tqdm
except Exception:
    def tqdm(iterable, **kwargs):
        return iterable
import io
import os

def download_google_sheet_as_xlsx(spreadsheet_id, output_path="/content/scheduler_input.xlsx"):
    """
    Download Google Sheet by spreadsheet_id as .xlsx for pandas/openpyxl workflow.
    เหมาะกับโค้ดเดิมที่ใช้ pd.read_excel(sheet_name=...)
    """
    print("Authenticating Google account...")
    auth.authenticate_user()

    drive_service = build("drive", "v3")

    request = drive_service.files().export_media(
        fileId=spreadsheet_id,
        mimeType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    with io.FileIO(output_path, "wb") as fh:
        downloader = MediaIoBaseDownload(fh, request)
        done = False
        while not done:
            status, done = downloader.next_chunk()
            if status:
                print(f"Download progress: {int(status.progress() * 100)}%")

    print(f"Downloaded Google Sheet as Excel: {output_path}")
    return output_path


SCHEDULE_SOURCES = {
    "1": {
        "label": "เภสัชกร",
        "spreadsheet_id": "1yQi8CBW0wonXWqHs2vvSxq8rz-Th1EWo0piUmir8yss",
        "employee_sheet_name": "employee",
        "output_prefix": "Pharmacist_Schedule",
        "gsheet_prefix": "Pharmacist Schedule",
    },
    "2": {
        "label": "ผู้ช่วยเภสัชกร",
        "spreadsheet_id": "1SU4iOfeEPzqJmuQrNHtsDk1ItYe-lEWUCLA4gaG0BNc",
        "employee_sheet_name": "employee",
        "output_prefix": "Assistant_Pharmacist_Schedule",
        "gsheet_prefix": "Assistant Pharmacist Schedule",
    },
}


def ask_int_input(prompt_text, default_value, min_value=None, max_value=None):
    while True:
        raw_value = input(f"{prompt_text} [{default_value}]: ").strip()
        if raw_value == "":
            return default_value
        try:
            value = int(raw_value)
            if min_value is not None and value < min_value:
                print(f"กรุณาใส่ค่าตั้งแต่ {min_value} ขึ้นไป")
                continue
            if max_value is not None and value > max_value:
                print(f"กรุณาใส่ค่าไม่เกิน {max_value}")
                continue
            return value
        except ValueError:
            print("กรุณาใส่เป็นตัวเลขจำนวนเต็ม")




def ask_yes_no_input(prompt_text, default_value="N"):
    """
    Ask a Y/N question.
    Default is N unless explicitly changed.
    Returns True for Y and False for N.
    """
    default_value = str(default_value).strip().upper()
    if default_value not in ["Y", "N"]:
        default_value = "N"

    while True:
        raw_value = input(f"{prompt_text} [Y/N, default {default_value}]: ").strip().upper()
        if raw_value == "":
            raw_value = default_value

        if raw_value in ["Y", "YES"]:
            return True
        if raw_value in ["N", "NO"]:
            return False

        print("กรุณาตอบเฉพาะ Y หรือ N")


def ask_schedule_source():
    print("\nเลือกชนิดของการรันข้อมูล")
    print("1 = ตารางเภสัชกร")
    print("2 = ตารางผู้ช่วยเภสัชกร")
    while True:
        selected = input("กรุณาเลือก 1 หรือ 2 [1]: ").strip() or "1"
        if selected in SCHEDULE_SOURCES:
            return SCHEDULE_SOURCES[selected]
        print("กรุณาเลือกเฉพาะ 1 หรือ 2")


class PharmacistScheduler:
    """
    Pharmacy shift scheduler with optimization and Excel export.
    Designed for multi-constraint pharmacist roster planning.
    """
    W_CONSECUTIVE = 10
    W_HOURS = 4
    W_PREFERENCE = 6
    MAX_CONSECUTIVE_DAYS = 3

    # Soft constraints for smoother monthly distribution and weekend-off protection.
    # ปรับเป็นคะแนน ไม่ block แข็ง เพื่อไม่ให้เกิด UNFILLED โดยไม่จำเป็น
    MIN_WEEKEND_OFF_DAYS = 4
    W_SHIFT_PACING = 900
    W_MONTH_SEGMENT_BALANCE = 180
    W_WEEKEND_OFF_PROTECTION = 2200

    def __init__(self, excel_file_path, employee_sheet_name='employee', staff_type='เภสัชกร'):
        self.pharmacists = {}
        self.shift_types = {}
        self.departments = {}
        self.pre_assignments = {}
        self.historical_scores = {}
        self.preference_multipliers = {}
        self.special_notes = {}
        self.shift_limits = {}
        self.employee_sheet_name = employee_sheet_name
        self.staff_type = staff_type
        self.employee_order = []
        self.no_preference_staff = set()
        self.excel_file_path = excel_file_path
        self.problem_days = set()
        self.run_logs = []
        self.run_config = {}
        self.soul_mates = {
            'ภก.ชานนท์ (บุ้ง)': 'ภญ.อาภาภัทร (มะปราง)'


        }

        self.read_data_from_excel(self.excel_file_path)
        self.load_historical_scores()
        self._calculate_preference_multipliers()

        self.night_shifts = {
            'I100-10', 'I100-12N', 'I400-12N', 'I400-10', 'O400ER-12N', 'O400ER-10'
        }
        self.holidays = {
            'specific_dates': ['2026-06-03','2026-06-01']
        }
        for pharmacist in self.pharmacists:
            self.pharmacists[pharmacist]['shift_counts'] = {
                shift_type: 0 for shift_type in self.shift_types
            }

    def _pre_check_staffing_levels(self, year, month):
        print("\nRunning pre-check for staffing levels (including all shifts + 3 buffer)...")
        start_date = datetime(year, month, 1)
        end_date = datetime(year, month + 1, 1) - timedelta(days=1) if month < 12 else datetime(year, 12, 31)
        dates = pd.date_range(start_date, end_date)

        all_ok = True
        for date in dates:
            available_pharmacists_count = sum(1 for p_name, p_info in self.pharmacists.items()
                                              if date.strftime('%Y-%m-%d') not in p_info['holidays'])
            required_shifts_base = sum(1 for st in self.shift_types
                                       if self.is_shift_available_on_date(st, date))
            total_required_shifts_with_buffer = required_shifts_base + 5

            if available_pharmacists_count < total_required_shifts_with_buffer:
                all_ok = False
                self.problem_days.add(date)
                print(f"WARNING: Potential shortage on {date.strftime('%Y-%m-%d')}. "
                      f"Available Pharmacists: {available_pharmacists_count}, "
                      f"Required Shifts (with +3 buffer): {total_required_shifts_with_buffer}")
        if all_ok:
            print("Pre-check complete. All days have sufficient staffing levels for the total workload.")
        else:
            print("Pre-check complete. Identified days with potential staff shortages. These will be prioritized.")
        return not all_ok


    def load_historical_scores(self):
        try:
            print("Attempting to load historical scores from sheet 'HistoricalScores'...")
            df = pd.read_excel(self.excel_file_path, sheet_name='HistoricalScores')
            if 'Pharmacist' in df.columns and 'Total Preference Score' in df.columns:
                for _, row in df.iterrows():
                    pharmacist = row['Pharmacist']
                    score = row['Total Preference Score']
                    if pharmacist in self.pharmacists:
                        self.historical_scores[pharmacist] = score
                print(f"Successfully loaded historical scores for {len(self.historical_scores)} pharmacists.")
            else:
                print("WARNING: 'HistoricalScores' sheet found, but required columns ('Pharmacist', 'Total Preference Score') are missing.")
        except ValueError:
            print("INFO: Sheet 'HistoricalScores' not found in the input file. Proceeding without historical data.")
        except Exception as e:
            print(f"An error occurred while loading historical scores: {e}")

    def _calculate_preference_multipliers(self):
        if not self.historical_scores:
            print("No historical scores found. All preference multipliers will be 1.0.")
            for pharmacist in self.pharmacists:
                self.preference_multipliers[pharmacist] = 1.0
            return
        min_score = min(self.historical_scores.values())
        max_score = max(self.historical_scores.values())
        if min_score == max_score:
            for pharmacist in self.pharmacists:
                self.preference_multipliers[pharmacist] = 1.0
            return
        for pharmacist, score in self.historical_scores.items():
            normalized_score = (score - min_score) / (max_score - min_score)
            min_multiplier = 0.7
            self.preference_multipliers[pharmacist] = min_multiplier + (1 - min_multiplier) * normalized_score
        for pharmacist in self.pharmacists:
            if pharmacist not in self.preference_multipliers:
                min_multiplier = 0.7
                self.preference_multipliers[pharmacist] = min_multiplier
                print(f"Pharmacist '{pharmacist}' not in historical data. Assigning a favorable multiplier of {min_multiplier}.")

    def _normalize_preference_value(self, value):
        if pd.isna(value):
            return None
        cleaned = str(value).strip()
        if cleaned == "" or cleaned.lower() == "none" or cleaned.lower() == "nan":
            return None
        return cleaned

    def _has_no_preferences(self, preferences):
        return all(self._normalize_preference_value(value) is None for value in preferences.values())

    def get_ordered_employees(self):
        ordered = [name for name in self.employee_order if name in self.pharmacists]
        missing = [name for name in self.pharmacists if name not in ordered]
        return ordered + missing

    def get_total_shift_count_in_schedule_dict(self, pharmacist, schedule_dict):
        count = 0
        for shifts in schedule_dict.values():
            for assigned in shifts.values():
                if assigned == pharmacist:
                    count += 1
        return count

    def get_unique_departments_worked(self, pharmacist, schedule_dict):
        departments = set()
        for shifts in schedule_dict.values():
            for shift_type, assigned in shifts.items():
                if assigned == pharmacist:
                    department = self.get_department_from_shift(shift_type)
                    if department:
                        departments.add(department)
        return departments

    def get_average_monthly_shift_target(self, year, month):
        """
        ประมาณจำนวนเวรเฉลี่ยต่อคนต่อเดือนจากจำนวนเวรที่เปิดจริงในเดือนนั้น
        ใช้เป็นฐาน pacing เพื่อไม่ให้คนเดิมถูกจัดหนักช่วงต้นเดือน/ปลายเดือน
        """
        cache_key = (year, month)
        if not hasattr(self, '_monthly_shift_target_cache'):
            self._monthly_shift_target_cache = {}
        if cache_key in self._monthly_shift_target_cache:
            return self._monthly_shift_target_cache[cache_key]

        start_date = datetime(year, month, 1)
        end_date = datetime(year + 1, 1, 1) - timedelta(days=1) if month == 12 else datetime(year, month + 1, 1) - timedelta(days=1)
        total_open_shifts = 0
        for date in pd.date_range(start_date, end_date):
            for shift_type in self.shift_types:
                if self.is_shift_available_on_date(shift_type, date):
                    total_open_shifts += 1

        staff_count = max(len(self.pharmacists), 1)
        target = max(total_open_shifts / staff_count, 1)
        self._monthly_shift_target_cache[cache_key] = target
        return target

    def get_month_segment(self, date):
        """แบ่งเดือนเป็น 3 ช่วง: ต้นเดือน / กลางเดือน / ปลายเดือน"""
        days_in_month = pd.Timestamp(date).days_in_month
        if date.day <= days_in_month / 3:
            return 'early'
        if date.day <= (2 * days_in_month) / 3:
            return 'middle'
        return 'late'

    def get_month_segment_shift_count(self, pharmacist, date, schedule_dict):
        """นับเวรของคนนี้ใน segment เดียวกับ date เพื่อกันการกระจุกในช่วงเดียวของเดือน"""
        target_segment = self.get_month_segment(date)
        count = 0
        for d, shifts in schedule_dict.items():
            if self.get_month_segment(d) != target_segment:
                continue
            if pharmacist in shifts.values():
                count += 1
        return count

    def get_total_weekend_days_in_schedule(self, schedule_dict):
        """นับจำนวนวันเสาร์-อาทิตย์ทั้งหมดในเดือนของ schedule_dict"""
        return sum(1 for d in schedule_dict if d.weekday() >= 5)

    def get_weekend_days_worked_until(self, pharmacist, schedule_dict):
        """นับจำนวนวันเสาร์-อาทิตย์ที่คนนี้ถูกจัดเวรแล้วใน schedule_dict ปัจจุบัน"""
        count = 0
        for d, shifts in schedule_dict.items():
            if d.weekday() >= 5 and pharmacist in shifts.values():
                count += 1
        return count

    def calculate_weekend_min_off_violations(self, schedule):
        """
        ตรวจสอบหลังจัดว่าใครมีวันหยุดเสาร์-อาทิตย์น้อยกว่า MIN_WEEKEND_OFF_DAYS
        คืนค่า dict: {pharmacist: shortfall}
        """
        weekend_dates = [d for d in schedule.index if d.weekday() >= 5]
        violations = {}
        for pharmacist in self.pharmacists:
            off_count = 0
            for d in weekend_dates:
                if pharmacist not in schedule.loc[d].values:
                    off_count += 1
            shortfall = max(0, self.MIN_WEEKEND_OFF_DAYS - off_count)
            if shortfall > 0:
                violations[pharmacist] = shortfall
        return violations

    def calculate_month_segment_variance(self, schedule):
        """
        วัดความกระจุกของเวรตามช่วงต้น/กลาง/ปลายเดือน
        ค่ายิ่งต่ำ = เวรของแต่ละคนกระจายตามเดือนดีกว่า
        """
        variances = []
        for pharmacist in self.pharmacists:
            counts = {'early': 0, 'middle': 0, 'late': 0}
            for d in schedule.index:
                if pharmacist in schedule.loc[d].values:
                    counts[self.get_month_segment(d)] += 1
            variances.append(np.var(list(counts.values())))
        return float(np.mean(variances)) if variances else 0

    def read_data_from_excel(self, file_path):
        # 1. โหลดข้อมูล Skill Group จากชีต 'Skill subset'
        skill_groups_map = {}
        try:
            print("Attempting to load skill subsets from sheet 'Skill subset'...")
            subset_df = pd.read_excel(file_path, sheet_name='Skill subset')

            # ตรวจสอบว่ามีคอลัมน์ที่ต้องการหรือไม่
            if 'Group Name' in subset_df.columns and 'Skills' in subset_df.columns:
                for _, row in subset_df.iterrows():
                    group_name = str(row['Group Name']).strip()
                    if group_name and group_name != 'nan':
                        # แยก Skill ด้วย comma และลบช่องว่างหน้าหลัง
                        skills_list = [s.strip() for s in str(row['Skills']).split(',') if s.strip() and s.strip() != 'nan']
                        skill_groups_map[group_name] = skills_list
                print(f"Successfully loaded {len(skill_groups_map)} skill groups.")
            else:
                print("WARNING: 'Skill subset' sheet found, but required columns ('Group Name', 'Skills') are missing.")
        except ValueError:
            print("INFO: Sheet 'Skill subset' not found. Proceeding without predefined skill groups.")
        except Exception as e:
            print(f"An error occurred while loading skill subsets: {e}")

        # 2. อ่านข้อมูลเจ้าหน้าที่จากชีต employee และ Map Skills
        pharmacists_df = pd.read_excel(file_path, sheet_name=self.employee_sheet_name)
        self.pharmacists = {}
        self.employee_order = []
        self.no_preference_staff = set()
        for _, row in pharmacists_df.iterrows():
            name = str(row['Name']).strip()
            if not name or name.lower() == 'nan':
                continue
            self.employee_order.append(name)
            max_hours = row.get('Max Hours', 250)
            if pd.isna(max_hours) or max_hours == '' or max_hours is None:
                max_hours = 250
            else:
                max_hours = float(max_hours)

            # ลอจิกการแตก Skill Group เป็น Skill ย่อย
            raw_skills = str(row['Skills']).split(',')
            expanded_skills = set()
            for s in raw_skills:
                s_clean = s.strip()
                # ถ้าเจอคำที่ตรงกับ Group Name ในชีต Skill subset ให้แตกเป็น List
                if s_clean in skill_groups_map:
                    expanded_skills.update(skill_groups_map[s_clean])
                elif s_clean and s_clean != 'nan':
                    expanded_skills.add(s_clean)

            preferences = {
                f'rank{i}': self._normalize_preference_value(row.get(f'Rank{i}'))
                for i in range(1, 9)
            }
            no_preference = self._has_no_preferences(preferences)
            if no_preference:
                self.no_preference_staff.add(name)

            self.pharmacists[name] = {
                'night_shift_count': 0,
                'skills': list(expanded_skills),
                'holidays': [date for date in str(row.get('Holidays', '')).split(',') if date != '1900-01-00' and date.strip() and date != 'nan'],
                'shift_counts': {},
                'preferences': preferences,
                'no_preference': no_preference,
                'max_hours': max_hours
            }

        # ... (โค้ดส่วนที่เหลือของฟังก์ชันตั้งแต่การอ่านชีต 'Shifts' ยังคงเหมือนเดิม) ...
        shifts_df = pd.read_excel(file_path, sheet_name='Shifts')
        # ...
        self.shift_types = {}
        for _, row in shifts_df.iterrows():
            shift_code = row['Shift Code']
            self.shift_types[shift_code] = {
                'description': row['Description'],
                'shift_type': row['Shift Type'],
                'start_time': row['Start Time'],
                'end_time': row['End Time'],
                'hours': row['Hours'],
                'required_skills': row['Required Skills'].split(','),
                'restricted_next_shifts': row['Restricted Next Shifts'].split(',') if pd.notna(row['Restricted Next Shifts']) else [],
            }
        departments_df = pd.read_excel(file_path, sheet_name='Departments')
        self.departments = {}
        for _, row in departments_df.iterrows():
            department = row['Department']
            self.departments[department] = row['Shift Codes'].split(',')
        pre_assign_df = pd.read_excel(file_path, sheet_name='PreAssignments')
        pre_assign_df['Date'] = pd.to_datetime(pre_assign_df['Date']).dt.strftime('%Y-%m-%d')
        self.pre_assignments = {}
        for pharmacist, group in pre_assign_df.groupby('Pharmacist'):
            date_dict = {}
            for date, g in group.groupby('Date'):
                shifts = []
                for shift_str in g['Shift']:
                    shifts.extend([s.strip() for s in str(shift_str).split(',') if s.strip()])
                date_dict[date] = shifts
            self.pre_assignments[pharmacist] = date_dict

        try:
            print("Attempting to load special notes from sheet 'SpecialNotes'...")
            notes_df = pd.read_excel(file_path, sheet_name='SpecialNotes', index_col=0)
            for pharmacist, row_data in notes_df.iterrows():
                if pharmacist in self.pharmacists:
                    for date_col, note in row_data.items():
                        if pd.notna(note) and str(note).strip():
                            date_str = pd.to_datetime(date_col).strftime('%Y-%m-%d')
                            if pharmacist not in self.special_notes:
                                self.special_notes[pharmacist] = {}
                            self.special_notes[pharmacist][date_str] = str(note).strip()
            print(f"Successfully loaded {sum(len(d) for d in self.special_notes.values())} special notes.")
        except ValueError:
            print("INFO: Sheet 'SpecialNotes' not found. Proceeding without special notes.")
        except Exception as e:
            print(f"An error occurred while loading special notes: {e}")

        try:
            print("Attempting to load shift limits from sheet 'ShiftLimits'...")
            limits_df = pd.read_excel(file_path, sheet_name='ShiftLimits')
            for _, row in limits_df.iterrows():
                pharmacist = row['Pharmacist']
                category = row['ShiftCategory']
                max_count = row['MaxCount']
                if pharmacist in self.pharmacists:
                    if pharmacist not in self.shift_limits:
                        self.shift_limits[pharmacist] = {}
                    self.shift_limits[pharmacist][category] = int(max_count)
            print(f"Successfully loaded {len(limits_df)} shift limit rules.")
        except ValueError:
            print("INFO: Sheet 'ShiftLimits' not found. Proceeding without shift limits.")
        except Exception as e:
            print(f"An error occurred while loading shift limits: {e}")
        # --- MIN SHIFT REQUIREMENTS ---
        self.min_shift_requirements = {}  # {pharmacist: {department: min_count}}
        try:
            print("Attempting to load minimum shift requirements from sheet 'MinShiftRequirements'...")
            min_req_df = pd.read_excel(file_path, sheet_name='MinShiftRequirements')
            for _, row in min_req_df.iterrows():
                pharmacist = str(row['Pharmacist']).strip()
                department = str(row['Department']).strip()
                min_count  = int(row['MinCount'])
                if pharmacist in self.pharmacists:
                    if pharmacist not in self.min_shift_requirements:
                        self.min_shift_requirements[pharmacist] = {}
                    self.min_shift_requirements[pharmacist][department] = min_count
            print(f"Successfully loaded min shift requirements for {len(self.min_shift_requirements)} pharmacists.")
        except ValueError:
            print("INFO: Sheet 'MinShiftRequirements' not found. Proceeding without min shift requirements.")
            self.min_shift_requirements = {}
        except Exception as e:
            print(f"An error occurred while loading min shift requirements: {e}")
            self.min_shift_requirements = {}

    def convert_time_to_minutes(self, time_input):
        if isinstance(time_input, str):
            hours, minutes = map(int, time_input.split(':'))
        elif isinstance(time_input, time):
            hours, minutes = time_input.hour, time_input.minute
        else:
            raise ValueError("Invalid input type. Expected string (HH:MM) or datetime.time object.")
        return hours * 60 + minutes

    def check_time_overlap(self, start1, end1, start2, end2):
        start1_mins = self.convert_time_to_minutes(start1)
        end1_mins = self.convert_time_to_minutes(end1)
        start2_mins = self.convert_time_to_minutes(start2)
        end2_mins = self.convert_time_to_minutes(end2)
        if end1_mins < start1_mins: end1_mins += 24 * 60
        if end2_mins < start2_mins: end2_mins += 24 * 60
        return start1_mins < end2_mins and end1_mins > start2_mins

    def check_mixing_expert_ratio_optimized(self, schedule_dict, date, current_shift=None, current_pharm=None):
        mixing_shifts = [p for s, p in schedule_dict[date].items()
                         if s.startswith('C8') and p not in ['NO SHIFT', 'UNASSIGNED', 'UNFILLED']]
        if current_shift and current_shift.startswith('C8') and current_pharm:
            mixing_shifts.append(current_pharm)
        if not mixing_shifts: return True
        total_mixing = len(mixing_shifts)
        expert_count = sum(1 for pharm in mixing_shifts
                           if pharm in self.pharmacists and 'mixing_expert' in self.pharmacists[pharm]['skills'])
        return expert_count >= (2 * total_mixing / 3)

    def count_consecutive_shifts(self, pharmacist, date, schedule, max_days=6):
        count = 0
        current_date = date - timedelta(days=1)
        for _ in range(max_days):
            if current_date in schedule.index and pharmacist in schedule.loc[current_date].values:
                count += 1
                current_date -= timedelta(days=1)
            else:
                break
        return count

    def is_holiday(self, date):
        return date.strftime('%Y-%m-%d') in self.holidays['specific_dates']

    def get_dept_shift_count(self, pharmacist, department, schedule_dict):
        """
        นับจำนวนเวรที่ pharmacist ทำในแผนก department แล้วใน schedule_dict ปัจจุบัน
        """
        count = 0
        for date, shifts in schedule_dict.items():
            for shift_type, assigned in shifts.items():
                if assigned == pharmacist:
                    if self.get_department_from_shift(shift_type) == department:
                        count += 1
        return count


    def _needs_min_shift(self, pharmacist, shift_type, schedule_dict):
        """
        คืนค่า True ถ้า pharmacist ยังไม่ถึง MinCount ของ department นี้
        ใช้เป็น priority boost ใน scoring
        """
        dept = self.get_department_from_shift(shift_type)
        if not dept:
            return False
        min_req = self.min_shift_requirements.get(pharmacist, {}).get(dept, 0)
        if min_req == 0:
            return False
        current_count = self.get_dept_shift_count(pharmacist, dept, schedule_dict)
        return current_count < min_req


    def calculate_weekend_off_variance(self, schedule, year, month):
        weekend_off_counts = {p: 0 for p in self.pharmacists}
        start_date = datetime(year, month, 1)
        end_date = datetime(year, month + 1, 1) - timedelta(days=1) if month < 12 else datetime(year, 12, 31)
        for date in pd.date_range(start_date, end_date):
            if date.weekday() >= 5:
                working_on_weekend = {schedule.loc[date, shift] for shift in schedule.columns if schedule.loc[date, shift] not in ['NO SHIFT', 'UNFILLED', 'UNASSIGNED']}
                for p_name in self.pharmacists:
                    if p_name not in working_on_weekend:
                        weekend_off_counts[p_name] += 1
        if len(weekend_off_counts) > 1:
            return np.var(list(weekend_off_counts.values()))
        return 0

    def is_night_shift(self, shift_type):
        return shift_type in self.night_shifts

    def _get_holiday_blocks(self, year, month):
        """
        คำนวณ block วันหยุดยาวทั้งหมดในเดือน
        คืนค่า dict: {last_day (datetime): set of all days in block}
        """
        import calendar
        start_date = datetime(year, month, 1)
        if month == 12:
            end_date = datetime(year, 12, 31)
        else:
            end_date = datetime(year, month + 1, 1) - timedelta(days=1)

        specific_holidays = set(
            datetime.strptime(d, '%Y-%m-%d')
            for d in self.holidays['specific_dates']
        )

        # สร้าง set วันที่เป็น "วันหยุด" ทุกประเภท (เสาร์, อาทิตย์, specific)
        all_off_days = set()
        current = start_date - timedelta(days=7)  # buffer ก่อนเดือน
        scan_end = end_date + timedelta(days=7)    # buffer หลังเดือน
        d = current
        while d <= scan_end:
            if d.weekday() >= 5 or d in specific_holidays:
                all_off_days.add(d)
            d += timedelta(days=1)

        # จัดกลุ่มเป็น block ติดกัน
        sorted_days = sorted(all_off_days)
        blocks = []
        if not sorted_days:
            return {}

        current_block = [sorted_days[0]]
        for day in sorted_days[1:]:
            if (day - current_block[-1]).days == 1:
                current_block.append(day)
            else:
                blocks.append(current_block)
                current_block = [day]
        blocks.append(current_block)

        # คืนค่า {last_day: set(block)}
        result = {}
        for block in blocks:
            # block ต้องมีเสาร์หรืออาทิตย์อยู่ด้วย (ไม่ใช่แค่ specific holiday กลางสัปดาห์โดด ๆ)
            has_weekend = any(d.weekday() >= 5 for d in block)
            if has_weekend and len(block) >= 1:
                last_day = block[-1]
                result[last_day] = set(block)

        return result


    def is_shift_available_on_date(self, shift_type, date):
        shift_info = self.shift_types[shift_type]
        is_holiday_date = self.is_holiday(date)  # specific_dates เท่านั้น

        weekday_num = date.weekday()
        is_saturday = (weekday_num == 5)
        is_sunday   = (weekday_num == 6)
        is_mon_to_thu = (0 <= weekday_num <= 3)

        s_type = str(shift_info['shift_type']).strip().lower()

        if s_type == 'weekday':
            # จันทร์-ศุกร์ ที่ไม่ใช่ specific holiday
            return not (is_holiday_date or is_saturday or is_sunday)

        elif s_type == 'saturday':
            # เสาร์ปกติ ไม่ใช่ specific holiday
            return is_saturday and not is_holiday_date

        elif s_type == 'sat-holiday':
            # เสาร์ + ตรงกับ specific holiday
            return is_saturday or is_holiday_date

        elif s_type == 'sunday':
            # อาทิตย์ปกติ ไม่ใช่ specific holiday
            return is_sunday and not is_holiday_date

        elif s_type == 'weekend':
            # Saturday, Sunday, OR any specific_dates holiday (incl. weekday holidays)
            return is_saturday or is_sunday or is_holiday_date

        elif s_type == 'sun-holiday':
            # อาทิตย์ + ตรงกับ specific holiday
            return is_sunday and is_holiday_date

        elif s_type in ['mon-thu', 'วันจันทร์-พฤหัส', 'วันจันทร์-พฤหัสบดี']:
            return is_mon_to_thu and not is_holiday_date

        elif s_type == 'holiday':
            # specific holiday ที่ไม่ใช่เสาร์/อาทิตย์ (weekday holiday)
            return is_holiday_date and not is_saturday and not is_sunday

        elif s_type == 'last-day-holiday':
            # วันสุดท้ายของ block หยุดยาวที่มีเสาร์/อาทิตย์รวมอยู่
            year  = date.year
            month = date.month
            # scan block รอบเดือนนี้ + buffer
            holiday_blocks = self._get_holiday_blocks(year, month)
            return date in holiday_blocks

        elif s_type == 'night':
            return True

        return False

    def get_department_from_shift(self, shift_type):
        if shift_type.startswith('I100'): return 'IPD100'
        elif shift_type.startswith('O100'): return 'OPD100'
        elif shift_type.startswith('Care'): return 'Care'
        elif shift_type.startswith('C8'): return 'Mixing'
        elif shift_type.startswith('I400'): return 'IPD400'
        elif shift_type.startswith('O400F1'): return 'OPD400F1'
        elif shift_type.startswith('O400F2'): return 'OPD400F2'
        elif shift_type.startswith('O400ER'): return 'ER'
        elif shift_type.startswith('ARI'): return 'ARI'
        elif shift_type.startswith('Refill'): return 'Refill'
        return None

    def _get_shift_category(self, shift_type):
        if self.is_night_shift(shift_type):
            return 'Night'
        if shift_type.startswith('C8'):
            return 'Mixing'
        return None

    def get_night_shift_count(self, pharmacist):
        return self.pharmacists[pharmacist]['night_shift_count']

    def get_preference_score(self, pharmacist, shift_type):
        p_skills = [skill.strip().lower() for skill in self.pharmacists[pharmacist]['skills']]
        if 'junior' in p_skills:
            return 1  # ให้คะแนนความชอบเป็น 1 (ดีที่สุด) เสมอ เพื่อให้กระจายไปได้ทุกที่

        # คนที่ไม่ได้ระบุ Preference เลย หรือระบุเป็น none/None ทุกช่อง
        # ให้ใช้คะแนนกลาง ไม่กดไปท้ายคิว และให้ logic department balance เป็นตัวกระจายงานแทน
        if self.pharmacists[pharmacist].get('no_preference', False):
            return 5

        department = self.get_department_from_shift(shift_type)
        for rank in range(1, 9):
            if self.pharmacists[pharmacist]['preferences'].get(f'rank{rank}') == department:
                return rank
        return 9

    def has_restricted_sequence_optimized(self, pharmacist, date, shift_type, schedule_dict):
        previous_date = date - timedelta(days=1)
        if previous_date in schedule_dict:
            for prev_shift, assigned_pharm in schedule_dict[previous_date].items():
                if assigned_pharm == pharmacist:
                    restricted = self.shift_types[prev_shift].get('restricted_next_shifts', [])
                    if shift_type in restricted: return True
        return False

    def has_overlapping_shift_optimized(self, pharmacist, date, new_shift_type, schedule_dict):
        if date not in schedule_dict: return False
        new_start = self.shift_types[new_shift_type]['start_time']
        new_end = self.shift_types[new_shift_type]['end_time']
        for existing_shift, assigned_pharm in schedule_dict[date].items():
            if assigned_pharm == pharmacist and existing_shift != new_shift_type:
                existing_start = self.shift_types[existing_shift]['start_time']
                existing_end = self.shift_types[existing_shift]['end_time']
                if self.check_time_overlap(new_start, new_end, existing_start, existing_end):
                    return True
        return False

    def has_nearby_night_shift_optimized(self, pharmacist, date, schedule_dict):
        for delta in [-2, -1, 1, 2]:
            check_date = date + timedelta(days=delta)
            if check_date in schedule_dict:
                for shift, assigned_pharm in schedule_dict[check_date].items():
                    if assigned_pharm == pharmacist and self.is_night_shift(shift):
                        return True
        return False

    def get_pharmacist_shifts(self, pharmacist, date, current_schedule):
        shifts = []
        if date in current_schedule.index:
            for shift_type in current_schedule.columns:
                if current_schedule.loc[date, shift_type] == pharmacist:
                    shifts.append(shift_type)
        return shifts

    def calculate_total_hours(self, pharmacist, schedule):
        total_hours = 0
        for date in schedule.index:
            for shift_type, assigned_pharm in schedule.loc[date].items():
                if assigned_pharm == pharmacist and shift_type in self.shift_types:
                    total_hours += self.shift_types[shift_type]['hours']
        return total_hours

    def _get_hour_imbalance_penalty(self, hours_dict):
        if not hours_dict or len(hours_dict) < 2:
            return 0
        hour_values = list(hours_dict.values())
        hour_stdev = stdev(hour_values)
        hour_range = max(hour_values) - min(hour_values)
        stdev_penalty = hour_stdev ** 2
        range_penalty = 0
        if hour_range > 10:
            range_penalty = (hour_range - 10) ** 2
        return stdev_penalty + range_penalty

    def calculate_schedule_metrics(self, schedule, year, month):
        hours = {p: self.calculate_total_hours(p, schedule) for p in self.pharmacists}
        night_counts = {p: self.pharmacists[p]['night_shift_count'] for p in self.pharmacists}
        weekend_off_var = self.calculate_weekend_off_variance(schedule, year, month)
        hour_penalty = self._get_hour_imbalance_penalty(hours)

        # --- ADDED: คำนวณความแปรปรวนของ Preference Score (%) ---
        pref_percentages = self.calculate_pharmacist_preference_scores(schedule)
        pref_variance = np.var(list(pref_percentages.values())) if len(pref_percentages) > 1 else 0

        weekend_min_off_violations = self.calculate_weekend_min_off_violations(schedule)
        metrics = {
            'hour_imbalance_penalty': hour_penalty,
            'night_variance': np.var(list(night_counts.values())) if night_counts else 0,
            'preference_score': sum(self.calculate_preference_penalty(p, schedule) for p in self.pharmacists),
            'preference_variance': pref_variance,
            'weekend_off_variance': weekend_off_var,
            'weekend_min_off_shortfall': sum(weekend_min_off_violations.values()),
            'month_segment_variance': self.calculate_month_segment_variance(schedule),
        }
        if len(hours) > 1:
            metrics['hour_diff_for_logging'] = stdev(hours.values())
        else:
            metrics['hour_diff_for_logging'] = 0
        return metrics

    def _log_schedule_event(self, event_type, message, **kwargs):
        """
        Store runtime logs in memory.
        These logs are exported only when enable_run_log=True.
        """
        log_row = {
            "Timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "Event Type": event_type,
            "Message": message,
        }
        for key, value in kwargs.items():
            log_row[key] = value
        self.run_logs.append(log_row)

    def _reset_runtime_shift_counters(self):
        """
        Reset counters used during schedule generation.
        """
        for pharmacist in self.pharmacists:
            self.pharmacists[pharmacist]['night_shift_count'] = 0
            self.pharmacists[pharmacist]['mixing_shift_count'] = 0
            self.pharmacists[pharmacist]['care_shift_count'] = 0
            self.pharmacists[pharmacist]['category_counts'] = {
                'Mixing': 0,
                'Night': 0
            }

    def _has_required_skills_for_shift(self, pharmacist, shift_type):
        """
        True when pharmacist has all required skills for shift_type.
        Empty Required Skills means everyone can be assigned.
        """
        p_skills = {
            str(skill).strip().lower()
            for skill in self.pharmacists.get(pharmacist, {}).get('skills', [])
            if str(skill).strip() and str(skill).strip().lower() != 'nan'
        }
        required_skills = {
            str(skill).strip().lower()
            for skill in self.shift_types.get(shift_type, {}).get('required_skills', [])
            if str(skill).strip() and str(skill).strip().lower() != 'nan'
        }
        return required_skills.issubset(p_skills)

    def _get_true_random_candidates(self, staff_pool, date, shift_type, schedule_dict):
        """
        Candidate filter for TRUE RANDOM OVERRIDE mode.

        This mode ignores scoring/fairness/preference/limits, but still keeps:
        1) no more than one shift per staff per day
        2) no assignment on personal holiday/leave
        3) required skill matching
        """
        date_str = pd.to_datetime(date).strftime('%Y-%m-%d')
        candidates = []

        for pharmacist in staff_pool:
            if pharmacist not in self.pharmacists:
                continue

            # 1) Do not assign multiple shifts in the same day.
            if pharmacist in schedule_dict[date].values():
                continue

            # 2) Do not assign staff on holiday/leave.
            if date_str in self.pharmacists[pharmacist].get('holidays', []):
                continue

            # 3) Required skill must match.
            if not self._has_required_skills_for_shift(pharmacist, shift_type):
                continue

            candidates.append(pharmacist)

        return candidates

    def _select_fair_random_candidate(self, candidates, pharmacist_hours, fairness_buffer_hours=8):
        """
        Select candidate using fairness-aware random logic.

        Logic:
        - Keep randomness, but do not random from the whole candidate pool.
        - First restrict pool to staff whose current total hours are close to the lowest current hours.
        - Then random inside that fair pool.

        fairness_buffer_hours=8 means:
        คนที่ชั่วโมงปัจจุบันไม่เกินคนชั่วโมงน้อยสุด + 8 ชั่วโมง จะมีสิทธิ์ถูกสุ่ม
        """
        if not candidates:
            return None, []

        min_hours = min(pharmacist_hours.get(p, 0) for p in candidates)
        fair_pool = [
            p for p in candidates
            if pharmacist_hours.get(p, 0) <= min_hours + fairness_buffer_hours
        ]

        if not fair_pool:
            fair_pool = candidates

        return random.choice(fair_pool), fair_pool

    def generate_monthly_schedule_true_random(self, year, month, iteration_num=1):
        """
        TRUE RANDOM OVERRIDE MODE - constrained + fairness-aware random.

        Randomly assigns open shifts while ignoring:
        - preference score
        - consecutive day limit
        - night spacing
        - restricted next shift rule
        - junior pairing rule
        - mixing ratio rule
        - shift category limits
        - historical score/multiplier
        - rescue/swap scoring

        Still enforced:
        - shift must be open on that date
        - PreAssignments are applied first and never overwritten
        - one staff cannot be assigned more than one shift in the same day
        - staff on holiday/leave cannot be assigned
        - staff must have all Required Skills for the shift
        - fairness / hour balance is preserved by randomizing only among lower-hour candidates
        """
        self._reset_runtime_shift_counters()

        start_date = datetime(year, month, 1)
        end_date = (
            datetime(year + 1, 1, 1) - timedelta(days=1)
            if month == 12
            else datetime(year, month + 1, 1) - timedelta(days=1)
        )
        dates = pd.date_range(start_date, end_date)

        schedule_dict = {
            date: {shift: 'NO SHIFT' for shift in self.shift_types}
            for date in dates
        }
        pharmacist_hours = {p: 0 for p in self.pharmacists}
        staff_pool = list(self.pharmacists.keys())

        unfilled_info = {
            'problem_days': [],
            'other_days': []
        }

        self._log_schedule_event(
            "TRUE_RANDOM_START",
            "Start constrained true random schedule generation",
            Year=year,
            Month=month,
            Iteration=iteration_num,
            StaffCount=len(staff_pool)
        )

        # 1) Apply locked PreAssignments first.
        for pharmacist, assignments in self.pre_assignments.items():
            if pharmacist not in self.pharmacists:
                self._log_schedule_event(
                    "PREASSIGN_SKIPPED",
                    "Preassigned staff not found in employee list",
                    Pharmacist=pharmacist
                )
                continue

            for date_str, shift_types in assignments.items():
                date = pd.to_datetime(date_str)
                if date not in schedule_dict:
                    continue

                for shift_type in shift_types:
                    if shift_type not in self.shift_types:
                        self._log_schedule_event(
                            "PREASSIGN_SKIPPED",
                            "Preassigned shift code not found in Shifts sheet",
                            Date=date.strftime("%Y-%m-%d"),
                            Shift=shift_type,
                            Pharmacist=pharmacist
                        )
                        continue

                    # Do not block PreAssignments even if they violate leave/skill/day rules.
                    # They are treated as manual locked assignments.
                    schedule_dict[date][shift_type] = pharmacist
                    self._update_shift_counts(pharmacist, shift_type)
                    pharmacist_hours[pharmacist] += self.shift_types[shift_type]['hours']

                    self._log_schedule_event(
                        "PREASSIGN_APPLIED",
                        "Applied locked preassignment",
                        Date=date.strftime("%Y-%m-%d"),
                        Shift=shift_type,
                        Pharmacist=pharmacist
                    )

        # 2) Randomly assign all open shifts with only hard safety filters.
        for date in tqdm(dates, desc=f"True Random Schedule (Iteration {iteration_num})", leave=False):
            shifts_to_process = list(self.shift_types.keys())
            random.shuffle(shifts_to_process)

            for shift_type in shifts_to_process:
                # Keep closed shifts as NO SHIFT.
                if not self.is_shift_available_on_date(shift_type, date):
                    schedule_dict[date][shift_type] = 'NO SHIFT'
                    continue

                # Do not overwrite PreAssignments or any existing assignment.
                if schedule_dict[date][shift_type] not in ['NO SHIFT', 'UNASSIGNED', 'UNFILLED']:
                    continue

                candidates = self._get_true_random_candidates(
                    staff_pool=staff_pool,
                    date=date,
                    shift_type=shift_type,
                    schedule_dict=schedule_dict
                )

                if not candidates:
                    schedule_dict[date][shift_type] = 'UNFILLED'
                    if date in self.problem_days:
                        unfilled_info['problem_days'].append((date, shift_type))
                    else:
                        unfilled_info['other_days'].append((date, shift_type))

                    self._log_schedule_event(
                        "UNFILLED_TRUE_RANDOM",
                        "No candidate after applying hard filters: one shift/day, not on leave, required skills",
                        Date=date.strftime("%Y-%m-%d"),
                        Shift=shift_type
                    )
                    continue

                chosen, fair_pool = self._select_fair_random_candidate(
                    candidates=candidates,
                    pharmacist_hours=pharmacist_hours,
                    fairness_buffer_hours=8
                )

                if chosen is None:
                    schedule_dict[date][shift_type] = 'UNFILLED'
                    if date in self.problem_days:
                        unfilled_info['problem_days'].append((date, shift_type))
                    else:
                        unfilled_info['other_days'].append((date, shift_type))
                    self._log_schedule_event(
                        "UNFILLED_TRUE_RANDOM",
                        "No chosen candidate after fairness filter",
                        Date=date.strftime("%Y-%m-%d"),
                        Shift=shift_type
                    )
                    continue

                before_hours = pharmacist_hours.get(chosen, 0)
                schedule_dict[date][shift_type] = chosen
                self._update_shift_counts(chosen, shift_type)
                pharmacist_hours[chosen] += self.shift_types[shift_type]['hours']

                self._log_schedule_event(
                    "TRUE_RANDOM_ASSIGN",
                    "Assigned by fairness-aware constrained true random override",
                    Date=date.strftime("%Y-%m-%d"),
                    Shift=shift_type,
                    Pharmacist=chosen,
                    CandidateCount=len(candidates),
                    FairPoolCount=len(fair_pool),
                    HoursBefore=before_hours,
                    ShiftHours=self.shift_types[shift_type]['hours'],
                    HoursAfter=pharmacist_hours[chosen]
                )

        final_schedule = pd.DataFrame.from_dict(schedule_dict, orient='index')
        final_schedule = final_schedule.reindex(columns=list(self.shift_types.keys()), fill_value='NO SHIFT')
        final_schedule.fillna('NO SHIFT', inplace=True)

        self._log_schedule_event(
            "TRUE_RANDOM_COMPLETE",
            "Completed constrained true random schedule generation",
            Year=year,
            Month=month,
            TotalUnfilled=len(unfilled_info['problem_days']) + len(unfilled_info['other_days'])
        )

        return final_schedule, unfilled_info

    def generate_monthly_schedule_shuffled(self, year, month, shuffled_shifts=None, shuffled_pharmacists=None, iteration_num=1):
        start_date = datetime(year, month, 1)
        end_date = datetime(year + 1, 1, 1) - timedelta(days=1) if month == 12 else datetime(year, month + 1, 1) - timedelta(days=1)
        dates = pd.date_range(start_date, end_date)
        schedule_dict = {date: {shift: 'NO SHIFT' for shift in self.shift_types} for date in dates}
        pharmacist_hours = {p: 0 for p in self.pharmacists}
        pharmacist_consecutive_days = {p: 0 for p in self.pharmacists}

        if shuffled_shifts is None:
            shuffled_shifts = list(self.shift_types.keys())
            random.shuffle(shuffled_shifts)
        if shuffled_pharmacists is None:
            shuffled_pharmacists = list(self.pharmacists.keys())
            random.shuffle(shuffled_pharmacists)

        for pharmacist in self.pharmacists:
            self.pharmacists[pharmacist]['night_shift_count'] = 0
            self.pharmacists[pharmacist]['mixing_shift_count'] = 0
            self.pharmacists[pharmacist]['care_shift_count'] = 0
            self.pharmacists[pharmacist]['category_counts'] = {
                'Mixing': 0,
                'Night': 0
            }

        for pharmacist, assignments in self.pre_assignments.items():
            if pharmacist not in self.pharmacists: continue
            for date_str, shift_types in assignments.items():
                date = pd.to_datetime(date_str)
                if date not in schedule_dict: continue
                for shift_type in shift_types:
                    if shift_type in self.shift_types:
                        schedule_dict[date][shift_type] = pharmacist
                        self._update_shift_counts(pharmacist, shift_type)
                        pharmacist_hours[pharmacist] += self.shift_types[shift_type]['hours']

        all_dates = list(pd.date_range(start_date, end_date))
        problem_dates_sorted = sorted([d for d in all_dates if d in self.problem_days])
        other_dates_sorted = sorted([d for d in all_dates if d not in self.problem_days])
        processing_order_dates = problem_dates_sorted + other_dates_sorted
        unfilled_info = {'problem_days': [], 'other_days': []}
        night_shifts_ordered = [s for s in shuffled_shifts if self.is_night_shift(s)]
        mixing_shifts_ordered = [s for s in shuffled_shifts if s.startswith('C8') and not self.is_night_shift(s)]
        care_shifts_ordered = [s for s in shuffled_shifts if s.startswith('Care') and not self.is_night_shift(s) and not s.startswith('C8')]
        other_shifts_ordered = [s for s in shuffled_shifts if not self.is_night_shift(s) and not s.startswith('C8') and not s.startswith('Care')]
        standard_shift_order = night_shifts_ordered + mixing_shifts_ordered + care_shifts_ordered + other_shifts_ordered
        problem_day_shift_order = mixing_shifts_ordered + care_shifts_ordered + night_shifts_ordered + other_shifts_ordered

        for date in tqdm(processing_order_dates, desc=f"Building Schedule (Iteration {iteration_num})", leave=False):
            pharmacists_working_yesterday = set()
            previous_date = date - timedelta(days=1)
            if previous_date in schedule_dict:
                   pharmacists_working_yesterday = {p for p in schedule_dict[previous_date].values() if p in self.pharmacists}
            for p_name in self.pharmacists:
                if p_name in pharmacists_working_yesterday:
                    pharmacist_consecutive_days[p_name] += 1
                else:
                    pharmacist_consecutive_days[p_name] = 0

            is_day_before_problem_day = (date + timedelta(days=1)) in self.problem_days

            shifts_to_process = problem_day_shift_order if date in self.problem_days else standard_shift_order
            for shift_type in shifts_to_process:
                if schedule_dict[date][shift_type] not in ['NO SHIFT', 'UNASSIGNED', 'UNFILLED'] or not self.is_shift_available_on_date(shift_type, date):
                    continue
                available = self._get_available_pharmacists_optimized(shuffled_pharmacists, date, shift_type, schedule_dict, pharmacist_hours, pharmacist_consecutive_days)
                if available:
                    chosen = self._select_best_pharmacist(available, shift_type, date, is_day_before_problem_day)
                    pharmacist_to_assign = chosen['name']
                    schedule_dict[date][shift_type] = pharmacist_to_assign
                    self._update_shift_counts(pharmacist_to_assign, shift_type)
                    pharmacist_hours[pharmacist_to_assign] += self.shift_types[shift_type]['hours']
                else:
                    # Rescue/SWAP step: สำหรับเวรที่ใช้ skill เฉพาะ เช่น Mixing/Care
                    # ถ้าคนมี skill ติดเวรอื่นอยู่ ให้ลองดึงคนนั้นมาลงเวรนี้
                    # แล้วหาคนอื่นที่เหมาะสมไปแทนเวรเดิมของเขา
                    rescued = self._try_rescue_assign_with_swap(
                        shuffled_pharmacists,
                        date,
                        shift_type,
                        schedule_dict,
                        pharmacist_hours,
                        pharmacist_consecutive_days,
                        is_day_before_problem_day
                    )

                    if not rescued:
                        schedule_dict[date][shift_type] = 'UNFILLED'
                        # บันทึกข้อมูลเวรที่ว่างไว้ แต่ปล่อยให้ลูปเดินหน้าต่อไปจนจบเดือน
                        if date in self.problem_days:
                            unfilled_info['problem_days'].append((date, shift_type))
                        else:
                            unfilled_info['other_days'].append((date, shift_type))

        # เมื่อจัดครบทุกวันแล้วค่อยสร้าง DataFrame
        final_schedule = pd.DataFrame.from_dict(schedule_dict, orient='index')
        final_schedule = final_schedule.reindex(columns=list(self.shift_types.keys()), fill_value='NO SHIFT')
        final_schedule.fillna('NO SHIFT', inplace=True)
        return final_schedule, unfilled_info

    def _update_shift_counts(self, pharmacist, shift_type):
        if self.is_night_shift(shift_type):
            self.pharmacists[pharmacist]['night_shift_count'] += 1
        if shift_type.startswith('C8'):
            self.pharmacists[pharmacist]['mixing_shift_count'] += 1
        if shift_type.startswith('Care'):
            self.pharmacists[pharmacist]['care_shift_count'] += 1
        category = self._get_shift_category(shift_type)
        if category and category in self.pharmacists[pharmacist]['category_counts']:
            self.pharmacists[pharmacist]['category_counts'][category] += 1

    def _revert_shift_counts(self, pharmacist, shift_type):
        """Undo shift counters when a pharmacist is moved out from an already assigned shift."""
        if pharmacist not in self.pharmacists or shift_type not in self.shift_types:
            return

        if self.is_night_shift(shift_type):
            self.pharmacists[pharmacist]['night_shift_count'] = max(
                0, self.pharmacists[pharmacist].get('night_shift_count', 0) - 1
            )
        if shift_type.startswith('C8'):
            self.pharmacists[pharmacist]['mixing_shift_count'] = max(
                0, self.pharmacists[pharmacist].get('mixing_shift_count', 0) - 1
            )
        if shift_type.startswith('Care'):
            self.pharmacists[pharmacist]['care_shift_count'] = max(
                0, self.pharmacists[pharmacist].get('care_shift_count', 0) - 1
            )

        category = self._get_shift_category(shift_type)
        if category and category in self.pharmacists[pharmacist].get('category_counts', {}):
            self.pharmacists[pharmacist]['category_counts'][category] = max(
                0, self.pharmacists[pharmacist]['category_counts'].get(category, 0) - 1
            )

    def _is_preassigned_shift(self, pharmacist, date, shift_type):
        """Return True if this exact assignment came from PreAssignments and should not be moved by rescue swap."""
        date_str = pd.to_datetime(date).strftime('%Y-%m-%d')
        return (
            pharmacist in self.pre_assignments
            and date_str in self.pre_assignments[pharmacist]
            and shift_type in self.pre_assignments[pharmacist][date_str]
        )

    def _is_skill_scarce_shift(self, shift_type):
        """Shifts with specific required skills should get rescue/swap priority."""
        required_skills = [
            str(s).strip().lower()
            for s in self.shift_types.get(shift_type, {}).get('required_skills', [])
            if str(s).strip()
        ]
        return bool(required_skills) or shift_type.startswith('C8') or shift_type.startswith('Care')

    def _can_replace_shift_after_temp_move(
        self,
        candidate,
        date,
        old_shift,
        schedule_dict,
        temp_hours,
        consecutive_days_dict
    ):
        """Check if candidate can cover old_shift after donor has been removed from it."""
        available = self._get_available_pharmacists_optimized(
            [candidate],
            date,
            old_shift,
            schedule_dict,
            temp_hours,
            consecutive_days_dict
        )
        return available[0] if available else None

    def _try_rescue_assign_with_swap(
        self,
        pharmacists,
        date,
        target_shift,
        schedule_dict,
        pharmacist_hours,
        pharmacist_consecutive_days,
        is_day_before_problem_day=False
    ):
        """
        Rescue logic for skill-scarce shifts.

        Use case:
        - target_shift เช่น C8/Mixing หรือ Care ต้องใช้ skill เฉพาะ
        - ไม่มีคนว่างเพราะคนที่มี skill ติดเวรอื่นในวันเดียวกัน
        - ระบบจะลองย้ายคนมี skill จากเวรเดิมมาลง target_shift
        - แล้วหาคนอื่นที่ทำเวรเดิมแทนได้

        This is intentionally conservative:
        - ไม่ย้ายเวรที่มาจาก PreAssignments
        - ไม่ย้ายจากเวร night
        - ไม่ย้ายจากเวร skill-scarce อีกเวรหนึ่งถ้ามีทางเลือกอื่น
        - ใช้ availability checker เดิมเพื่อลดการกระทบ constraint อื่น
        """
        if not self._is_skill_scarce_shift(target_shift):
            return False

        if not self.is_shift_available_on_date(target_shift, date):
            return False

        required_skills = [
            str(s).strip().lower()
            for s in self.shift_types[target_shift].get('required_skills', [])
            if str(s).strip()
        ]

        rescue_options = []

        # 1) หา donor: คนที่มี skill target แต่ติดเวรอื่นอยู่ในวันเดียวกัน
        for old_shift, donor in list(schedule_dict[date].items()):
            if donor not in self.pharmacists:
                continue
            if old_shift == target_shift:
                continue
            if schedule_dict[date].get(target_shift) not in ['NO SHIFT', 'UNASSIGNED', 'UNFILLED']:
                continue
            if self._is_preassigned_shift(donor, date, old_shift):
                continue
            if self.is_night_shift(old_shift):
                continue

            donor_skills = [s.strip().lower() for s in self.pharmacists[donor].get('skills', [])]
            if required_skills and not all(skill in donor_skills for skill in required_skills):
                continue

            # 2) ลอง remove donor จาก old_shift ชั่วคราว เพื่อเช็กว่า donor ลง target_shift ได้จริงหรือไม่
            original_old_assignee = schedule_dict[date][old_shift]
            original_target_assignee = schedule_dict[date][target_shift]
            schedule_dict[date][old_shift] = 'NO SHIFT'
            schedule_dict[date][target_shift] = 'NO SHIFT'

            temp_hours = pharmacist_hours.copy()
            temp_hours[donor] = max(0, temp_hours.get(donor, 0) - self.shift_types[old_shift]['hours'])

            donor_available = self._get_available_pharmacists_optimized(
                [donor],
                date,
                target_shift,
                schedule_dict,
                temp_hours,
                pharmacist_consecutive_days
            )

            if donor_available:
                # 3) หา replacement สำหรับ old_shift
                replacement_pool = [p for p in pharmacists if p != donor]
                replacement_candidates = self._get_available_pharmacists_optimized(
                    replacement_pool,
                    date,
                    old_shift,
                    schedule_dict,
                    temp_hours,
                    pharmacist_consecutive_days
                )

                if replacement_candidates:
                    replacement = self._select_best_pharmacist(
                        replacement_candidates,
                        old_shift,
                        date,
                        is_day_before_problem_day
                    )

                    old_shift_required_skills = [
                        str(s).strip().lower()
                        for s in self.shift_types[old_shift].get('required_skills', [])
                        if str(s).strip()
                    ]

                    # ยิ่ง old_shift ใช้ skill น้อย ยิ่งเหมาะกับการถูกดึง donor ออก
                    old_shift_scarcity_penalty = 1000 if self._is_skill_scarce_shift(old_shift) else 0
                    old_shift_skill_penalty = len(old_shift_required_skills) * 200
                    donor_target_score = self._calculate_suitability_score(donor_available[0])
                    replacement_score = self._calculate_suitability_score(replacement)
                    total_score = (
                        old_shift_scarcity_penalty
                        + old_shift_skill_penalty
                        + donor_target_score
                        + replacement_score
                    )

                    rescue_options.append({
                        'score': total_score,
                        'donor': donor,
                        'donor_from_shift': old_shift,
                        'target_shift': target_shift,
                        'replacement': replacement['name'],
                        'replacement_to_shift': old_shift,
                    })

            # restore temporary schedule
            schedule_dict[date][old_shift] = original_old_assignee
            schedule_dict[date][target_shift] = original_target_assignee

        if not rescue_options:
            return False

        # เลือก swap ที่กระทบ constraint น้อยที่สุด
        best = min(rescue_options, key=lambda x: x['score'])
        donor = best['donor']
        old_shift = best['donor_from_shift']
        replacement = best['replacement']

        # Apply swap จริง
        schedule_dict[date][old_shift] = replacement
        schedule_dict[date][target_shift] = donor

        self._revert_shift_counts(donor, old_shift)
        pharmacist_hours[donor] = max(0, pharmacist_hours.get(donor, 0) - self.shift_types[old_shift]['hours'])

        self._update_shift_counts(donor, target_shift)
        pharmacist_hours[donor] = pharmacist_hours.get(donor, 0) + self.shift_types[target_shift]['hours']

        self._update_shift_counts(replacement, old_shift)
        pharmacist_hours[replacement] = pharmacist_hours.get(replacement, 0) + self.shift_types[old_shift]['hours']

        print(
            f"RESCUE SWAP {date.strftime('%Y-%m-%d')}: "
            f"{donor} moved {old_shift} -> {target_shift}; "
            f"{replacement} assigned to {old_shift}"
        )
        return True

    def _get_available_pharmacists_optimized(self, pharmacists, date, shift_type, schedule_dict, current_hours_dict, consecutive_days_dict):
        available_pharmacists = []
        pharmacists_on_night_yesterday = set()
        previous_date = date - timedelta(days=1)
        if previous_date in schedule_dict:
            pharmacists_on_night_yesterday = {
                p for s, p in schedule_dict[previous_date].items()
                if p in self.pharmacists and self.is_night_shift(s)
            }

        new_start = self.shift_types[shift_type]['start_time']
        new_end = self.shift_types[shift_type]['end_time']
        new_dept = self.get_department_from_shift(shift_type)

        for pharmacist in pharmacists:
            if date.strftime('%Y-%m-%d') in self.pharmacists[pharmacist]['holidays']: continue
            if self.has_overlapping_shift_optimized(pharmacist, date, shift_type, schedule_dict): continue
            if pharmacist in pharmacists_on_night_yesterday: continue
            p_skills = [skill.strip().lower() for skill in self.pharmacists[pharmacist]['skills']]
            s_req_skills = [skill.strip().lower() for skill in self.shift_types[shift_type]['required_skills'] if skill.strip()]
            if not all(skill in p_skills for skill in s_req_skills): continue

            projected_hours = current_hours_dict[pharmacist] + self.shift_types[shift_type]['hours']
            if projected_hours > self.pharmacists[pharmacist].get('max_hours', 250): continue
            if self.has_restricted_sequence_optimized(pharmacist, date, shift_type, schedule_dict): continue

            # --- START: Junior Constraint Logic ---
            is_junior = 'junior' in p_skills
            if is_junior:
                junior_conflict = False

                # --- 1. Department Level Check (เช็คโควต้าพื้นฐานของแต่ละแผนกตามเดิม) ---
                current_juniors_in_dept = 0
                total_dept_shifts_at_time = 0

                for s_type, s_info in self.shift_types.items():
                    if self.get_department_from_shift(s_type) == new_dept and self.is_shift_available_on_date(s_type, date):
                        s_start = s_info['start_time']
                        s_end = s_info['end_time']
                        if self.check_time_overlap(new_start, new_end, s_start, s_end):
                            total_dept_shifts_at_time += 1

                for existing_shift, assigned_pharm in schedule_dict[date].items():
                    if assigned_pharm in self.pharmacists:
                        existing_is_junior = 'junior' in [s.strip().lower() for s in self.pharmacists[assigned_pharm]['skills']]
                        if existing_is_junior:
                            existing_dept = self.get_department_from_shift(existing_shift)
                            if new_dept == existing_dept:
                                existing_start = self.shift_types[existing_shift]['start_time']
                                existing_end = self.shift_types[existing_shift]['end_time']
                                if self.check_time_overlap(new_start, new_end, existing_start, existing_end):
                                    current_juniors_in_dept += 1

                max_juniors_allowed = 2 if total_dept_shifts_at_time >= 4 else 1

                if shift_type in ['O400F2-8/1', 'O400F2-8/2', 'O400F2-8/3']:
                    max_juniors_allowed = 2

                if shift_type in ['Care/1', 'Care/2']:
                    max_juniors_allowed = 2

                if current_juniors_in_dept + 1 > max_juniors_allowed:
                    junior_conflict = True

                # --- 2. Specific Pair Check (ห้าม Junior คู่กันในเวรที่กำหนด) ---
                if not junior_conflict:
                    # คู่ที่ 1: O400F2-6 กับ O400ER-6
                    pair1 = ('O400F2-6', 'O400ER-6')
                    if shift_type in pair1:
                        other_shift = pair1[0] if shift_type == pair1[1] else pair1[1]
                        assigned_pharm = schedule_dict[date].get(other_shift)

                        # ถ้าเวรคู่กันถูกจัดไปแล้ว ให้เช็กว่าเป็น Junior หรือไม่
                        if assigned_pharm in self.pharmacists:
                            other_is_junior = 'junior' in [s.strip().lower() for s in self.pharmacists[assigned_pharm]['skills']]
                            if other_is_junior:
                                junior_conflict = True

                    # คู่ที่ 2: I100-6 กับ I400-6
                    pair2 = ('I100-6', 'I400-6')
                    if shift_type in pair2:
                        other_shift = pair2[0] if shift_type == pair2[1] else pair2[1]
                        assigned_pharm = schedule_dict[date].get(other_shift)

                        if assigned_pharm in self.pharmacists:
                            other_is_junior = 'junior' in [s.strip().lower() for s in self.pharmacists[assigned_pharm]['skills']]
                            if other_is_junior:
                                junior_conflict = True

                if junior_conflict:
                    continue
            # --- END: Junior Constraint Logic ---

            category = self._get_shift_category(shift_type)
            if category:
                limit = self.shift_limits.get(pharmacist, {}).get(category)
                if limit is not None:
                    current_count = self.pharmacists[pharmacist]['category_counts'][category]
                    if current_count >= limit:
                        continue
            if self.is_night_shift(shift_type):
                if self.has_nearby_night_shift_optimized(pharmacist, date, schedule_dict): continue
                next_date = date + timedelta(days=1)
                if pharmacist in self.pre_assignments and next_date.strftime('%Y-%m-%d') in self.pre_assignments[pharmacist]: continue
            if shift_type.startswith('C8'):
                if not self.check_mixing_expert_ratio_optimized(schedule_dict, date, shift_type, pharmacist):
                    continue
            current_streak = self.get_dynamic_consecutive_days(pharmacist, date, schedule_dict)

            # Hard Constraint: ถ้าบวกเวรนี้เข้าไปแล้วทำงานติดกันเกินเพดาน ให้ตัดชื่อทิ้งทันที
            if current_streak >= self.MAX_CONSECUTIVE_DAYS:
                continue

            original_preference = self.get_preference_score(pharmacist, shift_type)
            multiplier = self.preference_multipliers.get(pharmacist, 1.0)

            # --- START: New Pacing Metrics ---
            max_hrs = self.pharmacists[pharmacist].get('max_hours', 250)
            current_hrs = current_hours_dict[pharmacist]

            days_in_month = pd.Timestamp(date).days_in_month
            time_elapsed_pct = date.day / days_in_month
            hours_used_pct = current_hrs / max_hrs if max_hrs > 0 else 1.0
            # --- END: New Pacing Metrics ---

            # --- START: New Soul Mate & Weekend Prep ---
            is_weekend = date.weekday() >= 5
            weekend_days_worked = 0
            if is_weekend:
                weekend_days_worked = self.get_weekend_days_worked(pharmacist, schedule_dict)

            soulmate = self.soul_mates.get(pharmacist)
            soulmate_working_today = False
            mate_on_holiday = False

            if soulmate:
                # เช็คว่าคู่หูถูกจัดให้ลงเวรในวันนี้ไปหรือยัง
                if soulmate in schedule_dict[date].values():
                    soulmate_working_today = True
                # เช็คว่าวันนี้คู่หูลาหยุด (Holiday) หรือไม่
                if soulmate in self.pharmacists and date.strftime('%Y-%m-%d') in self.pharmacists[soulmate]['holidays']:
                    mate_on_holiday = True
            # --- END: New Soul Mate & Weekend Prep ---

            pharmacist_data = {
                'name': pharmacist,
                'preference_score': original_preference * multiplier,
                'consecutive_days': current_streak,
                'night_count': self.pharmacists[pharmacist]['night_shift_count'],
                'mixing_count': self.pharmacists[pharmacist]['mixing_shift_count'],
                'care_count': self.pharmacists[pharmacist].get('care_shift_count', 0),
                'current_hours': current_hrs,
                'max_hours': max_hrs,
                'time_elapsed_pct': time_elapsed_pct,
                'hours_used_pct': hours_used_pct,

                # นำตัวแปรใหม่ส่งเข้าไปให้ฟังก์ชันคิดคะแนน
                'is_weekend': is_weekend,
                'weekend_days_worked': weekend_days_worked,
                'has_soulmate': bool(soulmate),
                'soulmate_working_today': soulmate_working_today,
                'mate_on_holiday': mate_on_holiday,
                'needs_min_shift': self._needs_min_shift(pharmacist, shift_type, schedule_dict),
                'no_preference': self.pharmacists[pharmacist].get('no_preference', False),
                'department_count': self.get_dept_shift_count(pharmacist, new_dept, schedule_dict) if new_dept else 0,
                'total_shift_count': self.get_total_shift_count_in_schedule_dict(pharmacist, schedule_dict),
                'has_worked_this_department': (new_dept in self.get_unique_departments_worked(pharmacist, schedule_dict)) if new_dept else True,
                'average_monthly_shift_target': self.get_average_monthly_shift_target(date.year, date.month),
                'month_segment_shift_count': self.get_month_segment_shift_count(pharmacist, date, schedule_dict),
                'total_weekend_days': self.get_total_weekend_days_in_schedule(schedule_dict),
                'weekend_days_worked_before': self.get_weekend_days_worked_until(pharmacist, schedule_dict),
            }
            available_pharmacists.append(pharmacist_data)
        return available_pharmacists
    def validate_min_shift_requirements(self, schedule):
        """
        ตรวจสอบหลัง generate ว่า pharmacist คนไหนยังไม่ถึง MinCount
        คืนค่า list of (pharmacist, department, required, actual)
        """
        violations = []
        schedule_dict = {date: schedule.loc[date].to_dict() for date in schedule.index}

        for pharmacist, dept_reqs in self.min_shift_requirements.items():
            for department, min_count in dept_reqs.items():
                actual = self.get_dept_shift_count(pharmacist, department, schedule_dict)
                if actual < min_count:
                    violations.append({
                        'pharmacist': pharmacist,
                        'department': department,
                        'required': min_count,
                        'actual': actual,
                        'shortfall': min_count - actual
                    })

        if violations:
            print("\n⚠️  MIN SHIFT REQUIREMENT VIOLATIONS:")
            print(f"{'Pharmacist':<30} {'Dept':<12} {'Required':>8} {'Actual':>8} {'Short':>8}")
            print("-" * 70)
            for v in violations:
                print(f"{v['pharmacist']:<30} {v['department']:<12} "
                      f"{v['required']:>8} {v['actual']:>8} {v['shortfall']:>8}")
        else:
            print("\n✅ All minimum shift requirements are satisfied.")

        return violations
    def _calculate_suitability_score(self, pharmacist_data):
        consecutive_penalty = self.W_CONSECUTIVE * (pharmacist_data['consecutive_days'] ** 2)
        hours_penalty = self.W_HOURS * (pharmacist_data['hours_used_pct'] * 100)
        preference_penalty = self.W_PREFERENCE * pharmacist_data['preference_score']

        min_shift_bonus = -200 if pharmacist_data.get('needs_min_shift', False) else 0

        no_preference_department_balance_penalty = 0
        if pharmacist_data.get('no_preference', False):
            # กลุ่มไม่มี Preference: บังคับแนวโน้มให้วนครบทุก Department ก่อน
            # และลดโอกาสลง Department เดิมซ้ำ ๆ เพื่อให้จำนวนแต่ละ Department เท่า ๆ กัน
            department_count = pharmacist_data.get('department_count', 0)
            total_shift_count = pharmacist_data.get('total_shift_count', 0)
            has_worked_this_department = pharmacist_data.get('has_worked_this_department', True)
            no_preference_department_balance_penalty = (department_count * 200) + (total_shift_count * 3)
            if not has_worked_this_department:
                no_preference_department_balance_penalty -= 300

        # 1) Hour pacing เดิม: กันชั่วโมงงานเกินสัดส่วนวันที่ผ่านไป
        pacing_penalty = 0
        if pharmacist_data['hours_used_pct'] > pharmacist_data['time_elapsed_pct']:
            diff = pharmacist_data['hours_used_pct'] - pharmacist_data['time_elapsed_pct']
            pacing_penalty = 500 * diff

        # 2) Shift-count pacing ใหม่: กันเวรกระจุกต้นเดือน/ปลายเดือน
        # ถ้าจำนวนเวรที่ได้ มากกว่าสัดส่วนของเดือนที่ผ่านไปมากเกิน จะโดน penalty
        target_monthly_shifts = pharmacist_data.get('average_monthly_shift_target', 1)
        current_shift_count = pharmacist_data.get('total_shift_count', 0)
        projected_shift_pct = (current_shift_count + 1) / target_monthly_shifts if target_monthly_shifts > 0 else 1
        month_progress_pct = pharmacist_data.get('time_elapsed_pct', 1)
        allowed_pacing_buffer = 0.12
        shift_pacing_penalty = 0
        if projected_shift_pct > month_progress_pct + allowed_pacing_buffer:
            shift_pacing_penalty = self.W_SHIFT_PACING * ((projected_shift_pct - month_progress_pct - allowed_pacing_buffer) ** 2)

        # 3) Segment balance ใหม่: กันคนเดิมไปกระจุกต้น/กลาง/ปลายเดือน
        segment_count = pharmacist_data.get('month_segment_shift_count', 0)
        month_segment_penalty = self.W_MONTH_SEGMENT_BALANCE * (segment_count ** 2)

        # 4) Weekend-off protection ใหม่: พยายามให้แต่ละคนมีวันหยุดเสาร์-อาทิตย์อย่างน้อย 4 วัน/เดือน
        weekend_off_protection_penalty = 0
        if pharmacist_data.get('is_weekend', False):
            total_weekend_days = pharmacist_data.get('total_weekend_days', 0)
            weekend_days_worked_before = pharmacist_data.get('weekend_days_worked_before', 0)
            max_weekend_work_days = max(total_weekend_days - self.MIN_WEEKEND_OFF_DAYS, 0)
            projected_weekend_work_days = weekend_days_worked_before + 1
            if projected_weekend_work_days > max_weekend_work_days:
                shortfall = projected_weekend_work_days - max_weekend_work_days
                weekend_off_protection_penalty = self.W_WEEKEND_OFF_PROTECTION * (shortfall ** 2)

        return (
            consecutive_penalty
            + hours_penalty
            + preference_penalty
            + min_shift_bonus
            + no_preference_department_balance_penalty
            + pacing_penalty
            + shift_pacing_penalty
            + month_segment_penalty
            + weekend_off_protection_penalty
        )

    def _select_best_pharmacist(self, available_pharmacists, shift_type, date, is_day_before_problem_day):
        if self.is_night_shift(shift_type) and is_day_before_problem_day:
            problem_day = date + timedelta(days=1)
            problem_day_str = problem_day.strftime('%Y-%m-%d')

            candidates_off_tomorrow = []
            for p_data in available_pharmacists:
                p_name = p_data['name']
                if problem_day_str in self.pharmacists[p_name]['holidays']:
                    candidates_off_tomorrow.append(p_data)

            if candidates_off_tomorrow:
                print(f"INFO: Prioritizing night shift on {date.strftime('%Y-%m-%d')} for pharmacists off on problem day {problem_day_str}.")
                return min(candidates_off_tomorrow, key=lambda x: (x['night_count'], self._calculate_suitability_score(x)))

        if self.is_night_shift(shift_type):
            return min(available_pharmacists, key=lambda x: (x['night_count'], self._calculate_suitability_score(x)))
        elif shift_type.startswith('C8'):
            return min(available_pharmacists, key=lambda x: (x['mixing_count'], self._calculate_suitability_score(x)))
        elif shift_type.startswith('Care'):                                                        # ← เพิ่ม
            return min(available_pharmacists, key=lambda x: (x['care_count'], self._calculate_suitability_score(x)))  # ← เพิ่ม
        else:
            return min(available_pharmacists, key=lambda x: self._calculate_suitability_score(x))

    def calculate_preference_penalty(self, pharmacist, schedule):
        penalty = 0
        for date in schedule.index:
            for shift_type, assigned_pharm in schedule.loc[date].items():
                if assigned_pharm == pharmacist:
                    penalty += self.get_preference_score(pharmacist, shift_type)
        return penalty

    def get_dynamic_consecutive_days(self, pharmacist, date, schedule_dict):
        """นับจำนวนวันทำงานต่อเนื่อง โดยกวาดดูทั้งอดีตและอนาคตจาก Schedule ปัจจุบัน"""
        streak = 0

        # เช็คย้อนหลัง (Backward)
        curr_date = date - timedelta(days=1)
        while curr_date in schedule_dict:
            worked_backward = any(p == pharmacist for p in schedule_dict[curr_date].values() if p not in ['NO SHIFT', 'UNFILLED', 'UNASSIGNED'])
            if worked_backward:
                streak += 1
                curr_date -= timedelta(days=1)
            else:
                break

        # เช็คเดินหน้า (Forward - จำเป็นเพราะเราจัด Problem Days ก่อน)
        curr_date = date + timedelta(days=1)
        while curr_date in schedule_dict:
            worked_forward = any(p == pharmacist for p in schedule_dict[curr_date].values() if p not in ['NO SHIFT', 'UNFILLED', 'UNASSIGNED'])
            if worked_forward:
                streak += 1
                curr_date += timedelta(days=1)
            else:
                break

        return streak

    def get_weekend_days_worked(self, pharmacist, schedule_dict):
        """นับจำนวนวันเสาร์-อาทิตย์ที่ทำงานไปแล้วในเดือนนี้"""
        weekend_days = 0
        for d, shifts in schedule_dict.items():
            if d.weekday() >= 5 and pharmacist in shifts.values():
                weekend_days += 1
        return weekend_days

    def is_schedule_better(self, current_metrics, best_metrics):
        current_unfilled = current_metrics.get('unfilled_problem_shifts', float('inf'))
        best_unfilled = best_metrics.get('unfilled_problem_shifts', float('inf'))
        if current_unfilled < best_unfilled: return True
        if current_unfilled > best_unfilled: return False
        weights = {
            'preference_score': 1.0,
            'preference_variance': 50.0,
            'hour_imbalance_penalty': 25.0,
            'night_variance': 800.0,
            'weekend_off_variance': 1000.0,
            'weekend_min_off_shortfall': 5000.0,
            'month_segment_variance': 2500.0,
        }
        current_score = sum(weights[k] * current_metrics.get(k, 0) for k in weights)
        best_score = sum(weights[k] * best_metrics.get(k, 0) for k in weights)
        return current_score < best_score

    def optimize_schedule(self, year, month, iterations=10, true_random_override=False, enable_run_log=False):
        self.run_logs = []
        self.run_config = {
            "Year": year,
            "Month": month,
            "Iterations": iterations,
            "True Random Override": true_random_override,
            "Enable Run Log": enable_run_log,
            "Staff Type": self.staff_type,
            "Run At": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        }

        self._log_schedule_event(
            "RUN_START",
            "Start schedule generation",
            Year=year,
            Month=month,
            Iterations=iterations,
            TrueRandomOverride=true_random_override,
            EnableRunLog=enable_run_log
        )

        if true_random_override:
            print("\n⚠️ TRUE RANDOM OVERRIDE MODE ENABLED")
            print("ระบบจะสุ่มจัดเวรแบบ constrained + fairness-aware random")
            print("ยังคงบังคับ: เวรเปิดจริง, ไม่ overwrite PreAssignments, ไม่จัดหลายเวรในวันเดียว, ไม่จัดคนลา, ต้องมี skill ตรง")
            print("ยังคงคุม Fairness / Hour Balance: สุ่มจากกลุ่มคนที่ชั่วโมงน้อยหรือใกล้เคียงคนชั่วโมงน้อยก่อน")
            print("iterations ยังใช้ได้: ระบบจะสุ่มหลายรอบและเลือกตารางที่ hour balance ดีกว่า")

            best_schedule = None
            best_unfilled_info = {}
            best_metrics = {
                'unfilled_problem_shifts': float('inf'),
                'hour_imbalance_penalty': float('inf'),
                'hour_diff_for_logging': float('inf')
            }

            for i in range(iterations):
                print(f"\n--- Fair Random Iteration {i + 1}/{iterations} ---")
                current_schedule, unfilled_info = self.generate_monthly_schedule_true_random(
                    year=year,
                    month=month,
                    iteration_num=i + 1
                )

                metrics = self.calculate_schedule_metrics(current_schedule, year, month)
                metrics['unfilled_problem_shifts'] = (
                    len(unfilled_info.get('problem_days', [])) + len(unfilled_info.get('other_days', []))
                )

                print(
                    f"Fair Random Results -> "
                    f"Unfilled Shifts: {metrics['unfilled_problem_shifts']} | "
                    f"Hour SD: {metrics.get('hour_diff_for_logging', 0):.2f} | "
                    f"Hour Penalty: {metrics.get('hour_imbalance_penalty', 0):.2f}"
                )

                self._log_schedule_event(
                    "TRUE_RANDOM_ITERATION_RESULT",
                    "Completed one fairness-aware random iteration",
                    Iteration=i + 1,
                    UnfilledShifts=metrics['unfilled_problem_shifts'],
                    HourSD=round(metrics.get('hour_diff_for_logging', 0), 4),
                    HourPenalty=round(metrics.get('hour_imbalance_penalty', 0), 4)
                )

                current_key = (
                    metrics['unfilled_problem_shifts'],
                    metrics.get('hour_imbalance_penalty', float('inf')),
                    metrics.get('hour_diff_for_logging', float('inf'))
                )
                best_key = (
                    best_metrics.get('unfilled_problem_shifts', float('inf')),
                    best_metrics.get('hour_imbalance_penalty', float('inf')),
                    best_metrics.get('hour_diff_for_logging', float('inf'))
                )

                if best_schedule is None or current_key < best_key:
                    best_schedule = current_schedule.copy()
                    best_unfilled_info = unfilled_info.copy()
                    best_metrics = metrics.copy()
                    print("*** Found a fairer random schedule! ***")

                    self._log_schedule_event(
                        "TRUE_RANDOM_BEST_UPDATED",
                        "Found a better fairness-aware random schedule",
                        Iteration=i + 1,
                        UnfilledShifts=metrics['unfilled_problem_shifts'],
                        HourSD=round(metrics.get('hour_diff_for_logging', 0), 4),
                        HourPenalty=round(metrics.get('hour_imbalance_penalty', 0), 4)
                    )

            self._log_schedule_event(
                "RUN_COMPLETE",
                "Completed fairness-aware constrained true random override mode",
                TotalUnfilled=len(best_unfilled_info.get('problem_days', [])) + len(best_unfilled_info.get('other_days', [])),
                FinalHourSD=round(best_metrics.get('hour_diff_for_logging', 0), 4),
                FinalHourPenalty=round(best_metrics.get('hour_imbalance_penalty', 0), 4)
            )

            return best_schedule, best_unfilled_info

        best_schedule = None
        best_metrics = {
            'unfilled_problem_shifts': float('inf'),
            'hour_imbalance_penalty': float('inf'),
            'night_variance': float('inf'),
            'preference_score': float('inf')
        }
        best_unfilled_info = {}

        self._pre_check_staffing_levels(year, month)
        print(f"\nStarting optimization with {iterations} iterations...")

        for i in range(iterations):
            print(f"\n--- Iteration {i + 1}/{iterations} ---")
            current_schedule, unfilled_info = self.generate_monthly_schedule_shuffled(
                year,
                month,
                iteration_num=i + 1
            )

            metrics = self.calculate_schedule_metrics(current_schedule, year, month)
            metrics['unfilled_problem_shifts'] = len(unfilled_info['problem_days']) + len(unfilled_info['other_days'])

            print(
                f"Iteration Results -> "
                f"Unfilled Shifts: {metrics['unfilled_problem_shifts']} | "
                f"Hour SD: {metrics.get('hour_diff_for_logging', 0):.2f} | "
                f"Night Var: {metrics.get('night_variance', 0):.2f} | "
                f"Weekend Shortfall: {metrics.get('weekend_min_off_shortfall', 0)} | "
                f"Month Segment Var: {metrics.get('month_segment_variance', 0):.2f} | "
                f"Pref Penalty: {metrics.get('preference_score', 0):.1f}"
            )

            self._log_schedule_event(
                "ITERATION_RESULT",
                "Completed one optimization iteration",
                Iteration=i + 1,
                UnfilledShifts=metrics['unfilled_problem_shifts'],
                HourSD=round(metrics.get('hour_diff_for_logging', 0), 4),
                NightVariance=round(metrics.get('night_variance', 0), 4),
                WeekendShortfall=metrics.get('weekend_min_off_shortfall', 0),
                MonthSegmentVariance=round(metrics.get('month_segment_variance', 0), 4),
                PreferencePenalty=round(metrics.get('preference_score', 0), 4)
            )

            if best_schedule is None or self.is_schedule_better(metrics, best_metrics):
                best_schedule = current_schedule.copy()
                best_metrics = metrics.copy()
                best_unfilled_info = unfilled_info.copy()
                print("*** Found a more balanced schedule! ***")

                self._log_schedule_event(
                    "BEST_UPDATED",
                    "Found a better schedule",
                    Iteration=i + 1,
                    UnfilledShifts=metrics['unfilled_problem_shifts'],
                    HourSD=round(metrics.get('hour_diff_for_logging', 0), 4),
                    NightVariance=round(metrics.get('night_variance', 0), 4),
                    WeekendShortfall=metrics.get('weekend_min_off_shortfall', 0),
                    MonthSegmentVariance=round(metrics.get('month_segment_variance', 0), 4),
                    PreferencePenalty=round(metrics.get('preference_score', 0), 4)
                )

        if best_schedule is not None:
            print("\nOptimization complete!\nFinal metrics for the best schedule found:")
            print(
                f"Unfilled Shifts: {best_metrics.get('unfilled_problem_shifts', 0)} | "
                f"Hour SD: {best_metrics.get('hour_diff_for_logging', 0):.2f} | "
                f"Night Var: {best_metrics.get('night_variance', 0):.2f} | "
                f"Weekend Shortfall: {best_metrics.get('weekend_min_off_shortfall', 0)} | "
                f"Month Segment Var: {best_metrics.get('month_segment_variance', 0):.2f} | "
                f"Pref Penalty: {best_metrics.get('preference_score', 0):.1f}"
            )
            self.validate_min_shift_requirements(best_schedule)

            self._log_schedule_event(
                "RUN_COMPLETE",
                "Completed optimization mode",
                FinalUnfilledShifts=best_metrics.get('unfilled_problem_shifts', 0),
                FinalHourSD=round(best_metrics.get('hour_diff_for_logging', 0), 4),
                FinalNightVariance=round(best_metrics.get('night_variance', 0), 4),
                FinalWeekendShortfall=best_metrics.get('weekend_min_off_shortfall', 0),
                FinalMonthSegmentVariance=round(best_metrics.get('month_segment_variance', 0), 4),
                FinalPreferencePenalty=round(best_metrics.get('preference_score', 0), 4)
            )
        else:
            print("\nOptimization failed to find any valid schedule.")
            self._log_schedule_event(
                "RUN_FAILED",
                "Optimization failed to find valid schedule"
            )

        return best_schedule, best_unfilled_info


    def export_to_excel(self, schedule, unfilled_info, filename, enable_run_log=False):
        wb = Workbook()
        ws = wb.active
        ws.title = 'Monthly Schedule'
        ws_daily = wb.create_sheet("Daily Summary")
        ws_daily_codes = wb.create_sheet("Daily Summary (Codes)")
        ws_pref = wb.create_sheet("Preference Scores")
        ws_negotiate = wb.create_sheet("Negotiation Suggestions")
        ws_signature = wb.create_sheet("Signature Sheet")
        ws_min_req = wb.create_sheet("Min Req Violations")
        ws_run_logs = wb.create_sheet("Run Logs") if enable_run_log else None
        header_fill = PatternFill(start_color='FFD3D3D3', end_color='FFD3D3D3', fill_type='solid')
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        ws.cell(row=1, column=1, value='Date').fill = header_fill
        for col, shift_type in enumerate(self.shift_types, 2):
            cell = ws.cell(row=1, column=col, value=f"{self.shift_types[shift_type]['description']}\n({self.shift_types[shift_type]['hours']} hrs)")
            cell.fill, cell.font, cell.alignment = header_fill, Font(bold=True), Alignment(wrap_text=True)
        # Sort the schedule by date before exporting
        schedule.sort_index(inplace=True)
        for row, date in enumerate(schedule.index, 2):
            ws.cell(row=row, column=1, value=date.strftime('%Y-%m-%d'))
            is_holiday = self.is_holiday(date)
            is_weekend = date.weekday() >= 5
            for col, shift_type in enumerate(self.shift_types, 2):
                cell = ws.cell(row=row, column=col, value=schedule.loc[date, shift_type])
                cell.border = border
                if schedule.loc[date, shift_type] == 'NO SHIFT': cell.fill = PatternFill(start_color='FFCCCCCC', fill_type='solid')
                elif is_holiday: cell.fill = PatternFill(start_color='FFFFB6C1', fill_type='solid')
                elif is_weekend: cell.fill = PatternFill(start_color='FFFFE4E1', fill_type='solid')
                elif schedule.loc[date, shift_type] == 'UNFILLED': cell.fill = PatternFill(start_color='FFFFFF00', fill_type='solid')
        ws.column_dimensions['A'].width = 22
        for col in range(2, len(self.shift_types) + 2):
            ws.column_dimensions[get_column_letter(col)].width = 20
        self.create_schedule_summaries(ws, schedule)
        self.create_daily_summary(ws_daily, schedule)
        self.create_preference_score_summary(ws_pref, schedule)
        self.create_daily_summary_with_codes(ws_daily_codes, schedule)
        self.create_negotiation_summary(ws_negotiate, schedule)

        self.create_signature_sheet(ws_signature, schedule)

        ws_min_req = wb.create_sheet("Min Req Violations")
        violations = self.validate_min_shift_requirements(schedule)

        headers = ["Pharmacist", "Department", "Required", "Actual", "Shortfall"]
        header_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
        border = Border(left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin'))

        for col, h in enumerate(headers, 1):
            cell = ws_min_req.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = Font(bold=True, color="FFFFFFFF")
            cell.border = border

        if not violations:
            ws_min_req.cell(row=2, column=1, value="✅ All minimum shift requirements satisfied.")
        else:
            for row, v in enumerate(violations, 2):
                for col, val in enumerate([
                    v['pharmacist'], v['department'],
                    v['required'], v['actual'], v['shortfall']
                ], 1):
                    cell = ws_min_req.cell(row=row, column=col, value=val)
                    cell.border = border
                    if v['shortfall'] > 0:
                        cell.fill = PatternFill(start_color='FFFFF2CC',
                                                end_color='FFFFF2CC', fill_type='solid')

        ws_min_req.column_dimensions['A'].width = 35
        for col in ['B','C','D','E']:
            ws_min_req.column_dimensions[col].width = 15


        if enable_run_log and ws_run_logs is not None:
            ws_run_logs.cell(row=1, column=1, value="Run Configuration").font = Font(bold=True)

            config_row = 2
            for key, value in self.run_config.items():
                ws_run_logs.cell(row=config_row, column=1, value=key)
                ws_run_logs.cell(row=config_row, column=2, value=str(value))
                config_row += 1

            log_start_row = config_row + 2
            ws_run_logs.cell(row=log_start_row, column=1, value="Run Logs").font = Font(bold=True)

            if self.run_logs:
                all_keys = []
                for log in self.run_logs:
                    for key in log.keys():
                        if key not in all_keys:
                            all_keys.append(key)

                header_row = log_start_row + 1
                for col_idx, key in enumerate(all_keys, 1):
                    cell = ws_run_logs.cell(row=header_row, column=col_idx, value=key)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color='FFD3D3D3', end_color='FFD3D3D3', fill_type='solid')
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

                for row_idx, log in enumerate(self.run_logs, header_row + 1):
                    for col_idx, key in enumerate(all_keys, 1):
                        ws_run_logs.cell(row=row_idx, column=col_idx, value=str(log.get(key, "")))

                for col_idx, key in enumerate(all_keys, 1):
                    ws_run_logs.column_dimensions[get_column_letter(col_idx)].width = min(max(len(str(key)) + 5, 15), 45)
            else:
                ws_run_logs.cell(row=log_start_row + 1, column=1, value="No logs recorded.")

        wb.save(filename)

    def create_signature_sheet(self, ws, schedule):
        # 1. กำหนด Style
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        header_fill = PatternFill(start_color='FFD3D3D3', end_color='FFD3D3D3', fill_type='solid')
        x_fill = PatternFill(start_color='FFD3D3D3', end_color='FFD3D3D3', fill_type='solid') # สีเทาอ่อนสำหรับช่อง X
        white_fill = PatternFill(start_color='FFFFFFFF', end_color='FFFFFFFF', fill_type='solid') # สีขาวสำหรับช่องว่าง

        # Mapping สีตามกลุ่มเวรให้ตรงกับ Daily Summary 100%
        shift_colors = {
            'I100': 'FF00B050',
            'O100': 'FF00B0F0',
            'Care': 'FFD40202',
            'C8': 'FFE6B8AF',
            'I400': 'FFFF00FF',
            'O400F1': 'FF0033CC',
            'O400F2': 'FFC78AF2',
            'O400ER': 'FFED7D31',
            'ARI': 'FF7030A0',
            'Refill': 'FF741b47'
        }

        # 2. สร้าง Header แถวบนสุด (วันที่)
        ws.cell(row=1, column=1, value='Shift / Date').fill = header_fill
        ws.cell(row=1, column=1).border = border
        ws.cell(row=1, column=1).font = Font(bold=True)

        sorted_dates = sorted(schedule.index)
        for col, date in enumerate(sorted_dates, 2):
            cell = ws.cell(row=1, column=col, value=date.strftime('%d/%m'))
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.font = Font(bold=True)

        # 3. สร้างข้อมูลแต่ละแถว (เรียงตามรหัสเวร)
        for row, shift_type in enumerate(self.shift_types.keys(), 2):
            shift_info = self.shift_types[shift_type]

            # จัด Format ชื่อเวร: ชื่อ (ชั่วโมง) \n (เวลาเริ่ม - เวลาจบ)
            shift_desc = f"{shift_info['description']} ({int(shift_info['hours'])} ชม.)\n({shift_info['start_time']} - {shift_info['end_time']})"

            # กำหนดสีพื้นหลังประจำแถวตามรหัสเวร
            row_color = 'FFFFFFFF'
            font_color = 'FF000000' # ค่าเริ่มต้นสีดำ

            # เรียง Prefix จากยาวไปสั้น เพื่อให้ระบบเช็ค O400F1, O400F2 ก่อน O400 ธรรมดา
            for prefix in sorted(shift_colors.keys(), key=len, reverse=True):
                if shift_type.startswith(prefix):
                    row_color = shift_colors[prefix]
                    # แผนกที่สีพื้นหลังเข้ม ให้ใช้ตัวหนังสือสีขาว (อิงตาม Daily Summary)
                    if prefix in ['Care', 'O400F1', 'ARI','Refill']:
                        font_color = 'FFFFFFFF'
                    break

            shift_label_fill = PatternFill(start_color=row_color, end_color=row_color, fill_type='solid')

            # ลงชื่อเวรในคอลัมน์แรก
            cell_first = ws.cell(row=row, column=1, value=shift_desc)
            cell_first.fill = shift_label_fill
            cell_first.border = border
            cell_first.alignment = Alignment(wrap_text=True, vertical='center', horizontal='left')
            cell_first.font = Font(color=font_color, bold=True)

            # 4. หยอดข้อมูล X หรือเว้นว่าง ในแต่ละวัน
            for col, date in enumerate(sorted_dates, 2):
                cell = ws.cell(row=row, column=col)
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center')

                status = schedule.loc[date, shift_type]

                # ถ้าเป็นวันนั้นไม่มีเวร ให้ใส่ X และทำพื้นหลังสีเทา
                if status == 'NO SHIFT':
                    cell.value = 'X'
                    cell.fill = x_fill
                else:
                    # ถ้ามีเวร ให้ปล่อยว่างไว้เซ็นชื่อ และใช้พื้นหลังสีขาวล้วน
                    cell.value = ''
                    cell.fill = white_fill

        # 5. ปรับขนาดความกว้างของคอลัมน์
        ws.column_dimensions['A'].width = 40
        for col in range(2, len(sorted_dates) + 2):
            ws.column_dimensions[get_column_letter(col)].width = 7

    def create_negotiation_summary(self, ws, schedule):
        header_fill = PatternFill(start_color='FF4F81BD', end_color='FF4F81BD', fill_type='solid')
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        bold_white_font = Font(bold=True, color="FFFFFFFF")
        alignment = Alignment(wrap_text=True, vertical='top')
        headers = ["Date", "Unfilled Shift", "Suggested Negotiation Candidates (Ranked)"]
        for col, header_text in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header_text)
            cell.fill, cell.font, cell.border, cell.alignment = header_fill, bold_white_font, border, alignment
        unfilled_shifts = []
        for date in schedule.index:
            for shift_type, assigned_pharm in schedule.loc[date].items():
                if assigned_pharm == 'UNFILLED':
                    unfilled_shifts.append((date, shift_type))
        if not unfilled_shifts:
            ws.cell(row=2, column=1, value="No unfilled shifts in the final schedule.")
            return
        current_row = 2
        for date, shift_type in unfilled_shifts:
            required_skills = self.shift_types[shift_type].get('required_skills', [])
            all_candidates = []
            for p_name, p_info in self.pharmacists.items():
                if not all(skill.strip() in p_info['skills'] for skill in required_skills if skill.strip()): continue
                if len(self.get_pharmacist_shifts(p_name, date, schedule)) > 0: continue
                is_on_holiday = date.strftime('%Y-%m-%d') in p_info['holidays']
                # --- START FIX: เพิ่มตัวแปรคำนวณ % การใช้ชั่วโมง ---
                max_hrs = p_info.get('max_hours', 250)
                current_hrs = self.calculate_total_hours(p_name, schedule)

                days_in_month = pd.Timestamp(date).days_in_month
                time_elapsed_pct = date.day / days_in_month
                hours_used_pct = current_hrs / max_hrs if max_hrs > 0 else 1.0

                pharmacist_data = {
                    'name': p_name,
                    'preference_score': self.get_preference_score(p_name, shift_type),
                    'consecutive_days': self.count_consecutive_shifts(p_name, date, schedule),
                    'night_count': p_info.get('night_shift_count', 0),
                    'mixing_count': p_info.get('mixing_shift_count', 0),
                    'current_hours': current_hrs,
                    'max_hours': max_hrs,
                    'time_elapsed_pct': time_elapsed_pct,
                    'hours_used_pct': hours_used_pct
                }
                # --- END FIX ---

                suitability_score = self._calculate_suitability_score(pharmacist_data)
                suitability_score = self._calculate_suitability_score(pharmacist_data)
                all_candidates.append({'name': p_name, 'is_on_holiday': is_on_holiday, 'score': suitability_score})
            sorted_candidates = sorted(all_candidates, key=lambda x: (x['is_on_holiday'], x['score']))
            suggestions_text = []
            for i, cand in enumerate(sorted_candidates[:3]):
                status = "(On Holiday)" if cand['is_on_holiday'] else "(Available)"
                suggestions_text.append(f"{i+1}. {cand['name']} {status}")
            final_text = "\n".join(suggestions_text) if suggestions_text else "No suitable candidate found"
            ws.cell(row=current_row, column=1, value=date.strftime('%Y-%m-%d')).border = border
            ws.cell(row=current_row, column=2, value=shift_type).border = border
            cell = ws.cell(row=current_row, column=3, value=final_text)
            cell.border, cell.alignment, ws.row_dimensions[current_row].height = border, alignment, 50
            current_row += 1
        ws.column_dimensions['A'].width, ws.column_dimensions['B'].width, ws.column_dimensions['C'].width = 15, 25, 45

    def create_schedule_summaries(self, ws, schedule):
        summary_row = len(schedule) + 3
        ws.cell(row=summary_row, column=1, value="Summary").font = Font(bold=True)
        hours_row = summary_row + 2
        ws.cell(row=hours_row, column=1, value="Working Hours Summary").font = Font(bold=True)
        for i, pharmacist in enumerate(self.pharmacists):
            hours = self.calculate_total_hours(pharmacist, schedule)
            ws.cell(row=hours_row + i + 1, column=1, value=pharmacist)
            ws.cell(row=hours_row + i + 1, column=2, value=f"Total Hours: {hours}")
        night_row = hours_row + len(self.pharmacists) + 2
        ws.cell(row=night_row, column=1, value="Night Shift Summary").font = Font(bold=True)
        for i, pharmacist in enumerate(self.pharmacists):
            row = night_row + i + 1
            ws.cell(row=row, column=1, value=pharmacist)
            ws.cell(row=row, column=2, value=f"Night Shifts: {self.pharmacists[pharmacist]['night_shift_count']}")
        shift_row = night_row + len(self.pharmacists) + 2
        ws.cell(row=shift_row, column=1, value="Shift Count Summary").font = Font(bold=True)
        shift_types_list = list(self.shift_types.keys())
        for col_idx, shift_type in enumerate(shift_types_list, 2):
            ws.cell(row=shift_row, column=col_idx, value=shift_type).font = Font(bold=True)
        for row_idx, pharmacist in enumerate(self.pharmacists, 1):
            row_num = shift_row + row_idx
            ws.cell(row=row_num, column=1, value=pharmacist)
            for col_idx, shift_type in enumerate(shift_types_list, 2):
                count = sum(1 for d in schedule.index if schedule.loc[d, shift_type] == pharmacist)
                ws.cell(row=row_num, column=col_idx, value=count)

    def _setup_daily_summary_styles(self):
        return {
            'header_fill': PatternFill(fill_type='solid', start_color='FFD3D3D3'),
            'weekend_fill': PatternFill(fill_type='solid', start_color='FFFFE4E1'),
            'holiday_fill': PatternFill(fill_type='solid', start_color='FFFFB6C1'),
            'holiday_empty_fill': PatternFill(fill_type='solid', start_color='FFFFFF00'),
            'off_fill': PatternFill(fill_type='solid', start_color='FFD3D3D3'),
            'border': Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')),
            'fills': {p: PatternFill(fill_type='solid', start_color=c) for p, c in [
                ('I100', 'FF00B050'), ('O100', 'FF00B0F0'), ('Care', 'FFD40202'), ('C8', 'FFE6B8AF'),
                ('I400', 'FFFF00FF'), ('O400F1', 'FF0033CC'), ('O400F2', 'FFC78AF2'),
                ('O400ER', 'FFED7D31'), ('ARI', 'FF7030A0'), ('Refill', 'FF741b47')]},
            'fonts': {
                'O400F1': Font(bold=True, color="FFFFFFFF"), 'ARI': Font(bold=True, color="FFFFFFFF"),
                'Refill': Font(bold=True, color="FFFFFFFF"),
                'default': Font(bold=True), 'header': Font(bold=True)
            }
        }

    def convert_excel_to_gsheet(self, excel_file_path, gsheet_name):
        print("\nAuthenticating and converting to Google Sheets...")
        auth.authenticate_user()
        service = build('drive', 'v3')

        file_metadata = {
            'name': gsheet_name,
            'mimeType': 'application/vnd.google-apps.spreadsheet' # คำสั่งนี้จะแปลง .xlsx เป็น Google Sheet อัตโนมัติ
        }

        media = MediaFileUpload(
            excel_file_path,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

        file = service.files().create(
            body=file_metadata,
            media_body=media,
            fields='id'
        ).execute()

        print(f"Successfully created Google Sheet: '{gsheet_name}' (ID: {file.get('id')})")
        return file.get('id')

    def create_daily_summary(self, ws, schedule):
        styles = self._setup_daily_summary_styles()
        ordered_pharmacists = self.get_ordered_employees()
        ws.cell(row=1, column=1, value='Pharmacist').fill = styles['header_fill']

        sorted_dates = sorted(schedule.index)

        for col, date in enumerate(sorted_dates, 2):
            cell = ws.cell(row=1, column=col, value=date.strftime('%d/%m'))
            cell.fill, cell.font = styles['header_fill'], styles['fonts']['header']
            if date.weekday() >= 5: cell.fill = styles['weekend_fill']
            if self.is_holiday(date): cell.fill = styles['holiday_fill']

        current_row = 2
        for pharmacist in ordered_pharmacists:
            if pharmacist not in self.pharmacists: continue
            ws.cell(row=current_row, column=1, value="").fill = styles['header_fill']
            ws.cell(row=current_row + 1, column=1, value=pharmacist).fill = styles['header_fill']
            ws.cell(row=current_row + 2, column=1, value="").fill = styles['header_fill']

            for col, date in enumerate(sorted_dates, 2):
                note_cell, cell1, cell2 = [ws.cell(row=current_row + r, column=col) for r in range(3)]
                all_cells = [note_cell, cell1, cell2]
                for cell in all_cells:
                    cell.border = styles['border']
                    cell.alignment = Alignment(horizontal="center", vertical="center")

                date_str = date.strftime('%Y-%m-%d')
                note_text = self.special_notes.get(pharmacist, {}).get(date_str)
                shifts = self.get_pharmacist_shifts(pharmacist, date, schedule)
                is_personal_holiday = date_str in self.pharmacists[pharmacist]['holidays']
                is_public_holiday_or_weekend = self.is_holiday(date) or date.weekday() >= 5

                # --- START OF LOGIC FIX ---
                # First, handle holiday or shift display
                if is_personal_holiday:
                    cell2.value = 'X'
                    cell1.value = None
                    for cell in all_cells:
                        cell.fill = styles['off_fill']
                else:
                    if len(shifts) > 0:
                        shift = shifts[0]
                        cell2.value = f"{int(self.shift_types[shift]['hours'])}N" if self.is_night_shift(shift) else int(self.shift_types[shift]['hours'])
                        prefix = next((p for p in styles['fills'] if shift.startswith(p)), None)
                        if prefix:
                            fill_color = styles['fills'][prefix]
                            cell2.fill, cell2.font = fill_color, styles['fonts'].get(prefix, Font(bold=True))
                            if len(shifts) == 1: cell1.fill = fill_color

                    if len(shifts) > 1:
                        shift = shifts[1]
                        cell1.value = f"{int(self.shift_types[shift]['hours'])}N" if self.is_night_shift(shift) else int(self.shift_types[shift]['hours'])
                        prefix = next((p for p in styles['fills'] if shift.startswith(p)), None)
                        if prefix: cell1.fill, cell1.font = styles['fills'][prefix], styles['fonts'].get(prefix, Font(bold=True))

                    if is_public_holiday_or_weekend:
                        note_cell.fill = styles['holiday_empty_fill']
                        if not shifts:
                            if not cell1.value: cell1.fill = styles['holiday_empty_fill']
                            if not cell2.value: cell2.fill = styles['holiday_empty_fill']

                # Second, ALWAYS apply the note if it exists. This overwrites nothing in note_cell
                # but ensures it is displayed regardless of holiday/shift status.
                if note_text:
                    note_cell.value = note_text
                    note_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                # --- END OF LOGIC FIX ---

            current_row += 3

        # ... (rest of the function is unchanged) ...
        total_row, unfilled_row = current_row + 1, current_row + 2
        ws.cell(row=total_row, column=1, value="Total Hours").fill = styles['header_fill']
        ws.cell(row=unfilled_row, column=1, value="Unfilled Shifts").fill = styles['header_fill']
        for col, date in enumerate(sorted_dates, 2):
            total_hours = sum(self.shift_types[st]['hours'] for st, p in schedule.loc[date].items() if p not in ['NO SHIFT', 'UNFILLED', 'UNASSIGNED'] and st in self.shift_types)
            unfilled_shifts = [st for st, p in schedule.loc[date].items() if p in ['UNFILLED', 'UNASSIGNED']]
            ws.cell(row=total_row, column=col, value=total_hours).border = styles['border']
            unfilled_cell = ws.cell(row=unfilled_row, column=col)
            unfilled_cell.border = styles['border']
            if unfilled_shifts:
                unfilled_cell.value, unfilled_cell.fill = "\n".join(unfilled_shifts), PatternFill(start_color='FFFFFF00', fill_type='solid')
            else:
                unfilled_cell.value = "0"
        ws.column_dimensions['A'].width = 25
        for col in range(2, len(schedule.index) + 3):
            ws.column_dimensions[get_column_letter(col)].width = 7

    def create_preference_score_summary(self, ws, schedule):
        header_fill = PatternFill(start_color='FFD3D3D3', end_color='FFD3D3D3', fill_type='solid')
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        bold_font = Font(bold=True)
        headers = ["Pharmacist", "Preference Score (%)", "Total Shifts Worked"]
        for col, header_text in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header_text)
            cell.fill, cell.font, cell.border = header_fill, bold_font, border
        preference_scores = self.calculate_pharmacist_preference_scores(schedule)
        pharmacist_list = sorted(self.pharmacists.keys())
        for row, pharmacist in enumerate(pharmacist_list, 2):
            total_shifts = sum(1 for date in schedule.index for p in schedule.loc[date] if p == pharmacist)
            score = preference_scores.get(pharmacist, 0)
            ws.cell(row=row, column=1, value=pharmacist).border = border
            score_cell = ws.cell(row=row, column=2, value=score)
            score_cell.border = border
            score_cell.number_format = '0.00"%"'
            ws.cell(row=row, column=3, value=total_shifts).border = border
        ws.column_dimensions['A'].width, ws.column_dimensions['B'].width, ws.column_dimensions['C'].width = 30, 25, 25

    def create_daily_summary_with_codes(self, ws, schedule):
        styles = self._setup_daily_summary_styles()
        ordered_pharmacists = self.get_ordered_employees()
        ws.cell(row=1, column=1, value='Pharmacist').fill = styles['header_fill']

        sorted_dates = sorted(schedule.index)

        for col, date in enumerate(sorted_dates, 2):
            cell = ws.cell(row=1, column=col, value=date.strftime('%d/%m'))
            cell.fill, cell.font = styles['header_fill'], styles['fonts']['header']
            if date.weekday() >= 5: cell.fill = styles['weekend_fill']
            if self.is_holiday(date): cell.fill = styles['holiday_fill']

        current_row = 2
        for pharmacist in ordered_pharmacists:
            if pharmacist not in self.pharmacists: continue
            ws.cell(row=current_row, column=1, value="").fill = styles['header_fill']
            ws.cell(row=current_row + 1, column=1, value=pharmacist).fill = styles['header_fill']
            ws.cell(row=current_row + 2, column=1, value="").fill = styles['header_fill']

            for col, date in enumerate(sorted_dates, 2):
                note_cell, cell1, cell2 = [ws.cell(row=current_row + r, column=col) for r in range(3)]
                all_cells = [note_cell, cell1, cell2]
                for cell in all_cells:
                    cell.border = styles['border']
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    if cell != note_cell: cell.font = Font(bold=True, size=9)

                date_str = date.strftime('%Y-%m-%d')
                note_text = self.special_notes.get(pharmacist, {}).get(date_str)
                shifts = self.get_pharmacist_shifts(pharmacist, date, schedule)
                is_personal_holiday = date_str in self.pharmacists[pharmacist]['holidays']
                is_public_holiday_or_weekend = self.is_holiday(date) or date.weekday() >= 5

                # --- START OF LOGIC FIX ---
                # First, handle holiday or shift display
                if is_personal_holiday:
                    cell2.value = 'OFF'
                    cell1.value = None
                    for cell in all_cells:
                        cell.fill = styles['off_fill']
                else:
                    if len(shifts) > 0:
                        shift_code, cell = shifts[0], cell2
                        cell.value = shift_code
                        prefix = next((p for p in styles['fills'] if shift_code.startswith(p)), None)
                        if prefix:
                            fill_color = styles['fills'][prefix]
                            cell.fill, cell.font = fill_color, styles['fonts'].get(prefix, styles['fonts']['default'])
                            if len(shifts) == 1: cell1.fill = fill_color

                    if len(shifts) > 1:
                        shift_code, cell = shifts[1], cell1
                        cell.value = shift_code
                        prefix = next((p for p in styles['fills'] if shift_code.startswith(p)), None)
                        if prefix: cell.fill, cell.font = styles['fills'][prefix], styles['fonts'].get(prefix, styles['fonts']['default'])

                    if is_public_holiday_or_weekend:
                        note_cell.fill = styles['holiday_empty_fill']
                        if not shifts:
                            if not cell1.value: cell1.fill = styles['holiday_empty_fill']
                            if not cell2.value: cell2.fill = styles['holiday_empty_fill']

                # Second, ALWAYS apply the note if it exists.
                if note_text:
                    note_cell.value = note_text
                    note_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                # --- END OF LOGIC FIX ---

            current_row += 3

        # ... (rest of the function is unchanged) ...
        total_row, unfilled_row = current_row + 1, current_row + 2
        ws.cell(row=total_row, column=1, value="Total Hours").fill = styles['header_fill']
        ws.cell(row=unfilled_row, column=1, value="Unfilled Shifts").fill = styles['header_fill']
        for col, date in enumerate(sorted_dates, 2):
            total_hours = sum(self.shift_types[st]['hours'] for st, p in schedule.loc[date].items() if p not in ['NO SHIFT', 'UNFILLED', 'UNASSIGNED'] and st in self.shift_types)
            unfilled_shifts = [st for st, p in schedule.loc[date].items() if p in ['UNFILLED', 'UNASSIGNED']]
            ws.cell(row=total_row, column=col, value=total_hours).border = styles['border']
            unfilled_cell = ws.cell(row=unfilled_row, column=col)
            unfilled_cell.border = styles['border']
            if unfilled_shifts:
                unfilled_cell.value, unfilled_cell.fill = "\n".join(unfilled_shifts), PatternFill(start_color='FFFFFF00', fill_type='solid')
            else:
                unfilled_cell.value = "0"
        ws.column_dimensions['A'].width = 25
        for col in range(2, len(schedule.index) + 2):
            ws.column_dimensions[get_column_letter(col)].width = 15

    def calculate_pharmacist_preference_scores(self, schedule):
        scores = {}
        MAX_POINTS_PER_SHIFT = 8
        for pharmacist in self.pharmacists:
            total_achieved_points = 0
            total_shifts_worked = 0
            for date in schedule.index:
                for shift_type, assigned_pharm in schedule.loc[date].items():
                    if assigned_pharm == pharmacist:
                        total_shifts_worked += 1
                        rank = self.get_preference_score(pharmacist, shift_type)
                        points = max(0, 9 - rank)
                        total_achieved_points += points
            if total_shifts_worked == 0:
                scores[pharmacist] = 0
            else:
                max_possible_points = total_shifts_worked * MAX_POINTS_PER_SHIFT
                if max_possible_points == 0: # Avoid division by zero
                    scores[pharmacist] = 0
                else:
                    percentage_score = (total_achieved_points / max_possible_points) * 100
                    scores[pharmacist] = percentage_score
        return scores

    # =========================================================================
    # ========== FUNCTIONS ADDED FOR SPECIFIC DATE SCHEDULING =================
    # =========================================================================

    def _pre_check_staffing_for_dates(self, dates_to_schedule):
        """
        New pre-check function that works with a list of specific dates.
        It does not modify the original _pre_check_staffing_levels function.
        """
        print("\nRunning pre-check for staffing levels for specific dates...")
        all_ok = True
        for date in dates_to_schedule:
            available_pharmacists_count = sum(1 for p_name, p_info in self.pharmacists.items()
                                              if date.strftime('%Y-%m-%d') not in p_info['holidays'])
            required_shifts_base = sum(1 for st in self.shift_types
                                       if self.is_shift_available_on_date(st, date))
            total_required_shifts_with_buffer = required_shifts_base + 3

            if available_pharmacists_count < total_required_shifts_with_buffer:
                all_ok = False
                self.problem_days.add(date)
                print(f"WARNING: Potential shortage on {date.strftime('%Y-%m-%d')}. "
                      f"Available Pharmacists: {available_pharmacists_count}, "
                      f"Required Shifts (with +3 buffer): {total_required_shifts_with_buffer}")
        if all_ok:
            print("Pre-check complete. All specified dates have sufficient staffing levels.")
        else:
            print("Pre-check complete. Identified specified dates with potential staff shortages.")
        return not all_ok

    def calculate_weekend_off_variance_for_dates(self, schedule):
        """
        New variance calculation for a specific set of dates present in a schedule.
        """
        weekend_off_counts = {p: 0 for p in self.pharmacists}
        for date in schedule.index:
            if date.weekday() >= 5: # 5 is Saturday, 6 is Sunday
                working_on_weekend = {schedule.loc[date, shift] for shift in schedule.columns if schedule.loc[date, shift] not in ['NO SHIFT', 'UNFILLED', 'UNASSIGNED']}
                for p_name in self.pharmacists:
                    if p_name not in working_on_weekend:
                        weekend_off_counts[p_name] += 1
        if len(weekend_off_counts) > 1:
            return np.var(list(weekend_off_counts.values()))
        return 0

    def calculate_metrics_for_schedule(self, schedule):
        """
        New metrics calculation function that works directly with a schedule DataFrame.
        This avoids the need to pass year and month.
        """
        hours = {p: self.calculate_total_hours(p, schedule) for p in self.pharmacists}
        night_counts = {p: self.pharmacists[p]['night_shift_count'] for p in self.pharmacists}
        weekend_off_var = self.calculate_weekend_off_variance_for_dates(schedule)
        hour_penalty = self._get_hour_imbalance_penalty(hours)

        pref_percentages = self.calculate_pharmacist_preference_scores(schedule)
        pref_variance = np.var(list(pref_percentages.values())) if len(pref_percentages) > 1 else 0

        weekend_min_off_violations = self.calculate_weekend_min_off_violations(schedule)
        metrics = {
            'hour_imbalance_penalty': hour_penalty,
            'night_variance': np.var(list(night_counts.values())) if night_counts else 0,
            'preference_score': sum(self.calculate_preference_penalty(p, schedule) for p in self.pharmacists),
            'preference_variance': pref_variance,
            'weekend_off_variance': weekend_off_var,
            'weekend_min_off_shortfall': sum(weekend_min_off_violations.values()),
            'month_segment_variance': self.calculate_month_segment_variance(schedule),
        }
        if len(hours) > 1 and len(hours.values()) > 1:
            metrics['hour_diff_for_logging'] = stdev(hours.values())
        else:
            metrics['hour_diff_for_logging'] = 0
        return metrics

    def generate_schedule_for_dates(self, dates_to_schedule, iteration_num=1):
        """
        New schedule generator that operates on a specific list of dates.
        """
        schedule_dict = {date: {shift: 'NO SHIFT' for shift in self.shift_types} for date in dates_to_schedule}
        pharmacist_hours = {p: 0 for p in self.pharmacists}
        pharmacist_consecutive_days = {p: 0 for p in self.pharmacists}

        shuffled_shifts = list(self.shift_types.keys())
        random.shuffle(shuffled_shifts)
        shuffled_pharmacists = list(self.pharmacists.keys())
        random.shuffle(shuffled_pharmacists)

        for pharmacist in self.pharmacists:
            self.pharmacists[pharmacist]['night_shift_count'] = 0
            self.pharmacists[pharmacist]['mixing_shift_count'] = 0
            self.pharmacists[pharmacist]['category_counts'] = {'Mixing': 0, 'Night': 0}

        for pharmacist, assignments in self.pre_assignments.items():
            if pharmacist not in self.pharmacists: continue
            for date_str, shift_types in assignments.items():
                date = pd.to_datetime(date_str).to_pydatetime().date()
                # Find the matching datetime object in the list
                matching_date = next((dt for dt in dates_to_schedule if dt.date() == date), None)
                if matching_date is None: continue
                for shift_type in shift_types:
                    if shift_type in self.shift_types:
                        schedule_dict[matching_date][shift_type] = pharmacist
                        self._update_shift_counts(pharmacist, shift_type)
                        pharmacist_hours[pharmacist] += self.shift_types[shift_type]['hours']

        all_dates = sorted(list(dates_to_schedule))
        problem_dates_sorted = sorted([d for d in all_dates if d in self.problem_days])
        other_dates_sorted = sorted([d for d in all_dates if d not in self.problem_days])
        processing_order_dates = problem_dates_sorted + other_dates_sorted
        unfilled_info = {'problem_days': [], 'other_days': []}

        night_shifts_ordered = [s for s in shuffled_shifts if self.is_night_shift(s)]
        mixing_shifts_ordered = [s for s in shuffled_shifts if s.startswith('C8') and not self.is_night_shift(s)]
        care_shifts_ordered = [s for s in shuffled_shifts if s.startswith('Care') and not self.is_night_shift(s) and not s.startswith('C8')]
        other_shifts_ordered = [s for s in shuffled_shifts if not self.is_night_shift(s) and not s.startswith('C8') and not s.startswith('Care')]
        standard_shift_order = night_shifts_ordered + mixing_shifts_ordered + care_shifts_ordered + other_shifts_ordered
        problem_day_shift_order = mixing_shifts_ordered + care_shifts_ordered + night_shifts_ordered + other_shifts_ordered

        for date in tqdm(processing_order_dates, desc=f"Building Schedule (Iteration {iteration_num})", leave=False):
            previous_date = date - timedelta(days=1)
            # We can only check consecutive days if the previous day was also in our scheduling set
            if previous_date in schedule_dict:
                pharmacists_working_yesterday = {p for p in schedule_dict[previous_date].values() if p in self.pharmacists}
                for p_name in self.pharmacists:
                    if p_name in pharmacists_working_yesterday:
                        pharmacist_consecutive_days[p_name] += 1
                    else:
                        pharmacist_consecutive_days[p_name] = 0
            else: # Reset if previous day wasn't scheduled
                 for p_name in self.pharmacists:
                    pharmacist_consecutive_days[p_name] = 0

            is_day_before_problem_day = (date + timedelta(days=1)) in self.problem_days
            shifts_to_process = problem_day_shift_order if date in self.problem_days else standard_shift_order

            for shift_type in shifts_to_process:
                if schedule_dict[date][shift_type] != 'NO SHIFT' or not self.is_shift_available_on_date(shift_type, date):
                    continue
                available = self._get_available_pharmacists_optimized(shuffled_pharmacists, date, shift_type, schedule_dict, pharmacist_hours, pharmacist_consecutive_days)
                if available:
                    chosen = self._select_best_pharmacist(available, shift_type, date, is_day_before_problem_day)
                    pharmacist_to_assign = chosen['name']
                    schedule_dict[date][shift_type] = pharmacist_to_assign
                    self._update_shift_counts(pharmacist_to_assign, shift_type)
                    pharmacist_hours[pharmacist_to_assign] += self.shift_types[shift_type]['hours']
                else:
                    schedule_dict[date][shift_type] = 'UNFILLED'
                    if date in self.problem_days:
                        unfilled_info['problem_days'].append((date, shift_type))
                    else:
                        unfilled_info['other_days'].append((date, shift_type))

        final_schedule = pd.DataFrame.from_dict(schedule_dict, orient='index')
        final_schedule.fillna('NO SHIFT', inplace=True)
        return final_schedule, unfilled_info

    def optimize_schedule_for_dates(self, dates_to_schedule, iterations=10):
        """
        New optimizer for scheduling a specific list of dates.
        """
        best_schedule = None
        best_metrics = {'unfilled_problem_shifts': float('inf'), 'preference_score': float('inf')}
        best_unfilled_info = {}

        # Use the new pre-check function
        self._pre_check_staffing_for_dates(dates_to_schedule)

        print(f"\nStarting optimization for {len(dates_to_schedule)} specific dates with {iterations} iterations...")

        for i in range(iterations):
            print(f"\n--- Iteration {i+1}/{iterations} ---")
            current_schedule, unfilled_info = self.generate_schedule_for_dates(dates_to_schedule, iteration_num=i+1)

            # Use the new metrics calculation function
            metrics = self.calculate_metrics_for_schedule(current_schedule)
            metrics['unfilled_problem_shifts'] = len(unfilled_info['problem_days']) + len(unfilled_info['other_days'])

            print(f"Iteration Results -> "
                  f"Unfilled Shifts: {metrics['unfilled_problem_shifts']} | "
                  f"Hour SD: {metrics.get('hour_diff_for_logging', 0):.2f} | "
                  f"Night Var: {metrics.get('night_variance', 0):.2f} | "
                  f"Weekend Shortfall: {metrics.get('weekend_min_off_shortfall', 0)} | "
                  f"Month Segment Var: {metrics.get('month_segment_variance', 0):.2f} | "
                  f"Pref Penalty: {metrics.get('preference_score', 0):.1f}")

            if best_schedule is None or self.is_schedule_better(metrics, best_metrics):
                best_schedule = current_schedule.copy()
                best_metrics = metrics.copy()
                best_unfilled_info = unfilled_info.copy()
                print("*** Found a more balanced schedule! ***")

        if best_schedule is not None:
            print("\nOptimization complete!\nFinal metrics for the best schedule found:")
            print(f"Unfilled Shifts: {best_metrics.get('unfilled_problem_shifts', 0)} | "
                  f"Hour SD: {best_metrics.get('hour_diff_for_logging', 0):.2f} | "
                  f"Night Var: {best_metrics.get('night_variance', 0):.2f} | "
                  f"Weekend Shortfall: {best_metrics.get('weekend_min_off_shortfall', 0)} | "
                  f"Month Segment Var: {best_metrics.get('month_segment_variance', 0):.2f} | "
                  f"Pref Penalty: {best_metrics.get('preference_score', 0):.1f}")
        else:
            print("\nOptimization failed to find any valid schedule.")

        return best_schedule, best_unfilled_info

# --- Main execution block ---


# ==============================================================================
# Streamlit UI
# ==============================================================================

import os
import tempfile
from datetime import datetime

import numpy as np
import pandas as pd
import streamlit as st



st.set_page_config(
    page_title="CRA Pharmacy Shift Scheduler",
    page_icon="💊",
    layout="wide",
    initial_sidebar_state="expanded",
)


CUSTOM_CSS = """
<style>
    .main {
        background: linear-gradient(135deg, #f7f9fc 0%, #eef3fb 100%);
    }
    .block-container {
        padding-top: 1.4rem;
        padding-bottom: 2rem;
        max-width: 1500px;
    }
    .hero-card {
        padding: 1.35rem 1.6rem;
        border-radius: 24px;
        background: linear-gradient(135deg, #002060 0%, #003087 62%, #0b4bb3 100%);
        color: white;
        box-shadow: 0 16px 40px rgba(0, 32, 96, 0.18);
        margin-bottom: 1rem;
    }
    .hero-title {
        font-size: 2rem;
        font-weight: 800;
        margin-bottom: 0.15rem;
        letter-spacing: -0.02em;
    }
    .hero-subtitle {
        color: rgba(255, 255, 255, 0.86);
        font-size: 0.98rem;
        margin-bottom: 0;
    }
    .metric-card {
        padding: 1rem;
        border-radius: 20px;
        background: white;
        border: 1px solid rgba(0, 32, 96, 0.08);
        box-shadow: 0 10px 28px rgba(15, 23, 42, 0.06);
    }
    .metric-label {
        font-size: 0.78rem;
        color: #64748b;
        font-weight: 700;
        text-transform: uppercase;
        letter-spacing: 0.04em;
    }
    .metric-value {
        font-size: 1.55rem;
        font-weight: 800;
        color: #002060;
        margin-top: 0.25rem;
    }
    .section-title {
        font-size: 1.2rem;
        font-weight: 800;
        color: #002060;
        margin: 0.7rem 0 0.45rem 0;
    }
    div[data-testid="stSidebar"] {
        background: #ffffff;
        border-right: 1px solid rgba(0, 32, 96, 0.08);
    }
    div[data-testid="stSidebar"] h1, div[data-testid="stSidebar"] h2, div[data-testid="stSidebar"] h3 {
        color: #002060;
    }
    .stButton > button {
        width: 100%;
        border-radius: 16px;
        background: linear-gradient(135deg, #f5921e 0%, #ffad42 100%);
        color: white;
        border: 0;
        font-weight: 800;
        padding: 0.7rem 1rem;
        box-shadow: 0 10px 24px rgba(245, 146, 30, 0.22);
    }
    .stDownloadButton > button {
        border-radius: 14px;
        font-weight: 700;
    }
    .warning-box {
        border-left: 5px solid #f5921e;
        background: #fff7ed;
        border-radius: 14px;
        padding: 0.85rem 1rem;
        color: #7c2d12;
        margin: 0.8rem 0;
    }
</style>
"""

st.markdown(CUSTOM_CSS, unsafe_allow_html=True)


@st.cache_data(show_spinner=False)
def read_excel_sheet_names(file_bytes: bytes):
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        tmp.write(file_bytes)
        tmp_path = tmp.name
    try:
        xls = pd.ExcelFile(tmp_path)
        return xls.sheet_names
    finally:
        try:
            os.remove(tmp_path)
        except OSError:
            pass


def save_uploaded_file(uploaded_file):
    suffix = ".xlsx"
    with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
        tmp.write(uploaded_file.getvalue())
        return tmp.name


def make_hour_summary(scheduler: PharmacistScheduler, schedule: pd.DataFrame) -> pd.DataFrame:
    rows = []
    for name in scheduler.get_ordered_employees():
        if name not in scheduler.pharmacists:
            continue
        total_hours = scheduler.calculate_total_hours(name, schedule)
        total_shifts = int((schedule == name).sum().sum())
        night_shifts = int(sum(
            1
            for date in schedule.index
            for shift, assigned in schedule.loc[date].items()
            if assigned == name and scheduler.is_night_shift(shift)
        ))
        weekend_days = int(sum(
            1
            for date in schedule.index
            if date.weekday() >= 5 and name in schedule.loc[date].values
        ))
        rows.append({
            "Name": name,
            "Total Hours": total_hours,
            "Total Shifts": total_shifts,
            "Night Shifts": night_shifts,
            "Weekend Days Worked": weekend_days,
            "Max Hours": scheduler.pharmacists[name].get("max_hours", 250),
        })
    return pd.DataFrame(rows)


def make_unfilled_df(unfilled_info) -> pd.DataFrame:
    rows = []
    for group_name, items in unfilled_info.items():
        for date, shift in items:
            rows.append({
                "Group": group_name,
                "Date": pd.to_datetime(date).strftime("%Y-%m-%d"),
                "Shift": shift,
            })
    return pd.DataFrame(rows)


def make_daily_long_df(schedule: pd.DataFrame) -> pd.DataFrame:
    rows = []
    for date in schedule.index:
        for shift, assigned in schedule.loc[date].items():
            if assigned not in ["NO SHIFT", "UNFILLED", "UNASSIGNED"]:
                rows.append({
                    "Date": pd.to_datetime(date).strftime("%Y-%m-%d"),
                    "Day": pd.to_datetime(date).strftime("%a"),
                    "Shift": shift,
                    "Assigned": assigned,
                })
    return pd.DataFrame(rows)


def render_metric_card(label: str, value: str):
    st.markdown(
        f"""
        <div class="metric-card">
            <div class="metric-label">{label}</div>
            <div class="metric-value">{value}</div>
        </div>
        """,
        unsafe_allow_html=True,
    )


def export_schedule_to_bytes(scheduler, schedule, unfilled_info, enable_run_log):
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        output_path = tmp.name
    try:
        scheduler.export_to_excel(
            schedule=schedule,
            unfilled_info=unfilled_info,
            filename=output_path,
            enable_run_log=enable_run_log,
        )
        with open(output_path, "rb") as f:
            return f.read()
    finally:
        try:
            os.remove(output_path)
        except OSError:
            pass


st.markdown(
    """
    <div class="hero-card">
        <div class="hero-title">💊 CRA Pharmacy Shift Scheduler</div>
        <p class="hero-subtitle">Upload Excel → run schedule optimization → review dashboard → download formatted Excel output.</p>
    </div>
    """,
    unsafe_allow_html=True,
)

with st.sidebar:
    st.title("⚙️ Settings")
    uploaded_file = st.file_uploader(
        "Upload scheduler Excel file",
        type=["xlsx"],
        help="ต้องมี sheet เช่น employee, Shifts, Departments, PreAssignments ตามโครงสร้างโปรแกรมเดิม",
    )

    sheet_names = []
    if uploaded_file is not None:
        try:
            sheet_names = read_excel_sheet_names(uploaded_file.getvalue())
        except Exception as exc:
            st.error(f"อ่านรายชื่อชีตไม่สำเร็จ: {exc}")

    default_employee_sheet = "employee"
    if sheet_names and default_employee_sheet not in sheet_names:
        default_employee_sheet = sheet_names[0]

    employee_sheet_name = st.selectbox(
        "Employee sheet",
        options=sheet_names if sheet_names else ["employee"],
        index=(sheet_names.index(default_employee_sheet) if sheet_names and default_employee_sheet in sheet_names else 0),
    )

    staff_type = st.selectbox(
        "Staff type",
        options=["เภสัชกร", "ผู้ช่วยเภสัชกร", "อื่น ๆ"],
        index=0,
    )

    current_year = datetime.now().year
    year = st.number_input("Year", min_value=2000, max_value=2100, value=max(current_year, 2026), step=1)
    month = st.number_input("Month", min_value=1, max_value=12, value=6, step=1)
    iterations = st.number_input("Iterations", min_value=1, max_value=500, value=20, step=1)

    st.divider()
    true_random_override = st.toggle(
        "True Random Override",
        value=False,
        help="สุ่มโดยยังคุม skill, leave, ไม่ซ้ำเวรในวันเดียว และ hour balance",
    )
    enable_run_log = st.toggle("Export Run Log sheet", value=False)

    run_button = st.button("🚀 Run Scheduler", type="primary")

if uploaded_file is None:
    st.info("อัปโหลดไฟล์ Excel ก่อน แล้วตั้งค่าเดือน/ปีเพื่อเริ่มจัดเวร")
    st.markdown(
        """
        <div class="warning-box">
        <b>Required core sheets:</b> employee, Shifts, Departments, PreAssignments<br>
        Optional sheets: Skill subset, HistoricalScores, SpecialNotes, ShiftLimits, MinShiftRequirements
        </div>
        """,
        unsafe_allow_html=True,
    )
    st.stop()

if run_button:
    input_path = save_uploaded_file(uploaded_file)

    try:
        with st.spinner("กำลังจัดเวรและคำนวณผลลัพธ์..."):
            scheduler = PharmacistScheduler(
                input_path,
                employee_sheet_name=employee_sheet_name,
                staff_type=staff_type,
            )

            schedule, unfilled_info = scheduler.optimize_schedule(
                year=int(year),
                month=int(month),
                iterations=int(iterations),
                true_random_override=bool(true_random_override),
                enable_run_log=bool(enable_run_log),
            )

        st.session_state["scheduler"] = scheduler
        st.session_state["schedule"] = schedule
        st.session_state["unfilled_info"] = unfilled_info
        st.session_state["enable_run_log"] = enable_run_log
        st.session_state["year"] = int(year)
        st.session_state["month"] = int(month)
        st.session_state["mode"] = "TRUE_RANDOM" if true_random_override else "OPTIMIZED"
        st.success("จัดเวรสำเร็จแล้ว")

    except Exception as exc:
        st.error("รันโปรแกรมไม่สำเร็จ")
        st.exception(exc)
    finally:
        try:
            os.remove(input_path)
        except OSError:
            pass

if "schedule" not in st.session_state:
    st.stop()

scheduler = st.session_state["scheduler"]
schedule = st.session_state["schedule"]
unfilled_info = st.session_state["unfilled_info"]
enable_run_log = st.session_state["enable_run_log"]
mode = st.session_state["mode"]

hour_summary = make_hour_summary(scheduler, schedule)
unfilled_df = make_unfilled_df(unfilled_info)
daily_long_df = make_daily_long_df(schedule)

st.markdown('<div class="section-title">Dashboard Summary</div>', unsafe_allow_html=True)
col1, col2, col3, col4, col5 = st.columns(5)

with col1:
    render_metric_card("Mode", mode)
with col2:
    render_metric_card("Staff", f"{len(scheduler.pharmacists):,}")
with col3:
    total_assigned = int((~schedule.isin(["NO SHIFT", "UNFILLED", "UNASSIGNED"])).sum().sum())
    render_metric_card("Assigned Shifts", f"{total_assigned:,}")
with col4:
    total_unfilled = len(unfilled_df)
    render_metric_card("Unfilled", f"{total_unfilled:,}")
with col5:
    if not hour_summary.empty:
        hour_range = hour_summary["Total Hours"].max() - hour_summary["Total Hours"].min()
    else:
        hour_range = 0
    render_metric_card("Hour Range", f"{hour_range:.1f}")

if not hour_summary.empty:
    hcol1, hcol2 = st.columns([1.2, 1])
    with hcol1:
        st.markdown('<div class="section-title">Total Hours by Staff</div>', unsafe_allow_html=True)
        chart_df = hour_summary.set_index("Name")[["Total Hours"]]
        st.bar_chart(chart_df, use_container_width=True)
    with hcol2:
        st.markdown('<div class="section-title">Fairness Snapshot</div>', unsafe_allow_html=True)
        st.dataframe(
            hour_summary.sort_values("Total Hours", ascending=True),
            use_container_width=True,
            hide_index=True,
        )

st.markdown('<div class="section-title">Schedule Result</div>', unsafe_allow_html=True)
tab1, tab2, tab3, tab4, tab5 = st.tabs([
    "📅 Calendar Matrix",
    "📋 Daily Long View",
    "👥 Staff Summary",
    "⚠️ Unfilled",
    "🧾 Run Logs",
])

with tab1:
    display_schedule = schedule.copy()
    display_schedule.index = [pd.to_datetime(d).strftime("%Y-%m-%d %a") for d in display_schedule.index]
    st.dataframe(display_schedule, use_container_width=True, height=620)

with tab2:
    if daily_long_df.empty:
        st.info("ยังไม่มีรายการเวรที่ถูก assign")
    else:
        selected_name = st.selectbox(
            "Filter by staff",
            options=["All"] + sorted(daily_long_df["Assigned"].dropna().unique().tolist()),
        )
        filtered_df = daily_long_df if selected_name == "All" else daily_long_df[daily_long_df["Assigned"] == selected_name]
        st.dataframe(filtered_df, use_container_width=True, hide_index=True, height=620)

with tab3:
    st.dataframe(hour_summary, use_container_width=True, hide_index=True, height=620)

with tab4:
    if unfilled_df.empty:
        st.success("ไม่มี UNFILLED shift")
    else:
        st.warning(f"พบ UNFILLED ทั้งหมด {len(unfilled_df)} รายการ")
        st.dataframe(unfilled_df, use_container_width=True, hide_index=True, height=500)

with tab5:
    if scheduler.run_logs:
        st.dataframe(pd.DataFrame(scheduler.run_logs), use_container_width=True, hide_index=True, height=620)
    else:
        st.info("ยังไม่มี run logs หรือไม่ได้เปิด Export Run Log")

st.markdown('<div class="section-title">Download Output</div>', unsafe_allow_html=True)
excel_bytes = export_schedule_to_bytes(scheduler, schedule, unfilled_info, enable_run_log)
file_name = f"{mode}_Schedule_{st.session_state['year']}_{st.session_state['month']:02d}.xlsx"
st.download_button(
    label="⬇️ Download formatted Excel schedule",
    data=excel_bytes,
    file_name=file_name,
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    use_container_width=True,
)
