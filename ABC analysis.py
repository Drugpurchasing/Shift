import streamlit as st
import pandas as pd
import io
import numpy as np
from openpyxl.styles import Alignment, Font
import time

# ==============================================================================
# Page Configuration (ต้องเป็นคำสั่งแรก)
# ==============================================================================
st.set_page_config(
    page_title="ABC Analysis Tool",
    page_icon="🔤",
    layout="wide"
)

# ==============================================================================
# Functions
# ==============================================================================
def process_abc_analysis(inventory_files, master_file_url, progress_bar):
    """
    Performs ABC analysis based on Net Consumption Value across different storage locations.

    Args:
        inventory_files: List of file objects containing consumption data.
        master_file_url: URL string for the Drug Master file.
        progress_bar: Streamlit progress bar object.

    Returns:
        Bytes of the classified Excel report, or None on error.
    """
    progress_bar.progress(10, text="[10%] กำลังรวมและจัดเตรียมข้อมูลการใช้งาน...")
    try:
        # Load and consolidate inventory/usage files
        all_dfs = [pd.read_excel(fp) for fp in inventory_files]
        consolidated_df = pd.concat(all_dfs, ignore_index=True)

        # Load Drug Master from URL
        master_df = pd.read_excel(master_file_url, sheet_name='Drug master', usecols=['Material', 'Drug group'])
        master_df['Material'] = master_df['Material'].astype(str)

    except Exception as e:
        st.error(f"เกิดข้อผิดพลาดในการโหลดไฟล์: {e}")
        return None

    progress_bar.progress(30, text="[30%] กำลังคำนวณมูลค่าการใช้งานรายเดือน...")
    df = consolidated_df
    df['Posting Date'] = pd.to_datetime(df['Posting Date'], errors='coerce')
    df.dropna(subset=['Posting Date'], inplace=True)
    df['MonthYear'] = df['Posting Date'].dt.to_period('M')
    df['Amt.in Loc.Cur.'] = pd.to_numeric(df['Amt.in Loc.Cur.'], errors='coerce').fillna(0)
    df['Quantity'] = pd.to_numeric(df['Quantity'], errors='coerce').fillna(0)
    df['Material'] = df['Material'].astype(str)

    # Aggregate monthly consumption data
    monthly_data = df.groupby(['Material', 'Material description', 'Storage location', 'MonthYear']).agg(
        MonthlyNetConsumption=('Amt.in Loc.Cur.', 'sum'),
        MonthlyNetQuantity=('Quantity', 'sum')
    ).reset_index()

    # Pivot for monthly quantities (to be included in the final report)
    monthly_qty_pivot = monthly_data.pivot_table(
        index=['Material', 'Material description', 'Storage location'],
        columns='MonthYear',
        values='MonthlyNetQuantity',
        fill_value=0
    )
    monthly_qty_pivot.columns = [f"Qty_{str(col)}" for col in monthly_qty_pivot.columns]
    monthly_qty_pivot = monthly_qty_pivot.abs()

    # Aggregate total and average data
    final_agg = monthly_data.groupby(['Material', 'Material description', 'Storage location']).agg(
        AvgMonthlyNetQuantity=('MonthlyNetQuantity', 'mean'),
        TotalNetConsumption=('MonthlyNetConsumption', 'sum')
    ).reset_index()

    final_agg['AvgMonthlyNetQuantity'] = final_agg['AvgMonthlyNetQuantity'].abs()

    # Merge monthly pivot back
    final_agg = pd.merge(final_agg, monthly_qty_pivot, on=['Material', 'Material description', 'Storage location'], how='left')
    final_agg['NetConsumptionValue'] = final_agg['TotalNetConsumption'].abs()
    abc_data_no_class = final_agg[final_agg['NetConsumptionValue'] > 0].copy()

    if abc_data_no_class.empty:
        st.warning("ไม่พบข้อมูลการใช้งาน (consumption data) ที่มีมูลค่ามากกว่า 0")
        return None

    # Merge with master data for drug group
    abc_data_no_class = pd.merge(abc_data_no_class, master_df, on='Material', how='left')
    abc_data_no_class['Drug group'].fillna('N/A', inplace=True)

    progress_bar.progress(60, text="[60%] กำลังจัดแบ่งกลุ่ม ABC ตามคลัง...")
    all_locations_classified = []

    # ABC Classification per Storage Location
    for location in abc_data_no_class['Storage location'].unique():
        loc_df = abc_data_no_class[abc_data_no_class['Storage location'] == location].copy()
        total_value_loc = loc_df['NetConsumptionValue'].sum()
        loc_df = loc_df.sort_values(by='NetConsumptionValue', ascending=False).reset_index(drop=True)
        loc_df['PercentageValue'] = loc_df['NetConsumptionValue'] / total_value_loc if total_value_loc > 0 else 0
        loc_df['CumulativePercentage'] = loc_df['PercentageValue'].cumsum()

        def assign_abc_class(cum_perc):
            if cum_perc <= 0.80:
                return 'A'
            elif cum_perc <= 0.95:
                return 'B'
            else:
                return 'C'

        loc_df['ABC_Class'] = loc_df['CumulativePercentage'].apply(assign_abc_class)
        all_locations_classified.append(loc_df)

    final_results = pd.concat(all_locations_classified)

    progress_bar.progress(80, text="[80%] กำลังสร้างชีตสรุปและจัดรูปแบบ Excel...")
    output_buffer = io.BytesIO()

    # Function to apply specific Excel formats and hide columns
    def apply_formats_and_hide_cols(writer, sheet_name, df):
        worksheet = writer.sheets[sheet_name]
        center_align = Alignment(horizontal='center', vertical='center')
        # Define number formats
        col_map = {'AvgMonthlyNetQuantity': '#,##0', 'NetConsumptionValue': '#,##0.00', 'PercentageValue': '0.00%',
                   'CumulativePercentage': '0.00%'}
        for col in df.columns:
            if isinstance(col, str) and col.startswith('Qty_'): col_map[col] = '#,##0'

        # Map column names to Excel column letters
        col_letters = {col_name: chr(65 + i) for i, col_name in enumerate(df.columns)}

        # Apply number formats
        for col_name, num_format in col_map.items():
            if col_name in col_letters:
                col_letter = col_letters[col_name]
                for row in range(2, worksheet.max_row + 1):
                    worksheet[f'{col_letter}{row}'].number_format = num_format

        # Apply center alignment for ABC_Class
        for row in range(2, worksheet.max_row + 1):
            worksheet[f'{col_letters["ABC_Class"]}{row}'].alignment = center_align

        # Hide monthly quantity columns
        for col_name in df.columns:
            if isinstance(col_name, str) and col_name.startswith('Qty_'):
                worksheet.column_dimensions[col_letters[col_name]].hidden = True

        # Auto-adjust column width for visible columns
        for col in worksheet.columns:
            if not worksheet.column_dimensions[col[0].column_letter].hidden:
                max_length = max(len(str(cell.value)) for cell in col if cell.value)
                worksheet.column_dimensions[col[0].column_letter].width = max_length + 2

    with pd.ExcelWriter(output_buffer, engine='openpyxl') as writer:
        # --- Executive Summary Sheet ---
        worksheet = writer.book.create_sheet("Executive Summary", 0)
        writer.sheets['Executive Summary'] = worksheet
        current_row = 1

        # 1. Summary ABC Count
        summary_abc_count = final_results.groupby(['Storage location', 'ABC_Class']).size().unstack(fill_value=0)
        for c in ['A', 'B', 'C']:
            if c not in summary_abc_count: summary_abc_count[c] = 0
        summary_abc_count = summary_abc_count[['A', 'B', 'C']]
        summary_abc_count['Total'] = summary_abc_count.sum(axis=1)
        summary_abc_count.loc['Total'] = summary_abc_count.sum()

        worksheet.cell(row=current_row, column=1, value='สรุปจำนวนรายการ A, B, C ในแต่ละคลัง').font = Font(bold=True)
        current_row += 1
        summary_abc_count.to_excel(writer, sheet_name='Executive Summary', startrow=current_row, startcol=0)
        current_row += summary_abc_count.shape[0] + 3

        # 2. Top 3 Drug Groups
        worksheet.cell(row=current_row, column=1,
                       value='กลุ่มยา (Drug Group) ที่มีมูลค่าการใช้งานสูงสุด 3 อันดับแรก (แยกตามคลัง)').font = Font(bold=True)
        current_row += 1
        top_groups = final_results.groupby('Storage location').apply(
            lambda x: x.groupby('Drug group')['NetConsumptionValue'].sum().nlargest(3)).reset_index()
        top_groups['NetConsumptionValue'] = top_groups['NetConsumptionValue'].map('{:,.2f}'.format)
        top_groups.to_excel(writer, sheet_name='Executive Summary', startrow=current_row, startcol=0, index=False)
        current_row += top_groups.shape[0] + 3

        # 3. Top 5 Items
        worksheet.cell(row=current_row, column=1,
                       value='รายการยาที่มีมูลค่าการใช้งานสูงสุด 5 อันดับแรก (แยกตามคลัง)').font = Font(bold=True)
        current_row += 1
        top_items = final_results.groupby('Storage location').apply(
            lambda x: x.groupby(['Material', 'Material description'])['NetConsumptionValue'].sum().nlargest(5)).reset_index()
        top_items['NetConsumptionValue'] = top_items['NetConsumptionValue'].map('{:,.2f}'.format)
        top_items.to_excel(writer, sheet_name='Executive Summary', startrow=current_row, startcol=0, index=False)

        # --- Detail Sheets per Storage Location ---
        for location in final_results['Storage location'].unique():
            sheet_df = final_results[final_results['Storage location'] == location].copy()
            sheet_name = f'SLoc_{location}'
            monthly_cols = sorted([col for col in sheet_df.columns if isinstance(col, str) and col.startswith('Qty_')])
            output_columns = ['Material', 'Material description', 'Storage location'] + monthly_cols + [
                'AvgMonthlyNetQuantity', 'NetConsumptionValue', 'PercentageValue', 'CumulativePercentage', 'ABC_Class',
                'Drug group']
            sheet_df = sheet_df[output_columns]
            sheet_df.to_excel(writer, sheet_name=sheet_name, index=False)
            apply_formats_and_hide_cols(writer, sheet_name, sheet_df)

    progress_bar.progress(95, text="[95%] กำลังจัดเตรียมการดาวน์โหลด...")
    return output_buffer.getvalue()

# ==============================================================================
# STREAMLIT USER INTERFACE (UI) for Standalone ABC Analysis
# ==============================================================================
st.title("🔤 ABC Analysis Tool (เครื่องมือวิเคราะห์ ABC)")
st.markdown("---")
st.markdown("เครื่องมือนี้ใช้สำหรับจัดกลุ่มรายการยาตามมูลค่าการใช้งาน (Net Consumption Value) แยกตามคลังสินค้า")

st.info("""
    **ขั้นตอนการใช้งาน:**
    1. **อัปโหลดไฟล์ข้อมูลการใช้งาน:** เลือกไฟล์ Excel (.xls, .xlsx) ที่มีข้อมูลการเบิกจ่ายยา/การใช้งาน (เลือกหลายไฟล์ได้)
    2. **กดปุ่ม 'เริ่มการวิเคราะห์ ABC'** (ไฟล์ Drug Master จะถูกดึงมาจากระบบโดยอัตโนมัติ)
""")

inventory_files = st.file_uploader("1. อัปโหลดไฟล์ข้อมูลการใช้งาน (Consumption Files)", type=["xlsx", "xls"],
                                     accept_multiple_files=True, key="abc_inventory_uploader")

# URL for the Drug Master file is now hardcoded
master_file_url = "https://docs.google.com/spreadsheets/d/e/2PACX-1vQJpIKf_q4h4h1VEIM0tT1MlMvoEw1PXLYMxMv_c3abXFvAIBS0tWHxLL0sDjuuBrPjbrTP7lJH-NQw/pub?output=xlsx"
st.success(f"✔️ ไฟล์ Drug Master จะถูกดึงข้อมูลจากระบบโดยอัตโนมัติ")


if st.button("🚀 เริ่มการวิเคราะห์ ABC", key="abc_button", use_container_width=True):
    if inventory_files:
        progress_bar = st.progress(0, text="กำลังเริ่มต้นการวิเคราะห์ ABC...")
        with st.spinner("กำลังทำการวิเคราะห์ ABC... กระบวนการนี้อาจใช้เวลาสักครู่"):
            # Pass the URL string directly to the function
            report_bytes = process_abc_analysis(inventory_files, master_file_url, progress_bar)

        if report_bytes:
            progress_bar.progress(100, text="[100%] การวิเคราะห์ ABC เสร็จสมบูรณ์")

            # แสดงปุ่มดาวน์โหลด
            st.download_button(
                label="📥 ดาวน์โหลดรายงาน ABC Analysis",
                data=report_bytes,
                file_name="Consolidated_ABC_Report.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

            # แสดงความสำเร็จหลังปุ่มดาวน์โหลด
            st.success("✅ การวิเคราะห์ ABC เสร็จสมบูรณ์และสร้างรายงานสำเร็จ! (ไฟล์พร้อมดาวน์โหลด)")
            progress_bar.empty()
    else:
        st.warning("⚠️ กรุณาอัปโหลดไฟล์ข้อมูลการใช้งาน")